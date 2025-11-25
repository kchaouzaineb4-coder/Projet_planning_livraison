import streamlit as st
import pandas as pd
from backend import DeliveryProcessor, TruckRentalProcessor, TruckTransferManager, SEUIL_POIDS, SEUIL_VOLUME 
import plotly.express as px


# =====================================================
# === Fonction show_df pour arrondir à 3 décimales ===
# =====================================================
def show_df(df, **kwargs):
    """
    Affiche un DataFrame avec tous les nombres arrondis à 3 décimales.
    kwargs sont transmis à st.dataframe.
    """
    if isinstance(df, pd.DataFrame):
        df_to_display = df.copy()
        df_to_display = df_to_display.round(3)
        st.dataframe(df_to_display, **kwargs)
    else:
        st.dataframe(df, **kwargs)

# =====================================================
# === Fonction show_df_multiline avec affichage HTML ===
# =====================================================
def show_df_multiline(df, column_to_multiline):
    """
    Affiche un DataFrame avec les articles multilignes dans la même cellule.
    Chaque 'No livraison' reste sur une seule ligne.
    """
    df_display = df.copy()

    # Grouper les lignes par livraison et concaténer les articles avec des <br>
    df_display = df_display.groupby(
        ['No livraison', 'Client', 'Ville', 'Représentant', 'Poids total', 'Volume total'],
        as_index=False
    ).agg({column_to_multiline: lambda x: "<br>".join(x.astype(str))})

    # CSS pour forcer l'affichage des <br> sur plusieurs lignes
    css = """
    <style>
    table {
        width: 100%;
        border-collapse: collapse;
    }
    th, td {
        border: 1px solid #555;
        padding: 8px;
        text-align: left;
        vertical-align: top;
        white-space: normal;
        word-wrap: break-word;
    }
    th {
        background-color: #EFF6FF;
        color: white;
    }
    td {
        color: #ddd;
    }
    </style>
    """

    html = df_display.to_html(escape=False, index=False)
    st.markdown(css + html, unsafe_allow_html=True)

# =====================================================
# 📌 Constantes pour les véhicules et chauffeurs
# =====================================================
VEHICULES_DISPONIBLES = [
    'SLG-VEH11', 'SLG-VEH14', 'SLG-VEH22', 'SLG-VEH19',
    'SLG-VEH10', 'SLG-VEH16', 'SLG-VEH23', 'SLG-VEH08', 'SLG-VEH20', 'code-Camion'
]

CHAUFFEURS_DETAILS = {
    '09254': 'DAMMAK Karim', '06002': 'MAAZOUN Bassem', '11063': 'SASSI Ramzi',
    '10334': 'BOUJELBENE Mohamed', '15144': 'GADDOUR Rami', '08278': 'DAMMAK Wissem',
    '18339': 'REKIK Ahmed', '07250': 'BARKIA Mustapha', '13321': 'BADRI Moez','99999': 'Chauffeur Camion'
}

# Configuration page
st.set_page_config(page_title="Planning Livraisons", layout="wide")

# CSS pour centrer le titre
st.markdown("""
<style>
/* Centrer le titre principal */
h1 {
    text-align: center !important;
    color: #1E3A8A;
    margin-bottom: 1rem;
}
</style>
""", unsafe_allow_html=True)

st.title("🚚 Planning de Livraisons & Optimisation des Tournées")
#st.markdown("---")
# =====================================================
# CSS PERSONNALISÉ POUR LA SECTION 1
# =====================================================
st.markdown("""
<style>

/* Style pour le header de la section 1 */
section[data-testid="stVerticalBlock"] > div:has(h1:contains("1. 📥 Importation des Données")) {
    background: linear-gradient(135deg, #1E3A8A 0%, #3B82F6 100%);
    padding: 1.5rem;
    border-radius: 10px;
    margin-bottom: 1rem;
    color: white;
}

/* Style pour le titre de la section 1 */
h1:contains("1. 📥 Importation des Données") {
    color: white !important;
    margin-bottom: 0 !important;
}

/* Style pour les colonnes de fichiers */
div[data-testid="stHorizontalBlock"] {
    background-color: #F8FAFC;
    padding: 1rem;
    border-radius: 8px;
    border: 2px solid #E2E8F0;
}

/* Style pour les file uploaders */
div[data-testid="stFileUploader"] {
    background-color: white;
    padding: 1rem;
    border-radius: 8px;
    border: 4px dashed #CBD5E1;
}

/* Style pour le bouton principal */
div[data-testid="stHorizontalBlock"] button[kind="primary"] {
    background: linear-gradient(135deg, #1E3A8A 0%, #2563EB 100%);
    border: none;
    color: white;
    font-weight: bold;
}

/* Style pour les labels des file uploaders */
div[data-testid="stFileUploader"] label {
    font-weight: bold;
    color: #1E40AF;
}

/* Style pour les fichiers uploadés */
div[data-testid="stFileUploader"] div[data-testid="stMarkdownContainer"] {
    background-color: #DBEAFE;
    padding: 0.5rem;
    border-radius: 5px;
    border-left: 4px solid #3B82F6;
}
</style>
""", unsafe_allow_html=True)
# =====================================================
# INITIALISATION DE L'ÉTAT DE SESSION
# =====================================================
if 'data_processed' not in st.session_state:
    st.session_state.data_processed = False
    st.session_state.df_grouped = None
    st.session_state.df_city = None
    st.session_state.df_grouped_zone = None
    st.session_state.df_zone = None 
    st.session_state.df_optimized_estafettes = None
    st.session_state.df_livraisons_original = None
    st.session_state.rental_processor = None
    st.session_state.propositions = None
    st.session_state.selected_client = None
    st.session_state.message = ""
    st.session_state.df_voyages = None
    st.session_state.df_livraisons = None

# =====================================================
# Fonctions de Callback pour la Location
# =====================================================

def update_propositions_view():
    """Met à jour le DataFrame de propositions après une action."""
    if st.session_state.rental_processor:
        st.session_state.propositions = st.session_state.rental_processor.detecter_propositions()
        
        # CORRECTION : Vérifier si le DataFrame de propositions n'est pas vide et contient la colonne 'Client'
        if (st.session_state.propositions is not None and 
            not st.session_state.propositions.empty and 
            'Client' in st.session_state.propositions.columns):
            
            # Réinitialiser la sélection si le client n'est plus dans les propositions ouvertes
            if (st.session_state.selected_client is not None and 
                st.session_state.selected_client not in st.session_state.propositions['Client'].astype(str).tolist()):
                st.session_state.selected_client = None
    else:
        st.session_state.propositions = pd.DataFrame()

def handle_location_action(accepter):
    """Gère l'acceptation ou le refus de la proposition de location."""
    if st.session_state.rental_processor and st.session_state.selected_client:
        try:
            # Assurer que le client est une chaîne valide
            client_to_process = str(st.session_state.selected_client)
            ok, msg, _ = st.session_state.rental_processor.appliquer_location(
                client_to_process, accepter=accepter
            )
            st.session_state.message = msg
            update_propositions_view()
        except Exception as e:
            st.session_state.message = f"❌ Erreur lors du traitement : {str(e)}"
    elif not st.session_state.selected_client:
        st.session_state.message = "⚠️ Veuillez sélectionner un client à traiter."
    else:
        st.session_state.message = "⚠️ Le processeur de location n'est pas initialisé."

def accept_location_callback():
    handle_location_action(True)

def refuse_location_callback():
    handle_location_action(False)

# =====================================================
# 1. UPLOAD DES FICHIERS INPUT (Section 1)
# =====================================================
st.header("1. 📥 Importation des Données")

col_file_1, col_file_2, col_file_3, col_button = st.columns([1, 1, 1, 1])
with col_file_1:
    liv_file = st.file_uploader("Fichier Livraisons (BL)", type=["xlsx"])
with col_file_2:
    ydlogist_file = st.file_uploader("Fichier Volumes (Articles)", type=["xlsx"])
with col_file_3:
    wcliegps_file = st.file_uploader("Fichier Clients/Zones", type=["xlsx"])
with col_button:
    st.markdown("<br>", unsafe_allow_html=True)
    if st.button("Exécuter le traitement complet", type="primary"):
        if liv_file and ydlogist_file and wcliegps_file:
            processor = DeliveryProcessor()
            try:
                with st.spinner("Traitement des données en cours..."):
                    # Récupération des 6 valeurs
                    df_grouped, df_city, df_grouped_zone, df_zone, df_optimized_estafettes, df_livraisons_original = processor.process_delivery_data(liv_file, ydlogist_file, wcliegps_file)
                
                # Stockage des résultats dans l'état de session
                st.session_state.df_optimized_estafettes = df_optimized_estafettes
                st.session_state.df_grouped = df_grouped
                st.session_state.df_city = df_city
                st.session_state.df_grouped_zone = df_grouped_zone
                st.session_state.df_zone = df_zone 
                st.session_state.df_livraisons_original = df_livraisons_original
                st.session_state.df_livraisons = df_grouped_zone  # Pour la section transfert
                
                # Initialisation avec les données originales
                st.session_state.rental_processor = TruckRentalProcessor(df_optimized_estafettes, df_livraisons_original)
                update_propositions_view()
                
                st.session_state.data_processed = True
                #st.session_state.message = "Traitement terminé avec succès ! Les résultats s'affichent ci-dessous."
                st.rerun()

            except Exception as e:
                st.error(f"❌ Erreur lors du traitement : {str(e)}")
                st.session_state.data_processed = False
        else:
            st.warning("Veuillez uploader tous les fichiers nécessaires.")
st.markdown("---")

# =====================================================
# AFFICHAGE DES RÉSULTATS (Se déclenche si les données sont traitées)
# =====================================================
if st.session_state.data_processed:
    
    # Affichage des messages d'opération
    if st.session_state.message.startswith("✅"):
        st.success(st.session_state.message)
    elif st.session_state.message.startswith("❌"):
        st.error(st.session_state.message)
    elif st.session_state.message.startswith("⚠️"):
        st.warning(st.session_state.message)
    #else:
        #st.info(st.session_state.message or "Prêt à traiter les propositions de location.")
    
    # Récupération du DF mis à jour à chaque fois
    df_optimized_estafettes = st.session_state.rental_processor.get_df_result() 
    
# =====================================================
# 2. ANALYSE DE LIVRAISON DÉTAILLÉE (Section 2)
# =====================================================
st.header("2. 🔍 Analyse de Livraison Détaillée")

tab_grouped, tab_city, tab_zone_group, tab_zone_summary, tab_charts = st.tabs([
    "Livraisons Client/Ville", 
    "Besoin Estafette par Ville", 
    "Livraisons Client/Zone", 
    "Besoin Estafette par Zone",
    "Graphiques"
])
# --- CSS PERSONNALISÉ POUR LES ONGLETS ---
st.markdown("""
<style>
    /* Style pour les onglets - COULEUR BLEUE */
    .stTabs [data-baseweb="tab-list"] {
        gap: 8px;
    }
    
    .stTabs [data-baseweb="tab"] {
        height: 50px;
        white-space: pre-wrap;
        background-color: #F0F2F6;
        border-radius: 8px 8px 0px 0px;
        gap: 8px;
        padding: 10px 16px;
        font-weight: 600;
    }
    
    .stTabs [data-baseweb="tab"]:hover {
        background-color: #E6F3FF;
        color: #0369A1;
    }
    
    /* ONGLET ACTIF - BLEU ROYAL */
    .stTabs [aria-selected="true"] {
        background-color: #0369A1 !important;
        color: white !important;
    }
    
    /* TEXTE DES ONGLETS */
    .stTabs [data-baseweb="tab"] p {
        font-size: 16px;
        font-weight: 600;
        margin: 0;
    }
    
    /* COULEUR DU TEXTE POUR ONGLET ACTIF */
    .stTabs [aria-selected="true"] p {
        color: white !important;
    }
</style>
""", unsafe_allow_html=True)


# --- Onglet Livraisons Client/Ville ---
with tab_grouped:
    st.subheader("Livraisons par Client & Ville")
    
    # Créer une copie du DataFrame et FILTRER TRIPOLI
    df_liv = st.session_state.df_grouped.drop(columns=["Zone"], errors='ignore').copy()
    df_liv = df_liv[df_liv["Ville"] != "TRIPOLI"]  # ← FILTRE TRIPOLI ICI
    
    # CSS pour un tableau organisé et professionnel
    st.markdown("""
    <style>
    /* Style général du tableau */
    .custom-table {
        width: 100%;
        border-collapse: collapse;
        font-family: Arial, sans-serif;
        font-size: 14px;
        box-shadow: 0 2px 8px rgba(0,0,0,0.1);
        border-radius: 8px;
        overflow: hidden;
    }
    
    /* En-têtes du tableau - BLEU ROYAL SANS DÉGRADÉ */
    .custom-table th {
        background-color: #0369A1;
        color: white;
        padding: 12px 8px;
        text-align: center;
        border: 2px solid #4682B4;
        font-weight: bold;
        font-size: 13px;
        vertical-align: middle;  /* CENTRAGE VERTICAL */
    }
    
    /* Cellules du tableau - TOUTES EN BLANC */
    .custom-table td {
        padding: 10px 8px;
        text-align: center;
        border: 1px solid #B0C4DE;
        background-color: white;
        color: #000000;
        vertical-align: middle;  /* CENTRAGE VERTICAL */
    }
    
    /* Bordures visibles pour toutes les cellules */
    .custom-table th, 
    .custom-table td {
        border: 1px solid #B0C4DE !important;
    }
    
    /* Bordures épaisses pour l'extérieur du tableau */
    .custom-table {
        border: 2px solid #4682B4 !important;
    }
    
    /* Style spécifique pour la colonne Article - CENTRÉ */
    .custom-table td:nth-child(5) {
        text-align: center;
        max-width: 200px;
        word-wrap: break-word;
        white-space: normal;
        vertical-align: middle;  /* CENTRAGE VERTICAL */
    }
    
    /* Style pour les cellules de poids et volume - NOIR */
    .custom-table td:nth-child(6),
    .custom-table td:nth-child(7) {
        
        color: #000000 !important;
        vertical-align: middle;  /* CENTRAGE VERTICAL */
    }
    
    /* Conteneur du tableau avec défilement horizontal */
    .table-container {
        overflow-x: auto;
        margin: 1rem 0;
        border-radius: 8px;
        border: 2px solid #4682B4;
    }
    
    /* Supprimer l'alternance des couleurs - TOUTES LES LIGNES BLANCHES */
    .custom-table tr:nth-child(even) td {
        background-color: white !important;
    }
    
    /* Survol des lignes - léger effet */
    .custom-table tr:hover td {
        background-color: #F0F8FF !important;
    }
    </style>
    """, unsafe_allow_html=True)
    
    # Vérifier si le DataFrame n'est pas vide après filtrage
    if df_liv.empty:
        st.info("ℹ️ Aucune livraison à afficher (TRIPOLI exclue)")
    else:
        # Préparer les données pour l'affichage HTML
        if "Article" in df_liv.columns:
            # Transformer les articles avec retours à la ligne HTML - SANS "•"
            df_liv["Article"] = df_liv["Article"].astype(str).apply(
                lambda x: "<br>".join(a.strip() for a in x.split(",") if a.strip())
            )
        
        # Formater les nombres - 3 chiffres après la virgule
        if "Poids total" in df_liv.columns:
            df_liv["Poids total"] = df_liv["Poids total"].map(lambda x: f"{x:.3f} kg" if pd.notna(x) else "")
        if "Volume total" in df_liv.columns:
            df_liv["Volume total"] = df_liv["Volume total"].map(lambda x: f"{x:.3f} m³" if pd.notna(x) else "")
        
        # Afficher le tableau avec le style CSS
        html_table = df_liv.to_html(
            escape=False, 
            index=False, 
            classes="custom-table",
            border=0
        )
        
        st.markdown(f"""
        <div class="table-container">
            {html_table}
        </div>
        """, unsafe_allow_html=True)
    
    # Métriques résumées - CORRECTION : Utiliser les données filtrées
    st.markdown("---")
    col1, col2, col3, col4 = st.columns(4)
    
    with col1:
        total_livraisons = len(df_liv)
        st.metric("📦 Total Livraisons", total_livraisons)
    
    with col2:
        total_clients = df_liv["Client"].nunique()
        st.metric("👥 Clients Uniques", total_clients)
    
    with col3:
        # CORRECTION : Calculer le poids total à partir des données filtrées (données originales)
        df_liv_original = st.session_state.df_grouped[st.session_state.df_grouped["Ville"] != "TRIPOLI"]
        total_poids = df_liv_original["Poids total"].sum()
        st.metric("⚖️ Poids Total", f"{total_poids:.3f} kg")
    
    with col4:
        # CORRECTION : Calculer le volume total à partir des données filtrées (données originales)
        total_volume = df_liv_original["Volume total"].sum()
        st.metric("📏 Volume Total", f"{total_volume:.3f} m³")
    
    # Information sur le filtrage
    #st.info("ℹ️ Les livraisons de TRIPOLI ont été exclues de ce tableau")
    
    # AJOUT DE L'IMPORT MANQUANT POUR BytesIO
    from io import BytesIO
    
    # Bouton de téléchargement (garder les données originales pour l'export)
    excel_buffer_grouped = BytesIO()
    with pd.ExcelWriter(excel_buffer_grouped, engine='openpyxl') as writer:
        st.session_state.df_grouped.drop(columns=["Zone"], errors='ignore').to_excel(writer, index=False, sheet_name="Livraisons Client Ville")
    excel_buffer_grouped.seek(0)
    
    st.download_button(
        label="💾 Télécharger Livraisons Client/Ville",
        data=excel_buffer_grouped,
        file_name="Livraisons_Client_Ville.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    
    # Stockage pour la section 5
    if "df_livraisons" not in st.session_state:
        st.session_state.df_livraisons = df_liv.copy()
# --- Onglet Besoin Estafette par Ville ---
with tab_city:
    st.subheader("Besoin Estafette par Ville")
    
    # Créer une copie du DataFrame et FILTRER TRIPOLI
    df_city_display = st.session_state.df_city.copy()
    df_city_display = df_city_display[df_city_display["Ville"] != "TRIPOLI"]  # ← FILTRE TRIPOLI ICI
    
    # Formater les nombres - 3 chiffres après la virgule
    if "Poids total" in df_city_display.columns:
        df_city_display["Poids total"] = df_city_display["Poids total"].map(lambda x: f"{x:.3f} kg" if pd.notna(x) else "")
    if "Volume total" in df_city_display.columns:
        df_city_display["Volume total"] = df_city_display["Volume total"].map(lambda x: f"{x:.3f} m³" if pd.notna(x) else "")
    if "Besoin estafette réel" in df_city_display.columns:
        df_city_display["Besoin estafette réel"] = df_city_display["Besoin estafette réel"].map(lambda x: f"{x:.1f}" if pd.notna(x) else "")
    
    # Vérifier si le DataFrame n'est pas vide après filtrage
    if df_city_display.empty:
        st.info("ℹ️ Aucune ville à afficher (TRIPOLI exclue)")
    else:
        # Afficher le tableau avec le style CSS
        html_table_city = df_city_display.to_html(
            escape=False, 
            index=False, 
            classes="custom-table",
            border=0
        )
        
        st.markdown(f"""
        <div class="table-container">
            {html_table_city}
        </div>
        """, unsafe_allow_html=True)
    
    # Métriques résumées - CORRECTION : Utiliser les données filtrées
    st.markdown("---")
    col1, col2, col3 = st.columns(3)
    
    with col1:
        total_villes = len(df_city_display)
        st.metric("🏙️ Total Villes", total_villes)
    
    with col2:
        # CORRECTION : Utiliser les données filtrées pour les calculs
        df_city_original_filtered = st.session_state.df_city[st.session_state.df_city["Ville"] != "TRIPOLI"]
        total_bls = df_city_original_filtered["Nombre de BLs"].sum() if "Nombre de BLs" in df_city_original_filtered.columns else 0
        st.metric("📦 Total BLs", int(total_bls))
    
    with col3:
        # CORRECTION : Utiliser les données filtrées pour les calculs
        total_estafettes = df_city_original_filtered["Besoin estafette réel"].sum() if "Besoin estafette réel" in df_city_original_filtered.columns else 0
        st.metric("🚐 Besoin Estafettes", f"{total_estafettes:.1f}")

    
    # Information sur le filtrage
    #st.info("ℹ️ La ville de TRIPOLI a été exclue de ce tableau")
    
    # Bouton de téléchargement (garder les données originales pour l'export)
    excel_buffer_city = BytesIO()
    with pd.ExcelWriter(excel_buffer_city, engine='openpyxl') as writer:
        st.session_state.df_city.to_excel(writer, index=False, sheet_name="Besoin Estafette Ville")
    excel_buffer_city.seek(0)
    
    st.download_button(
        label="💾 Télécharger Besoin par Ville",
        data=excel_buffer_city,
        file_name="Besoin_Estafette_Ville.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
# --- Onglet Livraisons Client & Ville + Zone ---
with tab_zone_group:
    st.subheader("Livraisons par Client & Ville + Zone")

    # Créer une copie du DataFrame
    df_liv_zone = st.session_state.df_grouped_zone.copy()
    
    # CSS pour un tableau organisé et professionnel
    st.markdown("""
    <style>
    /* Style général du tableau */
    .custom-table {
        width: 100%;
        border-collapse: collapse;
        font-family: Arial, sans-serif;
        font-size: 14px;
        box-shadow: 0 2px 8px rgba(0,0,0,0.1);
        border-radius: 8px;
        overflow: hidden;
    }
    
    /* En-têtes du tableau - BLEU ROYAL SANS DÉGRADÉ */
    .custom-table th {
        background-color: #0369A1;
        color: white;
        padding: 12px 8px;
        text-align: center;
        border: 2px solid #4682B4;
        font-weight: bold;
        font-size: 13px;
        vertical-align: middle;  /* CENTRAGE VERTICAL */
    }
    
    /* Cellules du tableau - TOUTES EN BLANC */
    .custom-table td {
        padding: 10px 8px;
        text-align: center;
        border: 1px solid #B0C4DE;
        background-color: white;
        color: #000000;
        vertical-align: middle;  /* CENTRAGE VERTICAL */
    }
    
    /* Bordures visibles pour toutes les cellules */
    .custom-table th, 
    .custom-table td {
        border: 1px solid #B0C4DE !important;
    }
    
    /* Bordures épaisses pour l'extérieur du tableau */
    .custom-table {
        border: 2px solid #4682B4 !important;
    }
    
    /* Style spécifique pour la colonne Article - CENTRÉ */
    .custom-table td:nth-child(5) {
        text-align: center;
        max-width: 200px;
        word-wrap: break-word;
        white-space: normal;
        vertical-align: middle;  /* CENTRAGE VERTICAL */
    }
    
    /* Style pour les cellules de poids et volume - NOIR */
    .custom-table td:nth-child(6),
    .custom-table td:nth-child(7) {
        font-weight: 600;
        color: #000000 !important;
        vertical-align: middle;  /* CENTRAGE VERTICAL */
    }
    
    /* Conteneur du tableau avec défilement horizontal */
    .table-container {
        overflow-x: auto;
        margin: 1rem 0;
        border-radius: 8px;
        border: 2px solid #4682B4;
    }
    
    /* Supprimer l'alternance des couleurs - TOUTES LES LIGNES BLANCHES */
    .custom-table tr:nth-child(even) td {
        background-color: white !important;
    }
    
    /* Survol des lignes - léger effet */
    .custom-table tr:hover td {
        background-color: #F0F8FF !important;
    }
    </style>
    """, unsafe_allow_html=True)
    
    # Préparer les données pour l'affichage HTML
    if "Article" in df_liv_zone.columns:
        # Transformer les articles avec retours à la ligne HTML - SANS "•"
        df_liv_zone["Article"] = df_liv_zone["Article"].astype(str).apply(
            lambda x: "<br>".join(a.strip() for a in x.split(",") if a.strip())
        )
    
    # Formater les nombres - 3 chiffres après la virgule
    if "Poids total" in df_liv_zone.columns:
        df_liv_zone["Poids total"] = df_liv_zone["Poids total"].map(lambda x: f"{x:.3f} kg" if pd.notna(x) else "")
    if "Volume total" in df_liv_zone.columns:
        df_liv_zone["Volume total"] = df_liv_zone["Volume total"].map(lambda x: f"{x:.3f} m³" if pd.notna(x) else "")
    
    # Afficher le tableau avec le style CSS
    html_table_zone_group = df_liv_zone.to_html(
        escape=False, 
        index=False, 
        classes="custom-table",
        border=0
    )
    
    st.markdown(f"""
    <div class="table-container">
        {html_table_zone_group}
    </div>
    """, unsafe_allow_html=True)
    
    # Métriques résumées
    st.markdown("---")
    col1, col2, col3 = st.columns(3)
    
    with col1:
        total_livraisons_zone = len(df_liv_zone)
        st.metric("📦 Total Livraisons", total_livraisons_zone)
    
    with col2:
        zones_count = df_liv_zone["Zone"].nunique()
        st.metric("🌍 Zones", zones_count)
    
    with col3:
        villes_count = df_liv_zone["Ville"].nunique()
        st.metric("🏙️ Villes", villes_count)
    
   
    
    # Bouton de téléchargement
    excel_buffer_zone_group = BytesIO()
    with pd.ExcelWriter(excel_buffer_zone_group, engine='openpyxl') as writer:
        st.session_state.df_grouped_zone.to_excel(writer, index=False, sheet_name="Livraisons Client Ville Zone")
    excel_buffer_zone_group.seek(0)
    
    st.download_button(
        label="💾 Télécharger Livraisons Client/Ville/Zone",
        data=excel_buffer_zone_group,
        file_name="Livraisons_Client_Ville_Zone.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
# --- Onglet Besoin Estafette par Zone ---
with tab_zone_summary:
    st.subheader("Besoin Estafette par Zone")
    
    # Créer une copie du DataFrame et renommer la colonne
    df_zone_display = st.session_state.df_zone.copy()
    
    # RENOMMER LA COLONNE "Nombre livraisons" en "Nombre de BLs"
    df_zone_display = df_zone_display.rename(columns={"Nombre livraisons": "Nombre de BLs"})
    
    # CSS pour un tableau organisé et professionnel
    st.markdown("""
    <style>
    /* Style général du tableau */
    .custom-table {
        width: 100%;
        border-collapse: collapse;
        font-family: Arial, sans-serif;
        font-size: 14px;
        box-shadow: 0 2px 8px rgba(0,0,0,0.1);
        border-radius: 8px;
        overflow: hidden;
    }
    
    /* En-têtes du tableau - BLEU ROYAL SANS DÉGRADÉ */
    .custom-table th {
        background-color: #0369A1;
        color: white;
        padding: 12px 8px;
        text-align: center;
        border: 2px solid #4682B4;
        font-weight: bold;
        font-size: 13px;
        vertical-align: middle;  /* CENTRAGE VERTICAL */
    }
    
    /* Cellules du tableau - TOUTES EN BLANC */
    .custom-table td {
        padding: 10px 8px;
        text-align: center;
        border: 1px solid #B0C4DE;
        background-color: white;
        color: #000000;
        vertical-align: middle;  /* CENTRAGE VERTICAL */
    }
    
    /* Bordures visibles pour toutes les cellules */
    .custom-table th, 
    .custom-table td {
        border: 1px solid #B0C4DE !important;
    }
    
    /* Bordures épaisses pour l'extérieur du tableau */
    .custom-table {
        border: 2px solid #4682B4 !important;
    }
    
    /* Style pour les cellules numériques */
    .custom-table td:nth-child(2),
    .custom-table td:nth-child(3),
    .custom-table td:nth-child(4),
    .custom-table td:nth-child(5) {
        font-weight: 600;
        color: #000000 !important;
        vertical-align: middle;  /* CENTRAGE VERTICAL */
    }
    
    /* Conteneur du tableau avec défilement horizontal */
    .table-container {
        overflow-x: auto;
        margin: 1rem 0;
        border-radius: 8px;
        border: 2px solid #4682B4;
    }
    
    /* Supprimer l'alternance des couleurs - TOUTES LES LIGNES BLANCHES */
    .custom-table tr:nth-child(even) td {
        background-color: white !important;
    }
    
    /* Survol des lignes - léger effet */
    .custom-table tr:hover td {
        background-color: #F0F8FF !important;
    }
    </style>
    """, unsafe_allow_html=True)
    
    # Formater les nombres - 3 chiffres après la virgule
    if "Poids total" in df_zone_display.columns:
        df_zone_display["Poids total"] = df_zone_display["Poids total"].map(lambda x: f"{x:.3f} kg" if pd.notna(x) else "")
    if "Volume total" in df_zone_display.columns:
        df_zone_display["Volume total"] = df_zone_display["Volume total"].map(lambda x: f"{x:.3f} m³" if pd.notna(x) else "")
    if "Besoin estafette réel" in df_zone_display.columns:
        df_zone_display["Besoin estafette réel"] = df_zone_display["Besoin estafette réel"].map(lambda x: f"{x:.1f}" if pd.notna(x) else "")
    # MAINTENANT ON UTILISE "Nombre de BLs" AU LIEU DE "Nombre livraisons"
    if "Nombre de BLs" in df_zone_display.columns:
        df_zone_display["Nombre de BLs"] = df_zone_display["Nombre de BLs"].map(lambda x: f"{int(x)}" if pd.notna(x) else "")
    
    # Afficher le tableau avec le style CSS
    html_table_zone = df_zone_display.to_html(
        escape=False, 
        index=False, 
        classes="custom-table",
        border=0
    )
    
    st.markdown(f"""
    <div class="table-container">
        {html_table_zone}
    </div>
    """, unsafe_allow_html=True)
    
    # Métriques résumées - CORRECTION : Utiliser les données originales pour les calculs
    st.markdown("---")
    col1, col2, col3 = st.columns(3)
    
    with col1:
        total_zones = len(df_zone_display)
        st.metric("🌍 Total Zones", total_zones)
    
    with col2:
        # Utiliser les données originales pour les calculs (avec l'ancien nom de colonne)
        if "Nombre livraisons" in st.session_state.df_zone.columns:
            total_bls_zone = st.session_state.df_zone["Nombre livraisons"].sum()
        else:
            total_bls_zone = 0
        st.metric("📦 Total BLs", int(total_bls_zone))
    
    with col3:
        # Utiliser les données originales pour les calculs
        total_estafettes_zone = st.session_state.df_zone["Besoin estafette réel"].sum() if "Besoin estafette réel" in st.session_state.df_zone.columns else 0
        st.metric("🚐 Besoin Estafettes", f"{total_estafettes_zone:.1f}")
    
    # Bouton de téléchargement
    excel_buffer_zone = BytesIO()
    with pd.ExcelWriter(excel_buffer_zone, engine='openpyxl') as writer:
        # Pour l'export Excel, on utilise les données originales
        st.session_state.df_zone.to_excel(writer, index=False, sheet_name="Besoin Estafette Zone")
    excel_buffer_zone.seek(0)
    
    st.download_button(
        label="💾 Télécharger Besoin par Zone",
        data=excel_buffer_zone,
        file_name="Besoin_Estafette_Zone.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

# --- Onglet Graphiques ---
with tab_charts:
    st.subheader("Statistiques par Ville")
    
    # FILTRER LES DONNÉES POUR EXCLURE TRIPOLI
    df_filtered = st.session_state.df_city[st.session_state.df_city["Ville"] != "TRIPOLI"]
    
    # Configuration commune pour tous les graphiques
    chart_config = {
        'color_discrete_sequence': ['#0369A1'],  # BLEU ROYAL
        'template': 'plotly_white',
    }
    
    col1, col2 = st.columns(2)
    with col1:
        fig1 = px.bar(df_filtered, x="Ville", y="Poids total", **chart_config)
        fig1.update_layout(title_text="Poids total livré par ville", title_x=0.5)
        st.plotly_chart(fig1, use_container_width=True)
        
    with col2:
        fig2 = px.bar(df_filtered, x="Ville", y="Volume total", **chart_config)
        fig2.update_layout(title_text="Volume total livré par ville (m³)", title_x=0.5)
        st.plotly_chart(fig2, use_container_width=True)

    col3, col4 = st.columns(2)
    with col3:
        # DIAGRAMME CORRIGÉ : Nombre de BL par ville
        df_chart = df_filtered.rename(columns={"Nombre livraisons": "Nombre de BLs"})
        fig3 = px.bar(df_chart, x="Ville", y="Nombre de BLs", **chart_config)
        fig3.update_layout(title_text="Nombre de BL par ville", title_x=0.5)
        st.plotly_chart(fig3, use_container_width=True)
        
    with col4:
        fig4 = px.bar(df_filtered, x="Ville", y="Besoin estafette réel", **chart_config)
        fig4.update_layout(title_text="Besoin en Estafettes par ville", title_x=0.5)
        st.plotly_chart(fig4, use_container_width=True)

st.markdown("---")
# =====================================================
# 3. PROPOSITION DE LOCATION DE CAMION (Section 3)
# =====================================================
st.header("3. 🚚 Proposition de location de camion")
st.markdown(f"🔸 Si un client dépasse **{SEUIL_POIDS} kg** ou **{SEUIL_VOLUME} m³**, une location est proposée (si non déjà décidée).")

# CSS POUR LES TABLEAUX DE LA SECTION 3
st.markdown("""
<style>
    /* Style général du tableau */
    .custom-table-rental {
        width: 100%;
        border-collapse: collapse;
        font-family: Arial, sans-serif;
        font-size: 14px;
        box-shadow: 0 2px 8px rgba(0,0,0,0.1);
        border-radius: 8px;
        overflow: hidden;
    }
    
    /* En-têtes du tableau - BLEU ROYAL SANS DÉGRADÉ */
    .custom-table-rental th {
        background-color: #0369A1;
        color: white;
        padding: 12px 8px;
        text-align: center;
        border: 2px solid #4682B4;
        font-weight: bold;
        font-size: 13px;
        vertical-align: middle;
    }
    
    /* Cellules du tableau - TOUTES EN BLANC */
    .custom-table-rental td {
        padding: 10px 8px;
        text-align: center;
        border: 1px solid #B0C4DE;
        background-color: white;
        color: #000000;
        vertical-align: middle;
    }
    
    /* Bordures visibles pour toutes les cellules */
    .custom-table-rental th, 
    .custom-table-rental td {
        border: 1px solid #B0C4DE !important;
    }
    
    /* Bordures épaisses pour l'extérieur du tableau */
    .custom-table-rental {
        border: 2px solid #4682B4 !important;
    }
    
    /* Style pour les cellules numériques */
    .custom-table-rental td:nth-child(2),
    .custom-table-rental td:nth-child(3),
    .custom-table-rental td:nth-child(4),
    .custom-table-rental td:nth-child(5),
    .custom-table-rental td:nth-child(6) {
        font-weight: 600;
        color: #000000 !important;
        vertical-align: middle;
    }
    
    /* Conteneur du tableau avec défilement horizontal */
    .table-container-rental {
        overflow-x: auto;
        margin: 1rem 0;
        border-radius: 8px;
        border: 2px solid #4682B4;
    }
    
    /* Supprimer l'alternance des couleurs - TOUTES LES LIGNES BLANCHES */
    .custom-table-rental tr:nth-child(even) td {
        background-color: white !important;
    }
    
    /* Survol des lignes - léger effet */
    .custom-table-rental tr:hover td {
        background-color: #F0F8FF !important;
    }
    
    /* Style spécifique pour les cellules multilignes (BL inclus) */
    .multiline-cell {
        line-height: 1.4;
        text-align: left !important;
        padding: 8px !important;
    }
</style>
""", unsafe_allow_html=True)

if st.session_state.propositions is not None and not st.session_state.propositions.empty:
    col_prop, col_details = st.columns([2, 3])
    
    with col_prop:
        st.markdown("### Propositions ouvertes")
        
        # CORRECTION : Vérifier si la colonne 'Client' existe
        if 'Client' in st.session_state.propositions.columns:
            # FORMATAGE DU TABLEAU DES PROPOSITIONS AVEC STYLE CSS
            propositions_display = st.session_state.propositions.copy()
            
            # Formater les nombres
            if "Poids total (kg)" in propositions_display.columns:
                propositions_display["Poids total (kg)"] = propositions_display["Poids total (kg)"].map(
                    lambda x: f"{float(x):.3f}" if pd.notna(x) else ""
                )
            if "Volume total (m³)" in propositions_display.columns:
                propositions_display["Volume total (m³)"] = propositions_display["Volume total (m³)"].map(
                    lambda x: f"{float(x):.3f}" if pd.notna(x) else ""
                )
            
            # Afficher le tableau avec le style CSS
            html_table_propositions = propositions_display.to_html(
                escape=False, 
                index=False, 
                classes="custom-table-rental",
                border=0
            )
            
            st.markdown(f"""
            <div class="table-container-rental">
                {html_table_propositions}
            </div>
            """, unsafe_allow_html=True)
            
           # MÉTRIQUES RÉSUMÉES
            st.markdown("---")
            col_metric1, col_metric2, col_metric3 = st.columns(3)

            with col_metric1:
                total_propositions = len(st.session_state.propositions)
                st.metric("📋 Propositions ouvertes", total_propositions)

            with col_metric2:
                # Calculer le nombre de clients dépassant le seuil de POIDS
                clients_poids = len(st.session_state.propositions[
                    st.session_state.propositions["Poids total (kg)"] >= SEUIL_POIDS
                ]) if "Poids total (kg)" in st.session_state.propositions.columns else 0
                st.metric("⚖️ Dépassement poids", clients_poids)

            with col_metric3:
                # CORRECTION : Calculer le nombre de clients dépassant le seuil de VOLUME
                clients_volume = len(st.session_state.propositions[
                    st.session_state.propositions["Volume total (m³)"] >= SEUIL_VOLUME
                ]) if "Volume total (m³)" in st.session_state.propositions.columns else 0
                st.metric("📦 Dépassement volume", clients_volume)

        
            # Sélection du client
            client_options = st.session_state.propositions['Client'].astype(str).tolist()
            client_options_with_empty = [""] + client_options
            
            # Index de sélection par défaut
            default_index = 0
            if st.session_state.selected_client in client_options:
                 default_index = client_options_with_empty.index(st.session_state.selected_client)
            elif len(client_options) > 0:
                 default_index = 1

            st.session_state.selected_client = st.selectbox(
                "Client à traiter :", 
                options=client_options_with_empty, 
                index=default_index,
                key='client_select' 
            )
        else:
            st.warning("⚠️ Format de données incorrect dans les propositions.")
            st.session_state.selected_client = None

        col_btn_acc, col_btn_ref = st.columns(2)
        is_client_selected = st.session_state.selected_client != "" and st.session_state.selected_client is not None
        
        with col_btn_acc:
            st.button(
                "✅ Accepter la location", 
                on_click=accept_location_callback, 
                disabled=not is_client_selected,
                use_container_width=True
            )
        with col_btn_ref:
            st.button(
                "❌ Refuser la proposition", 
                on_click=refuse_location_callback, 
                disabled=not is_client_selected,
                use_container_width=True
            )

    with col_details:
        st.markdown("### Détails de la commande client")
        if is_client_selected:
            try:
                resume, details_df = st.session_state.rental_processor.get_details_client(
                    st.session_state.selected_client
                )
                
                # Afficher le résumé
                st.markdown(f"**{resume}**")
                
                # FORMATAGE DU TABLEAU DES DÉTAILS AVEC STYLE CSS
                if not details_df.empty:
                    details_display = details_df.copy()
                    
                    # CORRECTION : Formatage simple et sécurisé des colonnes
                    def format_numeric_column(series, decimals, unit=""):
                        """Formate une colonne numérique avec le nombre de décimales et unité spécifiés"""
                        formatted_series = series.copy()
                        for i, value in enumerate(series):
                            if pd.notna(value) and value != "":
                                try:
                                    # Essayer de convertir en float
                                    if isinstance(value, str):
                                        # Nettoyer la valeur si c'est une string
                                        clean_value = value.replace(' kg', '').replace(' m³', '').replace('%', '').strip()
                                        num_value = float(clean_value)
                                    else:
                                        num_value = float(value)
                                    
                                    # Formater selon le nombre de décimales
                                    if decimals == 3:
                                        formatted_value = f"{num_value:.3f}"
                                    elif decimals == 2:
                                        formatted_value = f"{num_value:.2f}"
                                    elif decimals == 1:
                                        formatted_value = f"{num_value:.1f}"
                                    else:
                                        formatted_value = f"{num_value:.0f}"
                                    
                                    formatted_series.iloc[i] = f"{formatted_value}{unit}"
                                except (ValueError, TypeError):
                                    # Si conversion échoue, garder la valeur originale
                                    formatted_series.iloc[i] = str(value)
                            else:
                                formatted_series.iloc[i] = ""
                        return formatted_series
                    
                    # Formater les colonnes numériques
                    if "Poids total" in details_display.columns:
                        details_display["Poids total"] = format_numeric_column(details_display["Poids total"], 3, " kg")
                    
                    if "Volume total" in details_display.columns:
                        details_display["Volume total"] = format_numeric_column(details_display["Volume total"], 3, " m³")
                    
                    if "Taux d'occupation (%)" in details_display.columns:
                        details_display["Taux d'occupation (%)"] = format_numeric_column(details_display["Taux d'occupation (%)"], 2, "%")
                    
                    # Gestion spéciale pour "BL inclus" - format multiligne
                    if "BL inclus" in details_display.columns:
                        details_display["BL inclus"] = details_display["BL inclus"].astype(str).apply(
                            lambda x: "<br>".join(bl.strip() for bl in x.split(";")) if ";" in x else x
                        )
                    
                    # Afficher le tableau avec le style CSS
                    html_table_details = details_display.to_html(
                        escape=False, 
                        index=False, 
                        classes="custom-table-rental",
                        border=0
                    )
                    
                    st.markdown(f"""
                    <div class="table-container-rental">
                        {html_table_details}
                    </div>
                    """, unsafe_allow_html=True)
                    
                    # MÉTRIQUES POUR LES DÉTAILS - CORRECTION : Calculs sur données brutes
                    st.markdown("---")
                    col_det1, col_det2, col_det3 = st.columns(3)
                    
                    with col_det1:
                        total_camions = len(details_display)
                        st.metric("🚚 Nombre de camions", total_camions)
                    
                    with col_det2:
                        # Calculer le poids total à partir des données brutes
                        try:
                            if "Poids total" in details_df.columns:
                                poids_total = 0
                                for value in details_df["Poids total"]:
                                    if pd.notna(value):
                                        try:
                                            # Nettoyer la valeur si elle contient des unités
                                            if isinstance(value, str):
                                                clean_value = value.replace(' kg', '').replace('m³', '').strip()
                                            else:
                                                clean_value = str(value)
                                            poids_total += float(clean_value)
                                        except (ValueError, TypeError):
                                            continue
                                st.metric("📦 Poids total", f"{poids_total:.1f} kg")
                            else:
                                st.metric("📦 Poids total", "N/A")
                        except Exception as e:
                            st.metric("📦 Poids total", "Erreur")
                    
                    with col_det3:
                        # Calculer le volume total à partir des données brutes
                        try:
                            if "Volume total" in details_df.columns:
                                volume_total = 0
                                for value in details_df["Volume total"]:
                                    if pd.notna(value):
                                        try:
                                            # Nettoyer la valeur si elle contient des unités
                                            if isinstance(value, str):
                                                clean_value = value.replace(' kg', '').replace('m³', '').strip()
                                            else:
                                                clean_value = str(value)
                                            volume_total += float(clean_value)
                                        except (ValueError, TypeError):
                                            continue
                                st.metric("📏 Volume total", f"{volume_total:.3f} m³")
                            else:
                                st.metric("📏 Volume total", "N/A")
                        except Exception as e:
                            st.metric("📏 Volume total", "Erreur")
                        
            except Exception as e:
                st.error(f"❌ Erreur lors de la récupération des détails : {str(e)}")
                # Debug information
                st.write("Détails de l'erreur :")
                if 'details_df' in locals():
                    st.write("Colonnes disponibles :", details_df.columns.tolist())
                    if not details_df.empty:
                        st.write("Aperçu des données :")
                        st.dataframe(details_df.head())
        else:
            st.info("Sélectionnez un client pour afficher les détails de la commande/estafettes.")
else:
    st.success("✅ Aucune proposition de location de camion en attente de décision.")

st.markdown("---")

# =====================================================
# 4. VOYAGES PAR ESTAFETTE OPTIMISÉ (Section 4 - Résultat final)
# =====================================================
st.header("4. 🚐 Voyages par Estafette Optimisé (Inclut Camions Loués)")

try:
    # Récupération sécurisée du DataFrame
    if st.session_state.rental_processor:
        df_optimized_estafettes = st.session_state.rental_processor.get_df_result()
    elif "df_voyages" in st.session_state:
        df_optimized_estafettes = st.session_state.df_voyages.copy()
    else:
        st.error("❌ Données non disponibles. Veuillez exécuter le traitement complet.")
        st.stop()
    
    # Vérifier que le DataFrame n'est pas vide
    if df_optimized_estafettes.empty:
        st.warning("⚠️ Aucune donnée à afficher.")
        st.stop()
    
    # CORRECTION : Nettoyer les colonnes en double
    df_clean = df_optimized_estafettes.loc[:, ~df_optimized_estafettes.columns.duplicated()]
    
    # CORRECTION : TRIER PAR ZONE D'ABORD
    if "Zone" in df_clean.columns:
        # Extraire le numéro de zone pour un tri numérique
        df_clean["Zone_Num"] = df_clean["Zone"].str.extract('(\d+)').astype(float)
        df_clean = df_clean.sort_values("Zone_Num").drop("Zone_Num", axis=1)
        # Alternative plus simple si l'extraction échoue :
        # df_clean = df_clean.sort_values("Zone")
    
    # Définir l'ordre des colonnes pour l'affichage
    colonnes_ordre = [
        "Zone", "Véhicule N°", "Poids total chargé", "Volume total chargé",
        "Client(s) inclus", "Représentant(s) inclus", "BL inclus", 
        "Taux d'occupation (%)", "Location_camion", "Location_proposee", "Code Véhicule"
    ]
    
    # Filtrer seulement les colonnes qui existent
    colonnes_finales = [col for col in colonnes_ordre if col in df_clean.columns]
    
    # Créer le DataFrame d'affichage avec retours à la ligne POUR STREAMLIT
    df_display = df_clean[colonnes_finales].copy()
    
    # Transformer les colonnes avec retours à la ligne HTML pour l'affichage Streamlit
    if "Client(s) inclus" in df_display.columns:
        df_display["Client(s) inclus"] = df_display["Client(s) inclus"].astype(str).apply(
            lambda x: "<br>".join(client.strip() for client in x.split(",")) if x != "nan" else ""
        )
    
    if "Représentant(s) inclus" in df_display.columns:
        df_display["Représentant(s) inclus"] = df_display["Représentant(s) inclus"].astype(str).apply(
            lambda x: "<br>".join(rep.strip() for rep in x.split(",")) if x != "nan" else ""
        )
    
    if "BL inclus" in df_display.columns:
        df_display["BL inclus"] = df_display["BL inclus"].astype(str).apply(
            lambda x: "<br>".join(bl.strip() for bl in x.split(";")) if x != "nan" else ""
        )
    
    # Formater les colonnes numériques pour l'affichage
    if "Poids total chargé" in df_display.columns:
        df_display["Poids total chargé"] = df_display["Poids total chargé"].map(lambda x: f"{x:.3f} kg")
    if "Volume total chargé" in df_display.columns:
        df_display["Volume total chargé"] = df_display["Volume total chargé"].map(lambda x: f"{x:.3f} m³")
    if "Taux d'occupation (%)" in df_display.columns:
        df_display["Taux d'occupation (%)"] = df_display["Taux d'occupation (%)"].map(lambda x: f"{x:.3f}%")
    
    # CSS POUR UN TABLEAU PROFESSIONNEL (identique aux autres sections)
    st.markdown("""
    <style>
    /* Style général du tableau */
    .custom-table-voyages {
        width: 100%;
        border-collapse: collapse;
        font-family: Arial, sans-serif;
        font-size: 14px;
        box-shadow: 0 2px 8px rgba(0,0,0,0.1);
        border-radius: 8px;
        overflow: hidden;
    }
    
    /* En-têtes du tableau - BLEU ROYAL SANS DÉGRADÉ */
    .custom-table-voyages th {
        background-color: #0369A1;
        color: white;
        padding: 12px 8px;
        text-align: center;
        border: 2px solid #4682B4;
        font-weight: bold;
        font-size: 13px;
        vertical-align: middle;
    }
    
    /* Cellules du tableau - TOUTES EN BLANC */
    .custom-table-voyages td {
        padding: 10px 8px;
        text-align: center;
        border: 1px solid #B0C4DE;
        background-color: white;
        color: #000000;
        vertical-align: middle;
    }
    
    /* Bordures visibles pour toutes les cellules */
    .custom-table-voyages th, 
    .custom-table-voyages td {
        border: 1px solid #B0C4DE !important;
    }
    
    /* Bordures épaisses pour l'extérieur du tableau */
    .custom-table-voyages {
        border: 2px solid #4682B4 !important;
    }
    
    /* Conteneur du tableau avec défilement horizontal */
    .table-container-voyages {
        overflow-x: auto;
        margin: 1rem 0;
        border-radius: 8px;
        border: 2px solid #4682B4;
    }
    
    /* Supprimer l'alternance des couleurs - TOUTES LES LIGNES BLANCHES */
    .custom-table-voyages tr:nth-child(even) td {
        background-color: white !important;
    }
    
    /* Survol des lignes - léger effet */
    .custom-table-voyages tr:hover td {
        background-color: #F0F8FF !important;
    }
    
    /* Style pour les cellules multilignes */
    .custom-table-voyages td {
        line-height: 1.4;
    }
    </style>
    """, unsafe_allow_html=True)
    
    # Afficher le tableau avec le style CSS professionnel
    html_table = df_display.to_html(escape=False, index=False, classes="custom-table-voyages", border=0)
    
    st.markdown(f"""
    <div class="table-container-voyages">
        {html_table}
    </div>
    """, unsafe_allow_html=True)
    
    # MÉTRIQUES RÉSUMÉES
    st.markdown("---")
    col1, col2, col3, col4 = st.columns(4)
    
    with col1:
        total_voyages = len(df_display)
        st.metric("🚐 Total Voyages", total_voyages)
    
    with col2:
        total_zones = df_display["Zone"].nunique() if "Zone" in df_display.columns else 0
        st.metric("🌍 Zones couvertes", total_zones)
    
    with col3:
        camions_loues = df_display["Location_camion"].sum() if "Location_camion" in df_display.columns else 0
        st.metric("🚚 Camions loués", int(camions_loues))
    
    with col4:
        estafettes = total_voyages - camions_loues
        st.metric("📦 Estafettes", estafettes)
    
    # Préparer l'export Excel avec retours à la ligne \n
    df_export = df_clean.copy()
    
    # CORRECTION : S'assurer que l'export est aussi trié par zone
    if "Zone" in df_export.columns:
        df_export["Zone_Num"] = df_export["Zone"].str.extract('(\d+)').astype(float)
        df_export = df_export.sort_values("Zone_Num").drop("Zone_Num", axis=1)
    
    # Transformer les colonnes avec retours à la ligne \n pour Excel
    if "Client(s) inclus" in df_export.columns:
        df_export["Client(s) inclus"] = df_export["Client(s) inclus"].astype(str).apply(
            lambda x: "\n".join(client.strip() for client in x.split(",")) if x != "nan" else ""
        )
    
    if "Représentant(s) inclus" in df_export.columns:
        df_export["Représentant(s) inclus"] = df_export["Représentant(s) inclus"].astype(str).apply(
            lambda x: "\n".join(rep.strip() for rep in x.split(",")) if x != "nan" else ""
        )
    
    if "BL inclus" in df_export.columns:
        df_export["BL inclus"] = df_export["BL inclus"].astype(str).apply(
            lambda x: "\n".join(bl.strip() for bl in x.split(";")) if x != "nan" else ""
        )
    
    # Formater les colonnes numériques pour l'export
    if "Poids total chargé" in df_export.columns:
        df_export["Poids total chargé"] = df_export["Poids total chargé"].round(3)
    if "Volume total chargé" in df_export.columns:
        df_export["Volume total chargé"] = df_export["Volume total chargé"].round(3)
    
    # Bouton de téléchargement avec formatage Excel
    from io import BytesIO
    import openpyxl
    from openpyxl.styles import Alignment
    
    excel_buffer = BytesIO()
    
    with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
        df_export.to_excel(writer, index=False, sheet_name="Voyages Optimisés")
        
        # Récupérer le workbook et worksheet pour appliquer le formatage
        workbook = writer.book
        worksheet = writer.sheets["Voyages Optimisés"]
        
        # Appliquer le style wrap_text aux colonnes avec retours à la ligne
        wrap_columns = []
        if "Client(s) inclus" in df_export.columns:
            wrap_columns.append("Client(s) inclus")
        if "Représentant(s) inclus" in df_export.columns:
            wrap_columns.append("Représentant(s) inclus")
        if "BL inclus" in df_export.columns:
            wrap_columns.append("BL inclus")
        
        # Appliquer le format wrap_text à toutes les cellules des colonnes concernées
        for col_idx, col_name in enumerate(df_export.columns):
            if col_name in wrap_columns:
                col_letter = openpyxl.utils.get_column_letter(col_idx + 1)
                for row in range(2, len(df_export) + 2):  # Commence à la ligne 2 (après l'en-tête)
                    cell = worksheet[f"{col_letter}{row}"]
                    cell.alignment = Alignment(wrap_text=True, vertical='top')
        
        # Ajuster la largeur des colonnes pour une meilleure visibilité
        for column in worksheet.columns:
            max_length = 0
            column_letter = openpyxl.utils.get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)  # Largeur max de 50
            worksheet.column_dimensions[column_letter].width = adjusted_width
    
    excel_buffer.seek(0)
    
    st.download_button(
        label="💾 Télécharger Voyages Estafette Optimisés",
        data=excel_buffer,
        file_name="Voyages_Estafette_Optimises.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    
    # Mise à jour pour les sections suivantes
    st.session_state.df_voyages = df_clean

except KeyError as e:
    st.error(f"❌ Erreur de colonne manquante : {e}")
    st.info("🔄 Tentative de récupération des données...")
    
    # Tentative de récupération
    if st.session_state.rental_processor:
        st.session_state.df_voyages = st.session_state.rental_processor.df_base.copy()
        st.rerun()
        
except Exception as e:
    st.error(f"❌ Erreur lors de l'affichage des voyages optimisés: {str(e)}")
    # Afficher les données brutes pour debug
    st.write("Données brutes pour debug:")
    if st.session_state.rental_processor:
        st.write("Colonnes du df_base:", list(st.session_state.rental_processor.df_base.columns))
# =====================================================
# 5️⃣ TRANSFERT DES BLs ENTRE ESTAFETTES / CAMIONS - VERSION AMÉLIORÉE
# =====================================================
st.markdown("## 🔁 Transfert de BLs entre Estafettes / Camions")

MAX_POIDS = 1550  # kg
MAX_VOLUME = 4.608  # m³

if "df_voyages" not in st.session_state:
    st.warning("⚠️ Vous devez d'abord exécuter la section 4 (Voyages par Estafette Optimisé).")
elif "df_livraisons" not in st.session_state:
    st.warning("⚠️ Le DataFrame des livraisons détaillées n'est pas disponible.")
else:
    df_voyages = st.session_state.df_voyages.copy()
    df_livraisons = st.session_state.df_livraisons.copy()

    colonnes_requises = ["Zone", "Véhicule N°", "Poids total chargé", "Volume total chargé", "BL inclus"]

    if not all(col in df_voyages.columns for col in colonnes_requises):
        st.error(f"❌ Le DataFrame ne contient pas toutes les colonnes nécessaires : {', '.join(colonnes_requises)}")
    else:
        zones_disponibles = sorted(df_voyages["Zone"].dropna().unique().tolist())
        zone_selectionnee = st.selectbox("🌍 Sélectionner une zone", zones_disponibles)

        if zone_selectionnee:
            df_zone = df_voyages[df_voyages["Zone"] == zone_selectionnee]
            vehicules = sorted(df_zone["Véhicule N°"].dropna().unique().tolist())

            col1, col2 = st.columns(2)
            with col1:
                source = st.selectbox("🚐 Estafette / Camion source", vehicules)
            with col2:
                cible = st.selectbox("🎯 Estafette / Camion cible", [v for v in vehicules if v != source])

            if source and cible:
                df_source = df_zone[df_zone["Véhicule N°"] == source]
                if df_source.empty or df_source["BL inclus"].isna().all():
                    st.warning("⚠️ Aucun BL trouvé pour ce véhicule source.")
                else:
                    st.subheader(f"📦 BLs actuellement assignés à {source}")

                    # --- NOUVEAU : Créer un mapping BL → Client ---
                    bls_avec_clients = []
                    bls_simples = df_source["BL inclus"].iloc[0].split(";")
                    
                    for bl in bls_simples:
                        # Trouver le client correspondant à ce BL
                        client_info = df_livraisons[df_livraisons["No livraison"] == bl]
                        if not client_info.empty:
                            client_nom = client_info["Client de l'estafette"].iloc[0]
                            bl_affichage = f"{bl} - {client_nom}"
                        else:
                            bl_affichage = f"{bl} - Client non trouvé"
                        bls_avec_clients.append(bl_affichage)
                    
                    # Affichage formaté avec clients
                    df_source_display = df_source[["Véhicule N°", "Poids total chargé", "Volume total chargé"]].copy()
                    df_source_display["BL inclus (avec clients)"] = "<br>".join(bls_avec_clients)
                    
                    df_source_display["Poids total chargé"] = df_source_display["Poids total chargé"].map(lambda x: f"{x:.3f} kg")
                    df_source_display["Volume total chargé"] = df_source_display["Volume total chargé"].map(lambda x: f"{x:.3f} m³")
                    
                    # CSS AMÉLIORÉ pour un tableau plus visible et bien centré
                    st.markdown("""
                    <style>
                    .centered-table {
                        margin-left: auto;
                        margin-right: auto;
                        display: table;
                        width: 100%;
                    }
                    .centered-table table {
                        margin: 0 auto;
                        border-collapse: collapse;
                        width: 100%;
                        font-family: Arial, sans-serif;
                        box-shadow: 0 2px 8px rgba(0,0,0,0.1);
                    }
                    .centered-table th {
                        background-color: #0369A1;
                        color: white;
                        padding: 12px 8px;
                        text-align: center;
                        border: 2px solid #555;
                        font-weight: bold;
                        font-size: 14px;
                        vertical-align: middle;  /* ← CENTRAGE VERTICAL AJOUTÉ */
                    }
                    .centered-table td {
                        padding: 10px 8px;
                        text-align: center;
                        border: 2px solid #555;
                        background-color: #f9f9f9;
                        color: #333;
                        vertical-align: middle;  /* ← CENTRAGE VERTICAL AJOUTÉ */
                    }
                    .centered-table tr:nth-child(even) td {
                        background-color: #f0f0f0;
                    }
                    .centered-table tr:hover td {
                        background-color: #e6f3ff;
                    }
                    </style>
                    """, unsafe_allow_html=True)
                    # CSS SPÉCIFIQUE POUR LE MULTISELECT - VERSION GRIS
                    st.markdown("""
                    <style>
                    /* ===== STYLES POUR LE MULTISELECT DES BLs ===== */

                    /* APPROCHE 1 : Style général pour tous les multiselect */
                    .stMultiSelect > div > div {
                        background-color: #F8FAFC !important;  /* GRIS TRÈS CLAIR */
                        border: 2px solid #CBD5E1 !important;  /* GRIS CLAIR */
                        border-radius: 8px !important;
                    }

                    /* APPROCHE 2 : Style pour le conteneur du multiselect */
                    div[data-baseweb="select"] > div {
                        background-color: #F8FAFC !important;  /* GRIS TRÈS CLAIR */
                        border: 2px solid #CBD5E1 !important;  /* GRIS CLAIR */
                        border-radius: 8px !important;
                    }

                    /* APPROCHE 3 : Style spécifique pour l'input */
                    div[data-baseweb="select"] > div:first-child {
                        background-color: #F8FAFC !important;  /* GRIS TRÈS CLAIR */
                        border: 2px solid #CBD5E1 !important;  /* GRIS CLAIR */
                        border-radius: 8px !important;
                    }

                    /* Style pour les tags des éléments sélectionnés */
                    div[data-baseweb="select"] span[data-baseweb="tag"] {
                        background-color: #0369A1 !important;  /* GRIS MOYEN */
                        color: white !important;
                        border-radius: 12px !important;
                        font-weight: bold;
                    }

                    /* Style pour la dropdown */
                    div[role="listbox"] {
                        background-color: white !important;
                        border: 2px solid #CBD5E1 !important;  /* GRIS CLAIR */
                    }

                    /* Options sélectionnées dans la liste */
                    div[role="option"][aria-selected="true"] {
                        background-color: #F1F5F9 !important;  /* GRIS TRÈS CLAIR */
                        color: #475569 !important;  /* GRIS FONCÉ */
                    }

                    /* Options au survol */
                    div[role="option"]:hover {
                        background-color: #E2E8F0 !important;  /* GRIS CLAIR */
                    }
                    </style>
                    """, unsafe_allow_html=True)
                    # --- NOUVEAU : Sélection avec clients ---
                    st.subheader("📋 Sélectionner les BLs à transférer")
                    
                    # Créer les options avec format "BL - Client"
                    options_transfert = []
                    mapping_bl_original = {}  # Pour garder la correspondance BL original
                    
                    for bl in bls_simples:
                        client_info = df_livraisons[df_livraisons["No livraison"] == bl]
                        if not client_info.empty:
                            client_nom = client_info["Client de l'estafette"].iloc[0]
                            option_affichage = f"{bl} - {client_nom}"
                        else:
                            option_affichage = f"{bl} - Client non trouvé"
                        
                        options_transfert.append(option_affichage)
                        mapping_bl_original[option_affichage] = bl
                    
                    # Multiselect avec clients
                    bls_selectionnes_affichage = st.multiselect(
                        "Sélectionnez les BLs à transférer (avec clients) :", 
                        options_transfert,
                        format_func=lambda x: x  # Affiche tel quel le format "BL - Client"
                    )
                    
                    # Convertir la sélection en BLs simples pour le traitement
                    bls_selectionnes = [mapping_bl_original[bl_affichage] for bl_affichage in bls_selectionnes_affichage]

                    if bls_selectionnes and st.button("🔁 Exécuter le transfert"):
                        df_bls_selection = df_livraisons[df_livraisons["No livraison"].isin(bls_selectionnes)]
                        poids_bls = df_bls_selection["Poids total"].sum()
                        volume_bls = df_bls_selection["Volume total"].sum()

                        df_cible = df_zone[df_zone["Véhicule N°"] == cible]
                        poids_cible = df_cible["Poids total chargé"].sum()
                        volume_cible = df_cible["Volume total chargé"].sum()

                        if (poids_cible + poids_bls) > MAX_POIDS or (volume_cible + volume_bls) > MAX_VOLUME:
                            st.warning("⚠️ Le transfert dépasse les limites de poids ou volume du véhicule cible.")
                        else:
                            def transfer_bl(row):
                                bls = row["BL inclus"].split(";") if pd.notna(row["BL inclus"]) else []
                                bls_to_move = [b for b in bls if b in bls_selectionnes]

                                if row["Véhicule N°"] == source:
                                    new_bls = [b for b in bls if b not in bls_to_move]
                                    row["BL inclus"] = ";".join(new_bls)
                                    row["Poids total chargé"] = max(0, row["Poids total chargé"] - poids_bls)
                                    row["Volume total chargé"] = max(0, row["Volume total chargé"] - volume_bls)
                                elif row["Véhicule N°"] == cible:
                                    new_bls = bls + bls_to_move
                                    row["BL inclus"] = ";".join(new_bls)
                                    row["Poids total chargé"] += poids_bls
                                    row["Volume total chargé"] += volume_bls
                                return row

                            df_voyages = df_voyages.apply(transfer_bl, axis=1)
                            st.session_state.df_voyages = df_voyages
                            
                            # Afficher un résumé du transfert avec clients
                            clients_transferes = df_bls_selection["Client de l'estafette"].unique()
                            st.success(f"""
                            ✅ Transfert réussi !
                            - **{len(bls_selectionnes)} BL(s)** déplacé(s) de **{source}** vers **{cible}**
                            - **Clients concernés :** {', '.join(clients_transferes)}
                            - **Poids transféré :** {poids_bls:.1f} kg
                            - **Volume transféré :** {volume_bls:.3f} m³
                            """)

                            # --- Affichage Streamlit avec retours à la ligne ---
                            st.subheader("📊 Voyages après transfert (toutes les zones)")
                            df_display = df_voyages.sort_values(by=["Zone", "Véhicule N°"]).copy()
                            
                            # Transformer les colonnes avec retours à la ligne HTML
                            if "BL inclus" in df_display.columns:
                                df_display["BL inclus"] = df_display["BL inclus"].astype(str).apply(
                                    lambda x: "<br>".join(bl.strip() for bl in x.split(";")) if x != "nan" else ""
                                )
                            
                            df_display["Poids total chargé"] = df_display["Poids total chargé"].map(lambda x: f"{x:.3f} kg")
                            df_display["Volume total chargé"] = df_display["Volume total chargé"].map(lambda x: f"{x:.3f} m³")
                            
                            # Affichage avec HTML amélioré pour les retours à la ligne et centrage
                            html_content_after = f"""
                            <div class="centered-table">
                            {df_display[colonnes_requises].to_html(escape=False, index=False)}
                            </div>
                            """
                            st.markdown(html_content_after, unsafe_allow_html=True)

                            # --- Export Excel avec retours à la ligne \n ---
                            df_export = df_voyages.copy()
                            
                            # Transformer les BL avec retours à la ligne \n pour Excel
                            if "BL inclus" in df_export.columns:
                                df_export["BL inclus"] = df_export["BL inclus"].astype(str).apply(
                                    lambda x: "\n".join(bl.strip() for bl in x.split(";")) if x != "nan" else ""
                                )
                            
                            df_export["Poids total chargé"] = df_export["Poids total chargé"].round(3)
                            df_export["Volume total chargé"] = df_export["Volume total chargé"].round(3)

                            from io import BytesIO
                            import openpyxl
                            from openpyxl.styles import Alignment
                            
                            excel_buffer = BytesIO()
                            with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
                                df_export.to_excel(writer, index=False, sheet_name='Transfert BLs')
                                
                                # Appliquer le format wrap_text pour Excel
                                workbook = writer.book
                                worksheet = writer.sheets['Transfert BLs']
                                
                                # Appliquer le style wrap_text à la colonne BL inclus
                                if "BL inclus" in df_export.columns:
                                    for col_idx, col_name in enumerate(df_export.columns):
                                        if col_name == "BL inclus":
                                            col_letter = openpyxl.utils.get_column_letter(col_idx + 1)
                                            for row in range(2, len(df_export) + 2):
                                                cell = worksheet[f"{col_letter}{row}"]
                                                cell.alignment = Alignment(wrap_text=True, vertical='top')
                            
                            excel_buffer.seek(0)

                            st.download_button(
                                label="💾 Télécharger le tableau mis à jour (XLSX)",
                                data=excel_buffer,
                                file_name="voyages_apres_transfert.xlsx",
                                mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                            )
# =====================================================
# 6️⃣ AJOUT D'OBJETS MANUELS AUX VÉHICULES
# =====================================================
st.markdown("## 📦 AJOUT D'OBJETS MANUELS AUX VÉHICULES")

if "df_voyages" in st.session_state:
    # Initialiser le gestionnaire de transfert si pas déjà fait
    if "transfer_manager" not in st.session_state:
        st.session_state.transfer_manager = TruckTransferManager(
            st.session_state.df_voyages, 
            st.session_state.df_livraisons
        )
    
    df_voyages = st.session_state.df_voyages.copy()
    
    #st.info("""
    #**Fonctionnalité :** Ajouter des objets manuels (colis urgents, matériel supplémentaire) 
    #à un véhicule existant. Le système vérifie automatiquement la capacité disponible.
    #""")
    
    col1, col2, col3 = st.columns(3)
    
    with col1:
        # Sélection de la zone
        zones_disponibles = sorted(df_voyages["Zone"].dropna().unique().tolist())
        zone_objet = st.selectbox("🌍 Zone", zones_disponibles, key="zone_objet")
    
    with col2:
        # Sélection du véhicule dans la zone choisie
        if zone_objet:
            vehicules_zone = sorted(
                df_voyages[df_voyages["Zone"] == zone_objet]["Véhicule N°"].dropna().unique().tolist()
            )
            vehicule_objet = st.selectbox("🚚 Véhicule", vehicules_zone, key="vehicule_objet")
        else:
            vehicule_objet = st.selectbox("🚚 Véhicule", [], key="vehicule_objet")
    
    with col3:
        # Informations sur le véhicule sélectionné
        if zone_objet and vehicule_objet:
            vehicule_data = df_voyages[
                (df_voyages["Zone"] == zone_objet) & 
                (df_voyages["Véhicule N°"] == vehicule_objet)
            ].iloc[0]
            
            is_camion = vehicule_data.get("Code Véhicule", "") == "CAMION-LOUE"
            capacite_poids = 30500 if is_camion else 1550
            capacite_volume = 77.5 if is_camion else 4.608
            
            poids_actuel = vehicule_data.get("Poids total chargé", 0)
            volume_actuel = vehicule_data.get("Volume total chargé", 0)
            
            st.metric(
                "📊 Capacité utilisée", 
                f"{poids_actuel:.1f}kg / {capacite_poids}kg",
                f"{volume_actuel:.3f}m³ / {capacite_volume}m³"
            )
    
    # Formulaire d'ajout d'objet
    st.markdown("### 📝 Détails de l'objet à ajouter")
    
    # CSS personnalisé pour les couleurs
    st.markdown("""
    <style>
    .custom-border {
        border: 2px solid #1f77b4;
        border-radius: 5px;
        padding: 10px;
        margin: 5px 0px;
    }
    .custom-button {
        background-color: #1f77b4 !important;
        color: white !important;
        border: none !important;
    }
    .custom-button:hover {
        background-color: #1668a5 !important;
        color: white !important;
    }
    </style>
    """, unsafe_allow_html=True)
    
    col4, col5, col6 = st.columns(3)
    
    with col4:
        st.markdown('<div class="custom-border">', unsafe_allow_html=True)
        nom_objet = st.text_input("🏷️ Nom de l'objet", placeholder="Ex: Matériel urgent, Colis oublié...")
        st.markdown('</div>', unsafe_allow_html=True)
    
    with col5:
        st.markdown('<div class="custom-border">', unsafe_allow_html=True)
        poids_objet = st.number_input("⚖️ Poids (kg)", min_value=0.0, max_value=1000.0, value=10.0, step=0.1)
        st.markdown('</div>', unsafe_allow_html=True)
    
    with col6:
        st.markdown('<div class="custom-border">', unsafe_allow_html=True)
        volume_objet = st.number_input("📦 Volume (m³)", min_value=0.0, max_value=10.0, value=0.1, step=0.01)
        st.markdown('</div>', unsafe_allow_html=True)
    
    # Bouton d'ajout avec couleur de fond personnalisée
    st.markdown("""
    <style>
    div.stButton > button:first-child {
        background-color: #1f77b4;
        color: white;
        border: none;
    }
    div.stButton > button:first-child:hover {
        background-color: #1668a5;
        color: white;
    }
    </style>
    """, unsafe_allow_html=True)
    
    if st.button("➕ Ajouter l'objet au véhicule", type="primary"):
        if not nom_objet:
            st.error("❌ Veuillez donner un nom à l'objet.")
        elif zone_objet and vehicule_objet:
            try:
                # Appel de la méthode add_manual_object
                success, message, df_updated = st.session_state.transfer_manager.add_manual_object(
                    df_voyages=df_voyages,
                    vehicle=vehicule_objet,
                    zone=zone_objet,
                    name=nom_objet,
                    weight=poids_objet,
                    volume=volume_objet
                )
                
                if success:
                    st.success(message)
                    
                    # =====================================================
                    # MÉCANISME DE MISE À JOUR FORCÉE DE TOUTES LES DONNÉES
                    # =====================================================
                    
                    # 1. Mettre à jour le DataFrame principal dans session_state
                    st.session_state.df_voyages = df_updated
                    
                    # 2. Synchroniser le gestionnaire de transfert
                    st.session_state.transfer_manager.df_voyages = df_updated.copy()
                    
                    # 3. Synchroniser le processeur de location si disponible
                    if st.session_state.rental_processor:
                        try:
                            # Méthode 1 : Mettre à jour directement le df_base
                            st.session_state.rental_processor.df_base = df_updated.copy()
                            
                            # Méthode 2 : Recréer le processeur si nécessaire
                            st.session_state.rental_processor = TruckRentalProcessor(
                                df_updated, 
                                st.session_state.df_livraisons_original
                            )
                            
                            st.success("✅ Processeur de location synchronisé")
                        except Exception as e:
                            st.warning(f"⚠️ Synchronisation partielle du processeur : {str(e)}")
                    
                    # 4. Mettre à jour les propositions de location si elles existent
                    if st.session_state.propositions is not None:
                        try:
                            st.session_state.propositions = st.session_state.rental_processor.detecter_propositions()
                        except:
                            pass  # Ignorer si la mise à jour des propositions échoute
                    
                    # 5. Mettre à jour les voyages validés si ils existent
                    if 'df_voyages_valides' in st.session_state:
                        try:
                            # Recréer les voyages validés à partir des nouvelles données
                            mask_valides = df_updated["Véhicule N°"].isin(
                                st.session_state.df_voyages_valides["Véhicule N°"]
                            )
                            st.session_state.df_voyages_valides = df_updated[mask_valides].copy()
                        except:
                            pass  # Ignorer si la mise à jour des validations échoute
                    
                    # Afficher le véhicule mis à jour
                    vehicule_update = df_updated[
                        (df_updated["Zone"] == zone_objet) & 
                        (df_updated["Véhicule N°"] == vehicule_objet)
                    ].iloc[0]
                    
                    st.info(f"""
                    **Véhicule mis à jour :**
                    - Poids total : {vehicule_update['Poids total chargé']:.1f} kg
                    - Volume total : {vehicule_update['Volume total chargé']:.3f} m³
                    - Taux d'occupation : {vehicule_update['Taux d\'occupation (%)']:.1f}%
                    - BLs inclus : {vehicule_update['BL inclus']}
                    """)
                    
                    # Afficher un résumé des modifications
                    st.success("🔄 Toutes les données ont été mises à jour avec succès !")
                    
                    # FORCER L'ACTUALISATION COMPLÈTE DE L'APPLICATION
                    st.rerun()
                    
                else:
                    st.error(message)
                    
            except Exception as e:
                st.error(f"❌ Erreur lors de l'ajout de l'objet : {str(e)}")
                # Debug information
                st.error(f"Debug - Zone: {zone_objet}, Véhicule: {vehicule_objet}")
        else:
            st.error("❌ Veuillez sélectionner une zone et un véhicule.")
    
    # Affichage des objets ajoutés récemment
    st.markdown("### 📋 Historique des objets ajoutés")
    
    # Rechercher les objets manuels dans les BLs
    objets_manuels = []
    for idx, row in df_voyages.iterrows():
        bls = str(row.get("BL inclus", ""))
        if "OBJ-" in bls:
            for bl in bls.split(";"):
                if bl.startswith("OBJ-"):
                    # Trouver le véhicule correspondant dans les données mises à jour
                    vehicule_info = df_voyages[
                        (df_voyages["Zone"] == row["Zone"]) & 
                        (df_voyages["Véhicule N°"] == row["Véhicule N°"])
                    ]
                    if not vehicule_info.empty:
                        poids_vehicule = vehicule_info["Poids total chargé"].iloc[0]
                        volume_vehicule = vehicule_info["Volume total chargé"].iloc[0]
                        
                        objets_manuels.append({
                            "Véhicule": row["Véhicule N°"],
                            "Zone": row["Zone"],
                            "Objet": bl,
                            "Poids Véhicule": f"{poids_vehicule:.1f} kg",
                            "Volume Véhicule": f"{volume_vehicule:.3f} m³",
                            "Type": "Camion" if row.get("Code Véhicule", "") == "CAMION-LOUE" else "Estafette"
                        })
    
    if objets_manuels:
        df_objets = pd.DataFrame(objets_manuels)
        show_df(df_objets, use_container_width=True)
        
        # Bouton pour supprimer tous les objets (optionnel)
        col_clear1, col_clear2 = st.columns([3, 1])
        with col_clear2:
            if st.button("🗑️ Supprimer tous les objets", type="secondary"):
                # Réinitialiser les données sans objets manuels
                df_sans_objets = st.session_state.df_voyages.copy()
                for idx, row in df_sans_objets.iterrows():
                    bls_originaux = str(row["BL inclus"]).split(";")
                    bls_filtres = [bl for bl in bls_originaux if not bl.startswith("OBJ-")]
                    df_sans_objets.at[idx, "BL inclus"] = ";".join(bls_filtres)
                
                # Réappliquer la mise à jour forcée
                st.session_state.df_voyages = df_sans_objets
                st.session_state.transfer_manager.df_voyages = df_sans_objets.copy()
                if st.session_state.rental_processor:
                    st.session_state.rental_processor.df_base = df_sans_objets.copy()
                
                st.success("✅ Tous les objets manuels ont été supprimés")
                st.rerun()
    else:
        st.info(" Aucun objet manuel ajouté pour le moment.")

#else:
    #st.warning("⚠️ Vous devez d'abord exécuter la section 4 (Voyages par Estafette Optimisé).")
# =====================================================
# 7️⃣ VALIDATION DES VOYAGES APRÈS TRANSFERT
# =====================================================
st.markdown("## ✅ VALIDATION DES VOYAGES APRÈS TRANSFERT")

from io import BytesIO

# --- Fonction pour exporter DataFrame en Excel avec arrondi ---
def to_excel(df, sheet_name="Voyages Validés"):
    df_export = df.copy()
    if "Poids total chargé" in df_export.columns:
        df_export["Poids total chargé"] = df_export["Poids total chargé"].round(3)
    if "Volume total chargé" in df_export.columns:
        df_export["Volume total chargé"] = df_export["Volume total chargé"].round(3)

    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df_export.to_excel(writer, index=False, sheet_name=sheet_name)
    return output.getvalue()

# --- CSS pour améliorer l'apparence ---
st.markdown("""
<style>
.voyage-card {
    border: 1px solid #e0e0e0;
    border-radius: 10px;
    padding: 20px;
    margin: 10px 0;
    background: white;
    box-shadow: 0 2px 4px rgba(0,0,0,0.1);
}
.voyage-header {
    background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
    color: white;
    padding: 15px;
    border-radius: 8px;
    margin-bottom: 15px;
}
.metric-card {
    background: #f8f9fa;
    border-left: 4px solid #667eea;
    padding: 12px;
    margin: 8px 0;
    border-radius: 5px;
}
.bl-list {
    background: #fff3cd;
    border: 1px solid #ffeaa7;
    border-radius: 5px;
    padding: 10px;
    margin: 10px 0;
    max-height: 150px;
    overflow-y: auto;
}
.validation-buttons {
    display: flex;
    gap: 10px;
    margin-top: 15px;
}
</style>
""", unsafe_allow_html=True)

# --- Création du DataFrame de validation à partir du df_voyages ---
if "df_voyages" in st.session_state:
    voyages_apres_transfert = st.session_state.df_voyages.copy()
    df_validation = voyages_apres_transfert.copy()

    if "validations" not in st.session_state:
        st.session_state.validations = {}

    # --- Affichage amélioré des voyages ---
    st.markdown("### 📋 Liste des Voyages à Valider")
    
    for idx, row in df_validation.iterrows():
        # Création d'une carte pour chaque voyage
        with st.container():
            st.markdown(f"""
            <div class="voyage-card">
                <div class="voyage-header">
                    <h4>🚚 Voyage {row['Véhicule N°']} | Zone: {row['Zone']}</h4>
                </div>
            """, unsafe_allow_html=True)
            
            # Métriques principales
            col1, col2, col3 = st.columns(3)
            
            with col1:
                st.markdown(f"""
                <div class="metric-card">
                    <strong>⚖️ Poids Total</strong><br>
                    {row['Poids total chargé']:.3f} kg
                </div>
                """, unsafe_allow_html=True)
            
            with col2:
                st.markdown(f"""
                <div class="metric-card">
                    <strong>📏 Volume Total</strong><br>
                    {row['Volume total chargé']:.3f} m³
                </div>
                """, unsafe_allow_html=True)
            
            with col3:
                taux_occupation = row.get('Taux d\'occupation (%)', 'N/A')
                if taux_occupation != 'N/A':
                    taux_text = f"{taux_occupation:.1f}%"
                else:
                    taux_text = "N/A"
                st.markdown(f"""
                <div class="metric-card">
                    <strong>📊 Taux d'Occupation</strong><br>
                    {taux_text}
                </div>
                """, unsafe_allow_html=True)
            
            # Informations détaillées
            col4, col5 = st.columns(2)
            
            with col4:
                clients = row.get('Client(s) inclus', '')
                if clients:
                    st.markdown(f"**👥 Clients:** {clients}")
                
                representants = row.get('Représentant(s) inclus', '')
                if representants:
                    st.markdown(f"**👨‍💼 Représentants:** {representants}")
            
            with col5:
                location = "✅ Oui" if row.get('Location_camion') else "❌ Non"
                st.markdown(f"**🚛 Location:** {location}")
                
                code_vehicule = row.get('Code Véhicule', 'N/A')
                st.markdown(f"**🔧 Code Véhicule:** {code_vehicule}")
            
            # Liste des BL avec défilement
            bls = row.get('BL inclus', '')
            if bls:
                bls_list = bls.split(';')
                bls_html = "<br>".join([f"• {bl.strip()}" for bl in bls_list])
                st.markdown(f"""
                <div class="bl-list">
                    <strong>📋 BLs Inclus ({len(bls_list)}):</strong><br>
                    {bls_html}
                </div>
                """, unsafe_allow_html=True)
            
            # Boutons de validation côte à côte
            st.markdown("**✅ Validation du voyage:**")
            col_oui, col_non = st.columns(2)
            
            with col_oui:
                if st.button(f"✅ Valider {row['Véhicule N°']}", key=f"btn_oui_{idx}", 
                           use_container_width=True, type="primary" if st.session_state.validations.get(idx) == "Oui" else "secondary"):
                    st.session_state.validations[idx] = "Oui"
                    st.rerun()
            
            with col_non:
                if st.button(f"❌ Rejeter {row['Véhicule N°']}", key=f"btn_non_{idx}",
                           use_container_width=True, type="primary" if st.session_state.validations.get(idx) == "Non" else "secondary"):
                    st.session_state.validations[idx] = "Non"
                    st.rerun()
            
            # Afficher le statut actuel
            statut = st.session_state.validations.get(idx)
            if statut == "Oui":
                st.success(f"✅ Voyage {row['Véhicule N°']} validé")
            elif statut == "Non":
                st.error(f"❌ Voyage {row['Véhicule N°']} rejeté")
            else:
                st.info("⏳ En attente de validation")
            
            st.markdown("</div>", unsafe_allow_html=True)
            st.markdown("---")

    # --- Résumé des validations ---
    st.markdown("### 📊 Résumé des Validations")
    total_voyages = len(df_validation)
    valides = sum(1 for v in st.session_state.validations.values() if v == "Oui")
    rejetes = sum(1 for v in st.session_state.validations.values() if v == "Non")

    col1, col2, col3 = st.columns(3)

    with col1:
        st.metric("Total Voyages", total_voyages)
    with col2:
        st.metric("✅ Validés", valides)
    with col3:
        st.metric("❌ Rejetés", rejetes)

    # Information supplémentaire sur l'état des validations
    if valides + rejetes < total_voyages:
        st.info(f"ℹ️ {total_voyages - (valides + rejetes)} voyage(s) n'ont pas encore été validés")

    # --- Bouton pour appliquer les validations ---
    if st.button("🚀 Finaliser la Validation", type="primary", use_container_width=True):
        valid_indexes = [i for i, v in st.session_state.validations.items() if v == "Oui"]
        valid_indexes = [i for i in valid_indexes if i in df_validation.index]

        if valid_indexes:
            df_voyages_valides = df_validation.loc[valid_indexes].reset_index(drop=True)
            st.session_state.df_voyages_valides = df_voyages_valides

            st.success(f"✅ {len(df_voyages_valides)} voyage(s) validé(s) avec succès!")
            
            # Affichage des voyages validés
            st.markdown("### 🎉 Voyages Validés - Résumé Final")
            
            for idx, row_valide in df_voyages_valides.iterrows():
                with st.expander(f"🚚 {row_valide['Véhicule N°']} - Zone {row_valide['Zone']}", expanded=True):
                    col1, col2 = st.columns(2)
                    with col1:
                        st.metric("Poids", f"{row_valide['Poids total chargé']:.3f} kg")
                        st.metric("Clients", row_valide.get('Client(s) inclus', 'N/A'))
                    with col2:
                        st.metric("Volume", f"{row_valide['Volume total chargé']:.3f} m³")
                        st.metric("Représentants", row_valide.get('Représentant(s) inclus', 'N/A'))

            # --- Export Excel ---
            excel_data = to_excel(df_voyages_valides)
            st.download_button(
                label="💾 Télécharger les voyages validés (XLSX)",
                data=excel_data,
                file_name="Voyages_valides.xlsx",
                mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                use_container_width=True
            )
        else:
            st.warning("⚠️ Aucun voyage n'a été validé. Veuillez valider au moins un voyage.")

else:
    st.warning("⚠️ Vous devez d'abord exécuter la section 4 (Voyages par Estafette Optimisé).")

# =====================================================
# 8️⃣ ATTRIBUTION DES VÉHICULES ET CHAUFFEURS
# =====================================================
st.markdown("## 🚛 ATTRIBUTION DES VÉHICULES ET CHAUFFEURS")

if 'df_voyages_valides' in st.session_state and not st.session_state.df_voyages_valides.empty:

    df_attribution = st.session_state.df_voyages_valides.copy()

    # Fonction pour formatter les colonnes avec retours à la ligne POUR STREAMLIT
    def formater_colonnes_listes_streamlit(df):
        df_formate = df.copy()
        colonnes_a_formater = ['Client(s) inclus', 'Représentant(s) inclus', 'BL inclus']
        
        for col in colonnes_a_formater:
            if col in df_formate.columns:
                df_formate[col] = df_formate[col].apply(
                    lambda x: '\n'.join([elem.strip() for elem in str(x).replace(';', ',').split(',') if elem.strip()]) 
                    if pd.notna(x) else ""
                )
        return df_formate

    if "attributions" not in st.session_state:
        st.session_state.attributions = {}

    for idx, row in df_attribution.iterrows():
        with st.expander(f"🚚 Voyage {row['Véhicule N°']} | Zone : {row['Zone']}"):
            st.write("**Informations du voyage :**")
            
            # Créer un affichage personnalisé avec retours à ligne
            col1, col2, col3 = st.columns(3)
            
            with col1:
                st.write(f"**Zone:** {row['Zone']}")
                st.write(f"**Véhicule N°:** {row['Véhicule N°']}")
                if "Poids total chargé" in row:
                    st.write(f"**Poids total chargé:** {row['Poids total chargé']:.2f} kg")
                if "Volume total chargé" in row:
                    st.write(f"**Volume total chargé:** {row['Volume total chargé']:.3f} m³")
                if "Taux d'occupation (%)" in row:
                    st.write(f"**Taux d'occupation:** {row['Taux d\'occupation (%)']:.1f}%")
            
            with col2:
                # Afficher les clients avec retours à ligne
                if 'Client(s) inclus' in row and pd.notna(row['Client(s) inclus']):
                    st.write("**Clients:**")
                    clients = str(row['Client(s) inclus']).replace(';', ',').split(',')
                    for client in clients:
                        client_clean = client.strip()
                        if client_clean:
                            st.write(f"- {client_clean}")
                
                # Afficher les représentants avec retours à ligne
                if 'Représentant(s) inclus' in row and pd.notna(row['Représentant(s) inclus']):
                    st.write("**Représentants:**")
                    representants = str(row['Représentant(s) inclus']).replace(';', ',').split(',')
                    for rep in representants:
                        rep_clean = rep.strip()
                        if rep_clean:
                            st.write(f"- {rep_clean}")
            
            with col3:
                # Afficher les BL avec retours à ligne
                if 'BL inclus' in row and pd.notna(row['BL inclus']):
                    st.write("**BL associés:**")
                    bls = str(row['BL inclus']).replace(';', ',').split(',')
                    for bl in bls:
                        bl_clean = bl.strip()
                        if bl_clean:
                            st.write(f"- {bl_clean}")

            col_veh, col_chauf = st.columns(2)
            
            with col_veh:
                vehicule_selectionne = st.selectbox(
                    f"Véhicule pour le voyage {row['Véhicule N°']}",
                    VEHICULES_DISPONIBLES,
                    index=0 if st.session_state.attributions.get(idx, {}).get("Véhicule") else 0,
                    key=f"vehicule_{idx}"
                )
            
            with col_chauf:
                options_chauffeurs = [f"{matricule} - {nom}" for matricule, nom in CHAUFFEURS_DETAILS.items() if matricule != 'Matricule']
                
                default_index = 0
                chauffeur_actuel = st.session_state.attributions.get(idx, {}).get("Chauffeur_complet")
                if chauffeur_actuel and chauffeur_actuel in options_chauffeurs:
                    default_index = options_chauffeurs.index(chauffeur_actuel)
                
                chauffeur_selectionne_complet = st.selectbox(
                    f"Chauffeur pour le voyage {row['Véhicule N°']}",
                    options_chauffeurs,
                    index=default_index,
                    key=f"chauffeur_{idx}"
                )
                
                if chauffeur_selectionne_complet:
                    matricule_chauffeur = chauffeur_selectionne_complet.split(" - ")[0]
                    nom_chauffeur = chauffeur_selectionne_complet.split(" - ")[1]
                else:
                    matricule_chauffeur = ""
                    nom_chauffeur = ""

            st.session_state.attributions[idx] = {
                "Véhicule": vehicule_selectionne,
                "Chauffeur_complet": chauffeur_selectionne_complet,
                "Matricule_chauffeur": matricule_chauffeur,
                "Nom_chauffeur": nom_chauffeur
            }

    if st.button("✅ Appliquer les attributions"):

        df_attribution["Véhicule attribué"] = df_attribution.index.map(lambda i: st.session_state.attributions[i]["Véhicule"])
        df_attribution["Chauffeur attribué"] = df_attribution.index.map(lambda i: st.session_state.attributions[i]["Nom_chauffeur"])
        df_attribution["Matricule chauffeur"] = df_attribution.index.map(lambda i: st.session_state.attributions[i]["Matricule_chauffeur"])

        
        st.markdown("### 📦 Voyages avec Véhicule et Chauffeur")

        # --- Affichage Streamlit amélioré avec retours à ligne ---
        for idx, row in df_attribution.iterrows():
            with st.expander(f"📋 Voyage {row['Véhicule N°']} - Zone {row['Zone']} - Véhicule: {row.get('Véhicule attribué', 'N/A')} - Chauffeur: {row.get('Chauffeur attribué', 'N/A')}"):
                col1, col2, col3 = st.columns(3)
                
                with col1:
                    st.write("**Informations de base:**")
                    st.write(f"**Zone:** {row['Zone']}")
                    st.write(f"**Véhicule N°:** {row['Véhicule N°']}")
                    if "Poids total chargé" in row:
                        st.write(f"**Poids total chargé:** {row['Poids total chargé']:.3f} kg")
                    if "Volume total chargé" in row:
                        st.write(f"**Volume total chargé:** {row['Volume total chargé']:.3f} m³")
                    if "Taux d'occupation (%)" in row:
                        st.write(f"**Taux d'occupation:** {row['Taux d\'occupation (%)']:.3f}%")
                    if "Véhicule attribué" in row:
                        st.write(f"**Véhicule attribué:** {row['Véhicule attribué']}")
                    if "Chauffeur attribué" in row:
                        st.write(f"**Chauffeur attribué:** {row['Chauffeur attribué']}")
                    if "Matricule chauffeur" in row:
                        st.write(f"**Matricule chauffeur:** {row['Matricule chauffeur']}")
                
                with col2:
                    # Afficher les clients avec retours à ligne
                    if 'Client(s) inclus' in row and pd.notna(row['Client(s) inclus']):
                        st.write("**📋 Clients inclus:**")
                        clients = str(row['Client(s) inclus']).replace(';', ',').split(',')
                        for client in clients:
                            client_clean = client.strip()
                            if client_clean:
                                st.write(f"• {client_clean}")
                    
                    # Afficher les représentants avec retours à ligne
                    if 'Représentant(s) inclus' in row and pd.notna(row['Représentant(s) inclus']):
                        st.write("**👤 Représentants inclus:**")
                        representants = str(row['Représentant(s) inclus']).replace(';', ',').split(',')
                        for rep in representants:
                            rep_clean = rep.strip()
                            if rep_clean:
                                st.write(f"• {rep_clean}")
                
                with col3:
                    # Afficher les BL avec retours à ligne
                    if 'BL inclus' in row and pd.notna(row['BL inclus']):
                        st.write("**📄 BL associés:**")
                        bls = str(row['BL inclus']).replace(';', ',').split(',')
                        # Afficher en colonnes si beaucoup de BL
                        if len(bls) > 5:
                            cols = st.columns(2)
                            half = len(bls) // 2
                            for i, bl in enumerate(bls):
                                bl_clean = bl.strip()
                                if bl_clean:
                                    col_idx = 0 if i < half else 1
                                    with cols[col_idx]:
                                        st.write(f"• {bl_clean}")
                        else:
                            for bl in bls:
                                bl_clean = bl.strip()
                                if bl_clean:
                                    st.write(f"• {bl_clean}")

        # --- Export Excel avec retours à ligne et CENTRAGE ---
        from io import BytesIO
        import openpyxl

        def to_excel(df):
            df_export = df.copy()
            
            # Formater les colonnes avec retours à ligne pour Excel
            colonnes_a_formater = ['Client(s) inclus', 'Représentant(s) inclus', 'BL inclus']
            for col in colonnes_a_formater:
                if col in df_export.columns:
                    df_export[col] = df_export[col].apply(
                        lambda x: '\n'.join([elem.strip() for elem in str(x).replace(';', ',').split(',') if elem.strip()]) 
                        if pd.notna(x) else ""
                    )
            
            if "Poids total chargé" in df_export.columns:
                df_export["Poids total chargé"] = df_export["Poids total chargé"].round(3)
            if "Volume total chargé" in df_export.columns:
                df_export["Volume total chargé"] = df_export["Volume total chargé"].round(3)
            
            output = BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                df_export.to_excel(writer, index=False, sheet_name='Voyages_Attribués')
                
                # Appliquer le formatage des retours à ligne et CENTRAGE dans Excel
                workbook = writer.book
                worksheet = writer.sheets['Voyages_Attribués']
                
                # Style de centrage avec retours à ligne
                center_alignment = openpyxl.styles.Alignment(
                    horizontal='center', 
                    vertical='center', 
                    wrap_text=True
                )
                
                # Appliquer le centrage à TOUTES les cellules
                for row in worksheet.iter_rows(min_row=1, max_row=len(df_export) + 1, min_col=1, max_col=len(df_export.columns)):
                    for cell in row:
                        cell.alignment = center_alignment
                
                # Ajuster automatiquement la largeur des colonnes
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if cell.value:
                                # Calculer la longueur maximale en prenant en compte les retours à ligne
                                lines = str(cell.value).split('\n')
                                max_line_length = max(len(line) for line in lines)
                                max_length = max(max_length, max_line_length)
                        except:
                            pass
                    adjusted_width = min(50, (max_length + 2))  # Limiter à 50 caractères max
                    worksheet.column_dimensions[column_letter].width = adjusted_width
                
                # Ajuster la hauteur des lignes pour les retours à ligne
                for row in range(2, len(df_export) + 2):  # Commencer à la ligne 2 (après l'en-tête)
                    worksheet.row_dimensions[row].height = 60  # Hauteur fixe pour accommoder les retours à ligne
            
            return output.getvalue()

        # --- Export PDF avec données centrées verticalement ---
        from fpdf import FPDF

        def to_pdf_better_centered(df, title="Voyages Attribués"):
            pdf = FPDF(orientation='L')
            pdf.add_page()
            
            # Titre
            pdf.set_font("Arial", 'B', 16)
            pdf.cell(0, 15, title, ln=True, align="C")
            pdf.ln(5)
            
            # Créer une copie formatée pour le PDF
            df_pdf = df.copy()
            
            # Formater les nombres avec 3 chiffres après la virgule
            numeric_columns = {
                'Poids total chargé': 'kg',
                'Volume total chargé': 'm³', 
                'Taux d\'occupation (%)': '%'
            }
            
            for col, unit in numeric_columns.items():
                if col in df_pdf.columns:
                    df_pdf[col] = df_pdf[col].apply(
                        lambda x: f"{float(x):.3f} {unit}" if x and str(x).strip() and str(x).strip() != 'nan' else ""
                    )
            
            # Configuration des colonnes
            col_config = {
                'Zone': {'width': 14, 'header': 'Zone'},
                'Véhicule N°': {'width': 16, 'header': 'Véhicule'},
                'Poids total chargé': {'width': 22, 'header': 'Poids (kg)'},
                'Volume total chargé': {'width': 22, 'header': 'Volume (m³)'},
                'Client(s) inclus': {'width': 30, 'header': 'Clients'},
                'Représentant(s) inclus': {'width': 26, 'header': 'Représentants'},
                'BL inclus': {'width': 36, 'header': 'BL associés'},
                'Taux d\'occupation (%)': {'width': 18, 'header': 'Taux %'},
                'Véhicule attribué': {'width': 2, 'header': 'Véhicule Attribué'},
                'Chauffeur attribué': {'width': 28, 'header': 'Chauffeur'},
                'Matricule chauffeur': {'width': 18, 'header': 'Matricule'}
            }
            
            # Sélectionner seulement les colonnes existantes
            colonnes_existantes = [col for col in df_pdf.columns if col in col_config]
            widths = [col_config[col]['width'] for col in colonnes_existantes]
            headers = [col_config[col]['header'] for col in colonnes_existantes]
            
            # En-têtes
            pdf.set_font("Arial", 'B', 9)
            for i, header in enumerate(headers):
                pdf.cell(widths[i], 8, header, border=1, align='C')
            pdf.ln()
            
            # Données avec centrage vertical optimal
            pdf.set_font("Arial", '', 8)
            
            for voyage_idx, (_, row) in enumerate(df_pdf.iterrows()):
                # Déterminer le nombre de lignes nécessaires pour ce voyage
                list_columns = ['Client(s) inclus', 'Représentant(s) inclus', 'BL inclus']
                non_list_columns = [col for col in colonnes_existantes if col not in list_columns]
                
                max_lines = 1
                list_contents = {}
                
                for col in list_columns:
                    if col in colonnes_existantes:
                        content = str(row[col]) if pd.notna(row[col]) and str(row[col]) != 'nan' else ""
                        elements = content.replace(';', ',').split(',')
                        elements = [elem.strip() for elem in elements if elem.strip()]
                        list_contents[col] = elements
                        max_lines = max(max_lines, len(elements))
                
                # Pour les voyages avec peu de lignes, on centre sur la première ligne
                if max_lines <= 3:
                    display_line = 0  # Première ligne pour les petits blocs
                else:
                    display_line = max_lines // 2  # Milieu pour les grands blocs
                
                # Écrire le voyage
                for line_idx in range(max_lines):
                    for i, col in enumerate(colonnes_existantes):
                        if col in list_columns:
                            # Colonnes de liste
                            elements = list_contents.get(col, [])
                            content = elements[line_idx] if line_idx < len(elements) else ""
                        else:
                            # Colonnes non-liste - afficher sur la ligne de centrage
                            if line_idx == display_line:
                                content = str(row[col]) if pd.notna(row[col]) and str(row[col]) != 'nan' else ""
                            else:
                                content = ""
                        
                        # Bordures
                        border = ''
                        if line_idx == 0: border += 'T'
                        if line_idx == max_lines - 1: border += 'B'
                        if i == 0: border += 'L'
                        if i == len(colonnes_existantes) - 1: border += 'R'
                        
                        pdf.cell(widths[i], 6, content, border=border, align='C')
                    
                    pdf.ln()
            
            return pdf.output(dest='S').encode('latin-1')

        # Afficher les boutons de téléchargement côte à côte
        col1, col2 = st.columns(2)

        with col1:
            st.download_button(
                label="💾 Télécharger le tableau final (XLSX)",
                data=to_excel(df_attribution),
                file_name="Voyages_attribues.xlsx",
                mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )

        with col2:
            st.download_button(
                label="📄 Télécharger le tableau final (PDF)",
                data=to_pdf_better_centered(df_attribution),
                file_name="Voyages_attribues.pdf",
                mime='application/pdf'
            )
                
        # Mettre à jour le session state
        st.session_state.df_voyages_valides = df_attribution
        st.success("✅ Attributions appliquées avec succès !")
        
else:
    st.warning("⚠️ Vous devez d'abord valider les voyages dans la section 7.")

# =====================================================
# 9️⃣ RAPPORTS AVANCÉS ET ANALYTICS
# =====================================================
st.markdown("## 📊 RAPPORTS AVANCÉS ET ANALYTICS")

if "df_voyages" in st.session_state and "df_livraisons_original" in st.session_state:
    
    # Initialiser le générateur de rapports
    from backend import AdvancedReportGenerator
    report_generator = AdvancedReportGenerator(
        st.session_state.df_voyages, 
        st.session_state.df_livraisons_original
    )
    
    tab1, tab2, tab3, tab4 = st.tabs([
        "📈 Rapport Analytique", 
        "👤 Rapport Client", 
        "💰 Analyse Coûts", 
        "✅ Validation Données"
    ])
    
    with tab1:
        st.subheader("Rapport Analytique Complet")
        if st.button("🔄 Générer le rapport analytique"):
            with st.spinner("Génération du rapport en cours..."):
                rapport = report_generator.generer_rapport_analytique()
                st.text_area("Rapport détaillé", rapport, height=400)
    
    with tab2:
        st.subheader("Rapport Spécifique Client")
        clients_disponibles = sorted(st.session_state.df_livraisons_original["Client de l'estafette"].unique())
        client_rapport = st.selectbox("Sélectionner un client", clients_disponibles)
        
        if st.button("📋 Générer rapport client"):
            with st.spinner("Génération du rapport client..."):
                rapport_client = report_generator.generer_rapport_client(client_rapport)
                st.text_area(f"Rapport pour {client_rapport}", rapport_client, height=300)
    
    with tab3:
        st.subheader("Analyse des Coûts")
        col_cost1, col_cost2 = st.columns(2)
        
        with col_cost1:
            cout_estafette = st.number_input("Coût unitaire estafette (TND)", value=150, min_value=50, max_value=500)
        with col_cost2:
            cout_camion = st.number_input("Coût unitaire camion (TND)", value=800, min_value=300, max_value=2000)
        
        if st.button("💰 Calculer les coûts"):
            from backend import calculer_couts_estimation
            couts = calculer_couts_estimation(
                st.session_state.df_voyages, 
                cout_estafette, 
                cout_camion
            )
            
            if 'erreur' not in couts:
                st.success(couts['cout_estimation'])
                
                # Graphique des coûts
                import plotly.express as px
                df_couts = pd.DataFrame({
                    'Type': ['Estafettes', 'Camions'],
                    'Coût Total (TND)': [
                        couts['estafettes'] * couts['cout_estafette_unitaire'],
                        couts['camions'] * couts['cout_camion_unitaire']
                    ]
                })
                
                fig = px.pie(df_couts, values='Coût Total (TND)', names='Type', 
                            title='Répartition des coûts par type de véhicule')
                st.plotly_chart(fig, use_container_width=True)
            else:
                st.error(couts['erreur'])
    
    with tab4:
        st.subheader("Validation d'Intégrité des Données")
        if st.button("🔍 Vérifier l'intégrité des données"):
            from backend import verifier_integrite_donnees
            resultat_validation = verifier_integrite_donnees(
                st.session_state.df_voyages,
                st.session_state.df_livraisons_original
            )
            
            if "✅" in resultat_validation:
                st.success(resultat_validation)
            else:
                st.warning(resultat_validation)

else:
    st.warning("⚠️ Vous devez d'abord traiter les données .")

st.markdown("---")

# =====================================================
# 🔟 GÉNÉRATION DES CODES VOYAGE
# =====================================================
st.markdown("## 🏷️ GÉNÉRATION DES CODES VOYAGE")

if "df_voyages_valides" in st.session_state and not st.session_state.df_voyages_valides.empty:
    
    df_final = st.session_state.df_voyages_valides.copy()
    
    #st.info("""
    #**Génération automatique des codes voyage uniques pour chaque mission.**
    #Le format : **Véhicule/Date/NuméroSéquentiel**
    #""")
    
    # Configuration des paramètres de génération
    col1, col2, col3 = st.columns(3)
    
    with col1:
        date_voyage = st.date_input(
            "📅 Date de voyage",
            value=pd.Timestamp.now().date(),
            help="Date prévue pour les livraisons"
        )
    
    with col2:
        numero_debut = st.number_input(
            "🔢 Numéro séquentiel de départ",
            min_value=1,
            max_value=1000,
            value=1,
            help="Numéro de départ pour la séquence"
        )
    
    with col3:
        st.markdown("<br>", unsafe_allow_html=True)
        generer_codes = st.button("🏷️ Générer les codes voyage", type="primary")
    
    if generer_codes:
        try:
            # Préparation des données pour le code voyage
            df_final['Date Voyage Format'] = date_voyage.strftime('%Y%m%d')
            
            # Création du numéro séquentiel pour chaque voyage
            df_final['Numero Séquentiel'] = range(numero_debut, numero_debut + len(df_final))
            df_final['Numero Séquentiel Formatted'] = df_final['Numero Séquentiel'].apply(lambda x: f"{x:03d}")
            
            # Création du Code voyage
            df_final['Code voyage'] = (
                df_final['Véhicule N°'].astype(str) + '/' +
                df_final['Date Voyage Format'].astype(str) + '/' +
                df_final['Numero Séquentiel Formatted'].astype(str)
            )
            
            # Mettre à jour le session state
            st.session_state.df_voyages_valides = df_final
            
            st.success(f"✅ {len(df_final)} codes voyage générés avec succès !")
            
            # Afficher un aperçu des codes générés
            st.markdown("### 📋 Aperçu des codes voyage générés")
            df_apercu = df_final[['Véhicule N°', 'Zone', 'Code voyage']].copy()
            show_df(df_apercu, use_container_width=True)
            
            print("✅ Colonne 'Code voyage' créée avec succès.")
            
        except Exception as e:
            st.error(f"❌ Erreur lors de la génération des codes voyage : {str(e)}")
    
    # Afficher les codes existants si déjà générés
    elif 'Code voyage' in df_final.columns:
        st.success("✅ Codes voyage déjà générés")
        df_apercu = df_final[['Véhicule N°', 'Zone', 'Code voyage']].copy()
        show_df(df_apercu, use_container_width=True)
        
        # Option pour regénérer les codes
        if st.button("🔄 Regénérer les codes voyage"):
            del df_final['Code voyage']
            del df_final['Date Voyage Format']
            del df_final['Numero Séquentiel']
            del df_final['Numero Séquentiel Formatted']
            st.session_state.df_voyages_valides = df_final
            st.rerun()

else:
    st.warning("⚠️ Vous devez d'abord valider les voyages.")

# =====================================================
# 📤 EXPORT FINAL ET PLANNING COMPLET - VERSION OPTIMISÉE
# =====================================================
st.markdown("## 📤 EXPORT FINAL ET PLANNING COMPLET")

if "df_voyages_valides" in st.session_state and not st.session_state.df_voyages_valides.empty:
    
    df_export_final = st.session_state.df_voyages_valides.copy()
    
    # =====================================================
    # GARANTIR QUE TOUTES LES COLONNES REQUISES EXISTENT
    # =====================================================
    
    # Vérifier et créer la colonne "Chauffeur" si nécessaire
    if "Chauffeur" not in df_export_final.columns:
        # Priorité 1 : Utiliser "Chauffeur attribué"
        if "Chauffeur attribué" in df_export_final.columns:
            df_export_final["Chauffeur"] = df_export_final["Chauffeur attribué"]
            st.success("✅ Colonne 'Chauffeur' créée à partir de 'Chauffeur attribué'")
        # Priorité 2 : Utiliser "Matricule chauffeur" avec format
        elif "Matricule chauffeur" in df_export_final.columns:
            df_export_final["Chauffeur"] = df_export_final["Matricule chauffeur"].apply(
                lambda x: f"Chauffeur {x}" if pd.notna(x) and x != "" else "À attribuer"
            )
        # Fallback
        else:
            df_export_final["Chauffeur"] = "À attribuer"
            st.warning("⚠️ Colonne 'Chauffeur' créée vide")
    
    # Vérifier que "Code voyage" existe
    if "Code voyage" not in df_export_final.columns:
        st.error("❌ La colonne 'Code voyage' est manquante. Veuillez d'abord générer les codes voyage dans la section 10.")
        st.stop()
    
    # =====================================================
    # FONCTION POUR FORMATER LES COLONNES AVEC RETOURS À LA LIGNE
    # =====================================================
    def formater_colonnes_retours_ligne(df):
        df_formate = df.copy()
        colonnes_a_formater = ['BL inclus', 'Client(s) inclus', 'Représentant(s) inclus']
        
        for col in colonnes_a_formater:
            if col in df_formate.columns:
                df_formate[col] = df_formate[col].apply(
                    lambda x: '\n'.join([elem.strip() for elem in str(x).replace(';', ',').split(',') if elem.strip()]) 
                    if pd.notna(x) else ""
                )
        return df_formate
    
    # =====================================================
    # AFFICHAGE DÉTAILLÉ AVEC RETOURS À LA LIGNE
    # =====================================================
    st.markdown("### 📊 Planning de Livraisons Détaillé")
    
    # Appliquer le formatage pour l'affichage Streamlit
    df_affichage_formate = formater_colonnes_retours_ligne(df_export_final)
    
    # Afficher chaque voyage avec expanders détaillés
    for idx, row in df_affichage_formate.iterrows():
        with st.expander(f"🚚 Voyage {row['Véhicule N°']} | Zone : {row['Zone']} | Véhicule: {row.get('Véhicule attribué', 'N/A')} | Chauffeur: {row.get('Chauffeur', 'N/A')}"):
            col1, col2, col3 = st.columns(3)
            
            with col1:
                st.write("**Informations de base:**")
                st.write(f"**Code voyage:** {row['Code voyage']}")
                st.write(f"**Zone:** {row['Zone']}")
                st.write(f"**Véhicule N°:** {row['Véhicule N°']}")
                if "Poids total chargé" in row:
                    st.write(f"**Poids total chargé:** {row['Poids total chargé']:.3f} kg")
                if "Volume total chargé" in row:
                    st.write(f"**Volume total chargé:** {row['Volume total chargé']:.3f} m³")
                if "Taux d'occupation (%)" in row:
                    st.write(f"**Taux d'occupation:** {row['Taux d\'occupation (%)']:.3f}%")
                if "Véhicule attribué" in row:
                    st.write(f"**Véhicule attribué:** {row['Véhicule attribué']}")
                if "Chauffeur" in row:
                    st.write(f"**Chauffeur:** {row['Chauffeur']}")
            
            with col2:
                # Afficher les clients avec retours à ligne
                if 'Client(s) inclus' in row and pd.notna(row['Client(s) inclus']):
                    st.write("**📋 Clients inclus:**")
                    clients = str(row['Client(s) inclus']).split('\n')
                    for client in clients:
                        client_clean = client.strip()
                        if client_clean:
                            st.write(f"• {client_clean}")
                
                # Afficher les représentants avec retours à ligne
                if 'Représentant(s) inclus' in row and pd.notna(row['Représentant(s) inclus']):
                    st.write("**👤 Représentants inclus:**")
                    representants = str(row['Représentant(s) inclus']).split('\n')
                    for rep in representants:
                        rep_clean = rep.strip()
                        if rep_clean:
                            st.write(f"• {rep_clean}")
            
            with col3:
                # Afficher les BL avec retours à ligne
                if 'BL inclus' in row and pd.notna(row['BL inclus']):
                    st.write("**📄 BL associés:**")
                    bls = str(row['BL inclus']).split('\n')
                    # Afficher en colonnes si beaucoup de BL
                    if len(bls) > 5:
                        cols = st.columns(2)
                        half = len(bls) // 2
                        for i, bl in enumerate(bls):
                            bl_clean = bl.strip()
                            if bl_clean:
                                col_idx = 0 if i < half else 1
                                with cols[col_idx]:
                                    st.write(f"• {bl_clean}")
                    else:
                        for bl in bls:
                            bl_clean = bl.strip()
                            if bl_clean:
                                st.write(f"• {bl_clean}")

    # =====================================================
    # EXPORT EXCEL AVEC RETOURS À LA LIGNE
    # =====================================================
    col_export1, col_export2 = st.columns(2)
    
    with col_export1:
        nom_fichier = st.text_input(
            "📝 Nom du fichier d'export", 
            value=f"Planning_Livraisons_{pd.Timestamp.now().strftime('%Y%m%d_%H%M')}",
            help="Le fichier sera sauvegardé avec l'extension .xlsx"
        )
    
    with col_export2:
        st.markdown("<br>", unsafe_allow_html=True)
        if st.button("🚀 Générer l'export complet", type="primary"):
            try:
                from backend import exporter_planning_excel
                
                # Préparer les données supplémentaires
                donnees_supplementaires = {}
                
                # Ajouter les données de base si disponibles
                if st.session_state.df_grouped is not None:
                    donnees_supplementaires['Livraisons_Client_Ville'] = st.session_state.df_grouped
                if st.session_state.df_city is not None:
                    donnees_supplementaires['Besoin_Estafette_Ville'] = st.session_state.df_city
                if st.session_state.df_zone is not None:
                    donnees_supplementaires['Besoin_Estafette_Zone'] = st.session_state.df_zone
                
                # Appliquer le formatage avec retours à ligne avant l'export
                df_export_formate = formater_colonnes_retours_ligne(df_export_final)
                
                # Générer l'export
                success, message = exporter_planning_excel(
                    df_export_formate,  # Utiliser le DataFrame formaté avec retours à ligne
                    f"{nom_fichier}.xlsx",
                    donnees_supplementaires,
                    st.session_state.df_livraisons_original
                )
                                
                if success:
                    st.success(message)
                    
                    # Aperçu du format d'export
                    st.subheader("👁️ Aperçu du format d'export")
                    colonnes_apercu = ["Code voyage", "Zone", "Véhicule N°", "Chauffeur", "BL inclus", "Client(s) inclus", "Poids total chargé", "Volume total chargé"]
                    colonnes_apercu = [col for col in colonnes_apercu if col in df_export_formate.columns]
                    
                    df_apercu = df_export_formate[colonnes_apercu].head(5).copy()
                    
                    # Formater l'affichage
                    if "Poids total chargé" in df_apercu.columns:
                        df_apercu["Poids total chargé"] = df_apercu["Poids total chargé"].map(lambda x: f"{x:.1f} kg")
                    if "Volume total chargé" in df_apercu.columns:
                        df_apercu["Volume total chargé"] = df_apercu["Volume total chargé"].map(lambda x: f"{x:.3f} m³")
                    
                    show_df(df_apercu, use_container_width=True)
                    
                    # Proposer le téléchargement
                    with open(f"{nom_fichier}.xlsx", "rb") as file:
                        btn = st.download_button(
                            label="💾 Télécharger le planning complet",
                            data=file,
                            file_name=f"{nom_fichier}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )
                else:
                    st.error(message)
                    
            except Exception as e:
                st.error(f"❌ Erreur lors de l'export : {str(e)}")

    # =====================================================
    # APERÇU DU PLANNING FINAL (TABLEAU SIMPLE)
    # =====================================================
    st.markdown("### 👁️ Aperçu du Planning Final (Vue Tableau)")
    
    df_apercu_final = df_export_final.copy()
    
    # Appliquer le formatage pour l'aperçu
    df_apercu_final = formater_colonnes_retours_ligne(df_apercu_final)
    
    # Colonnes à afficher (format d'export final)
    colonnes_apercu = ["Code voyage", "Zone", "Véhicule N°", "Chauffeur", "BL inclus", "Client(s) inclus", "Poids total chargé", "Volume total chargé"]
    colonnes_apercu = [col for col in colonnes_apercu if col in df_apercu_final.columns]
    
    # Formater l'affichage pour l'aperçu
    if "Poids total chargé" in df_apercu_final.columns:
        df_apercu_final["Poids total chargé"] = df_apercu_final["Poids total chargé"].map(lambda x: f"{x:.1f} kg")
    if "Volume total chargé" in df_apercu_final.columns:
        df_apercu_final["Volume total chargé"] = df_apercu_final["Volume total chargé"].map(lambda x: f"{x:.3f} m³")
    
    show_df(df_apercu_final[colonnes_apercu], use_container_width=True)

else:
    st.warning("⚠️ Vous devez d'abord valider les voyages et générer les codes voyage.")

# =====================================================
# 🎯 RÉSUMÉ ET TABLEAU DE BORD FINAL
# =====================================================
st.markdown("## 🎯 RÉSUMÉ DU PLANNING")

if "df_voyages" in st.session_state:
    df_final = st.session_state.df_voyages.copy()
    
    # Calcul des métriques principales
    total_vehicules = len(df_final)
    estafettes = len(df_final[df_final["Code Véhicule"] == "ESTAFETTE"])
    camions = len(df_final[df_final["Code Véhicule"] == "CAMION-LOUE"])
    poids_total = df_final["Poids total chargé"].sum()
    volume_total = df_final["Volume total chargé"].sum()
    taux_moyen = df_final["Taux d'occupation (%)"].mean() if "Taux d'occupation (%)" in df_final.columns else 0
    
    # Affichage des métriques
    col_metric1, col_metric2, col_metric3, col_metric4 = st.columns(4)
    
    with col_metric1:
        st.metric("🚚 Total Véhicules", total_vehicules)
    
    with col_metric2:
        st.metric("🚐 Estafettes", estafettes)
    
    with col_metric3:
        st.metric("🚛 Camions", camions)
    
    with col_metric4:
        st.metric("⚖️ Poids Total", f"{poids_total:.0f} kg")
    
    col_metric5, col_metric6, col_metric7, col_metric8 = st.columns(4)
    
    with col_metric5:
        st.metric("📦 Volume Total", f"{volume_total:.1f} m³")
    
    with col_metric6:
        st.metric("📊 Taux Occupation Moyen", f"{taux_moyen:.1f}%")
    
    with col_metric7:
        # Calcul de l'efficacité
        efficacite = " Bonne" if taux_moyen > 70 else " Moyenne" if taux_moyen > 50 else " Faible"
        st.metric("🎯 Efficacité", efficacite)
    
    with col_metric8:
        # Statut de complétion
        status = "✅ Complet" if 'df_voyages_valides' in st.session_state else "🟡 En cours"
        st.metric("📋 Statut", status)
    
    # Graphique de répartition par zone
    # Dans la section "RÉSUMÉ ET TABLEAU DE BORD FINAL", remplacez le code problématique par :

    st.subheader("📊 Répartition par Zone")
    if 'Zone' in df_final.columns:
        repartition_zone = df_final.groupby("Zone").size().reset_index(name="Nombre de véhicules")
        
        if not repartition_zone.empty:
            import plotly.express as px

            fig_zone = px.bar(
                repartition_zone, 
                x="Zone", 
                y="Nombre de véhicules",
                title="Nombre de véhicules par zone",
                color="Nombre de véhicules",
                color_continuous_scale=[
                    "#ADE8F4",   # Bleu clair visible
                    "#90E0EF",
                    "#4EA8DE",
                    "#3A86FF",
                    "#1E6091",
                    "#0A3D62" 
                ],
                text="Nombre de véhicules"
            )

            fig_zone.update_layout(coloraxis_colorbar=dict(title="Nb Véhicules"))
            st.plotly_chart(fig_zone, use_container_width=True)

    # Graphique de répartition par type de véhicule
    st.subheader("🚗 Répartition par Type de Véhicule")
    if "Code Véhicule" in df_final.columns:
        repartition_type = df_final["Code Véhicule"].value_counts().reset_index()
        repartition_type.columns = ["Type Véhicule", "Nombre"]
        
        fig_type = px.pie(
            repartition_type, 
            values="Nombre", 
            names="Type Véhicule",
            title="Répartition des types de véhicules"
        )
        st.plotly_chart(fig_type, use_container_width=True)

else:
    st.warning("⚠️ Le planning n'est pas encore généré.")


# =====================================================
# 🏁 PIED DE PAGE
# =====================================================
st.markdown("---")
st.markdown(
    """
    <div style='text-align: center; color: #666;'>
        <p>🚚 <strong>Système d'Optimisation des Livraisons</strong> - Développé par Zaineb KCHAOU</p>
        <p>📧 Support : Zaineb.KCHAOU@sopal.com | 📞 Hotline : +216 23 130 088</p>
    </div>
    """,
    unsafe_allow_html=True
)
# =====================================================
# 📱 STYLE RESPONSIVE ET AMÉLIORATIONS VISUELLES
# =====================================================
st.markdown("""
<style>
    .main-header {
        font-size: 2.5rem;
        color: #1f77b4;
        text-align: center;
        margin-bottom: 2rem;
    }
    .metric-card {
        background-color: #f0f2f6;
        padding: 1rem;
        border-radius: 10px;
        border-left: 5px solid #1f77b4;
    }
    
    .success-box {
        background-color: #d4edda;
        border: 1px solid #c3e6cb;
        border-radius: 5px;
        padding: 1rem;
        margin: 1rem 0;
    }
    .warning-box {
        background-color: #fff3cd;
        border: 1px solid #ffeaa7;
        border-radius: 5px;
        padding: 1rem;
        margin: 1rem 0;
    }
    .info-box {
        background-color: #d1ecf1;
        border: 1px solid #bee5eb;
        border-radius: 5px;
        padding: 1rem;
        margin: 1rem 0;
    }
</style>
""", unsafe_allow_html=True)