import streamlit as st
import pandas as pd
import plotly.express as px
from io import BytesIO
from backend import DeliveryProcessor, TruckRentalProcessor, TruckTransferManager, SEUIL_POIDS, SEUIL_VOLUME
import openpyxl
from openpyxl.styles import Alignment
# =====================================================
# CONFIGURATION DE LA PAGE ET CSS GLOBAL
# =====================================================
st.set_page_config(
    page_title="🚚 Planning de Livraisons",
    page_icon="🚚",
    layout="wide",
    initial_sidebar_state="expanded"
)

# CSS global pour toute l'application
st.markdown("""
<style>
    /* Navigation sidebar */
    [data-testid="stSidebarNav"] {
        padding-top: 20px;
    }
    
    /* Style des liens de navigation */
    [data-testid="stSidebarNav"] a {
        padding: 10px 15px;
        margin: 5px 0;
        border-radius: 8px;
        display: flex;
        align-items: center;
        gap: 10px;
        font-weight: 500;
        transition: all 0.3s ease;
    }
    
    [data-testid="stSidebarNav"] a:hover {
        background-color: #E6F3FF;
        color: #0369A1;
    }
    
    /* En-tête principal */
    .main-header {
        font-size: 2.5rem;
        color: #1E3A8A;
        text-align: center;
        margin-bottom: 1rem;
        padding-bottom: 1rem;
        border-bottom: 3px solid #1E3A8A;
    }
    
    /* Cartes de métriques */
    .metric-card {
        background: linear-gradient(135deg, #F8FAFC 0%, #EFF6FF 100%);
        padding: 1.5rem;
        border-radius: 10px;
        border-left: 5px solid #0369A1;
        box-shadow: 0 2px 4px rgba(0,0,0,0.1);
    }
    
    /* Tables */
    .custom-table {
        width: 100%;
        border-collapse: collapse;
        font-family: Arial, sans-serif;
        font-size: 14px;
        box-shadow: 0 2px 8px rgba(0,0,0,0.1);
        border-radius: 8px;
        overflow: hidden;
    }
    
    .custom-table th {
        background-color: #0369A1;
        color: white;
        padding: 12px 8px;
        text-align: center;
        border: 2px solid #4682B4;
        font-weight: normal;
        font-size: 13px;
        vertical-align: middle;
    }
    
    .custom-table td {
        padding: 10px 8px;
        text-align: center;
        border: 1px solid #B0C4DE;
        background-color: white;
        color: #000000;
        vertical-align: middle;
        font-weight: normal;
    }
    
    /* Onglets */
    .stTabs [data-baseweb="tab-list"] {
        gap: 8px;
    }
    
    .stTabs [data-baseweb="tab"] {
        height: 50px;
        white-space: pre-wrap;
        background-color: #F0F2F6;
        border-radius: 8px 8px 0px 0px;
        gap: 8px;
        padding: 10px 16px;
        font-weight: 600;
    }
    
    .stTabs [aria-selected="true"] {
        background-color: #0369A1 !important;
        color: white !important;
    }
</style>
""", unsafe_allow_html=True)

# =====================================================
# FONCTIONS UTILITAIRES COMMUNES
# =====================================================
def show_df(df, **kwargs):
    """Affiche un DataFrame avec arrondi à 3 décimales."""
    if isinstance(df, pd.DataFrame):
        df_to_display = df.copy()
        df_to_display = df_to_display.round(3)
        st.dataframe(df_to_display, **kwargs)
    else:
        st.dataframe(df, **kwargs)

def to_excel(df, sheet_name="Données"):
    """Export DataFrame to Excel."""
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name=sheet_name)
    return output.getvalue()

# =====================================================
# INITIALISATION SESSION STATE
# =====================================================
if 'data_processed' not in st.session_state:
    st.session_state.data_processed = False
    st.session_state.df_grouped = None
    st.session_state.df_city = None
    st.session_state.df_grouped_zone = None
    st.session_state.df_zone = None
    st.session_state.df_optimized_estafettes = None
    st.session_state.df_livraisons_original = None
    st.session_state.rental_processor = None
    st.session_state.propositions = None
    st.session_state.selected_client = None
    st.session_state.message = ""
    st.session_state.df_voyages = None
    st.session_state.df_livraisons = None
    st.session_state.df_voyages_valides = None
    st.session_state.transfer_manager = None
    st.session_state.attributions = {}
    st.session_state.validations = {}
# =====================================================
# 📌 Constantes pour les véhicules et chauffeurs
# =====================================================
VEHICULES_DISPONIBLES = [
    'SLG-VEH11', 'SLG-VEH14', 'SLG-VEH22', 'SLG-VEH19',
    'SLG-VEH10', 'SLG-VEH16', 'SLG-VEH23', 'SLG-VEH08', 'SLG-VEH20', 'code-Camion'
]

CHAUFFEURS_DETAILS = {
    '09254': 'DAMMAK Karim', '06002': 'MAAZOUN Bassem', '11063': 'SASSI Ramzi',
    '10334': 'BOUJELBENE Mohamed', '15144': 'GADDOUR Rami', '08278': 'DAMMAK Wissem',
    '18339': 'REKIK Ahmed', '07250': 'BARKIA Mustapha', '13321': 'BADRI Moez','99999': 'Chauffeur Camion'
}
# =====================================================
# Fonctions de Callback pour la Location
# =====================================================

def update_propositions_view():
    """Met à jour le DataFrame de propositions après une action."""
    if st.session_state.rental_processor:
        st.session_state.propositions = st.session_state.rental_processor.detecter_propositions()
        
        # CORRECTION : Vérifier si le DataFrame de propositions n'est pas vide et contient la colonne 'Client'
        if (st.session_state.propositions is not None and 
            not st.session_state.propositions.empty and 
            'Client' in st.session_state.propositions.columns):
            
            # Réinitialiser la sélection si le client n'est plus dans les propositions ouvertes
            if (st.session_state.selected_client is not None and 
                st.session_state.selected_client not in st.session_state.propositions['Client'].astype(str).tolist()):
                st.session_state.selected_client = None
    else:
        st.session_state.propositions = pd.DataFrame()

def handle_location_action(accepter):
    """Gère l'acceptation ou le refus de la proposition de location."""
    if st.session_state.rental_processor and st.session_state.selected_client:
        try:
            # Assurer que le client est une chaîne valide
            client_to_process = str(st.session_state.selected_client)
            ok, msg, _ = st.session_state.rental_processor.appliquer_location(
                client_to_process, accepter=accepter
            )
            st.session_state.message = msg
            update_propositions_view()
        except Exception as e:
            st.session_state.message = f"❌ Erreur lors du traitement : {str(e)}"
    elif not st.session_state.selected_client:
        st.session_state.message = "⚠️ Veuillez sélectionner un client à traiter."
    else:
        st.session_state.message = "⚠️ Le processeur de location n'est pas initialisé."

def accept_location_callback():
    handle_location_action(True)

def refuse_location_callback():
    handle_location_action(False)

# =====================================================
# PAGE 1: Importation des données 
# =====================================================
def page_import():
    st.markdown("<h1 class='main-header'>1. 📥 Importation des données</h1>", unsafe_allow_html=True)
    
    st.markdown("""
    <div style='background: linear-gradient(135deg, #1E3A8A 0%, #3B82F6 100%);
                padding: 2rem;
                border-radius: 10px;
                color: white;
                margin-bottom: 2rem;
                text-align: center;
                box-shadow: 0 4px 12px rgba(0, 0, 0, 0.1);'>
        <h3 style='color: white; margin-bottom: 1rem;'>📋 Instructions d'importation</h3>
        <p style='margin-bottom: 0; font-size: 16px;'>Téléchargez les 3 fichiers requis pour commencer l'analyse des livraisons.</p>
    </div>
    """, unsafe_allow_html=True)
    
    col1, col2, col3 = st.columns(3)
    
    with col1:
        st.markdown("""
        <div class='metric-card'>
            <h4 style='color: #0369A1; margin-bottom: 10px;'>📄 Fichier Livraisons</h4>
            <p style='font-size: 14px; color: #666;'>Format Excel (.xlsx)</p>
        </div>
        """, unsafe_allow_html=True)
        liv_file = st.file_uploader("BL", type=["xlsx"], key="file1")
        
    with col2:
        st.markdown("""
        <div class='metric-card'>
            <h4 style='color: #0369A1; margin-bottom: 10px;'>📦 Fichier Volumes</h4>
            <p style='font-size: 14px; color: #666;'>Format Excel (.xlsx)</p>
        </div>
        """, unsafe_allow_html=True)
        ydlogist_file = st.file_uploader("Articles", type=["xlsx"], key="file2")
        
    with col3:
        st.markdown("""
        <div class='metric-card'>
            <h4 style='color: #0369A1; margin-bottom: 10px;'>🏢 Fichier Clients</h4>
            <p style='font-size: 14px; color: #666;'>Format Excel (.xlsx)</p>
        </div>
        """, unsafe_allow_html=True)
        wcliegps_file = st.file_uploader("Clients/Zones", type=["xlsx"], key="file3")
    
    st.markdown("<br>", unsafe_allow_html=True)
    
    # Bouton de traitement
    col_btn1, col_btn2, col_btn3 = st.columns([1, 2, 1])
    with col_btn2:
        if st.button("🚀 Exécuter le traitement complet", type="primary", use_container_width=True):
            if liv_file and ydlogist_file and wcliegps_file:
                processor = DeliveryProcessor()
                try:
                    with st.spinner("🔍 Traitement des données en cours..."):
                        df_grouped, df_city, df_grouped_zone, df_zone, df_optimized_estafettes, df_livraisons_original = processor.process_delivery_data(
                            liv_file, ydlogist_file, wcliegps_file
                        )
                    
                    # Stockage dans session_state
                    st.session_state.df_grouped = df_grouped
                    st.session_state.df_city = df_city
                    st.session_state.df_grouped_zone = df_grouped_zone
                    st.session_state.df_zone = df_zone
                    st.session_state.df_optimized_estafettes = df_optimized_estafettes
                    st.session_state.df_livraisons_original = df_livraisons_original
                    st.session_state.df_livraisons = df_grouped_zone
                    
                    # Initialisation des processeurs
                    st.session_state.rental_processor = TruckRentalProcessor(
                        df_optimized_estafettes, df_livraisons_original
                    )
                    st.session_state.data_processed = True
                    
                    st.markdown("""
                <div style="text-align: center; padding: 10px; margin-top: 20px;">
                    <p style="color: green; font-size: 16px; font-weight: bold;">✅ Données importées et traitées avec succès !</p>
                </div>
                """, unsafe_allow_html=True)
                
                    
                except Exception as e:
                    st.error(f"❌ Erreur lors du traitement : {str(e)}")
            else:
                st.warning("⚠️ Veuillez télécharger les 3 fichiers requis.")
    
    # Afficher les résultats si disponibles
    if st.session_state.data_processed:
        st.markdown("---")
        st.subheader("📊 Aperçu des données importées")
        
        col_info1, col_info2, col_info3 = st.columns(3)
        with col_info1:
            if st.session_state.df_grouped is not None:
                st.metric("📦 Livraisons", f"{len(st.session_state.df_grouped)}")
        with col_info2:
            if st.session_state.df_city is not None:
                villes = st.session_state.df_city['Ville'].nunique()
                st.metric("🏙️ Villes", f"{villes}")
        with col_info3:
            if st.session_state.df_zone is not None:
                zones = st.session_state.df_zone['Zone'].nunique()
                st.metric("🌍 Zones", f"{zones}")
        
        # Boutons de navigation
        st.markdown("---")
        col_nav1, col_nav2 = st.columns(2)
        with col_nav1:
            if st.button("📋 Voir l'analyse détaillée →", use_container_width=True):
                st.session_state.page = "analyse"
                st.rerun()
        with col_nav2:
            if st.button("🚚 Passer à l'optimisation →", use_container_width=True, type="secondary"):
                st.session_state.page = "optimisation"
                st.rerun()

# =====================================================
# PAGE 2: Analyse détaillée 
# =====================================================
def page_analyse():
    st.markdown("<h1 class='main-header'>2. 🔍 Analyse détaillée</h1>", unsafe_allow_html=True)
    
    if not st.session_state.data_processed:
        st.warning("⚠️ Veuillez d'abord importer les données dans la page 1.")
        if st.button("📥 Retour à l'importation"):
            st.session_state.page = "import"
            st.rerun()
        return
    
    # CSS PERSONNALISÉ POUR LES ONGLETS
    st.markdown("""
    <style>
        /* Style pour les onglets - COULEUR BLEUE */
        .stTabs [data-baseweb="tab-list"] {
            gap: 8px;
        }
        
        .stTabs [data-baseweb="tab"] {
            height: 50px;
            white-space: pre-wrap;
            background-color: #F0F2F6;
            border-radius: 8px 8px 0px 0px;
            gap: 8px;
            padding: 10px 16px;
            font-weight: 600;
        }
        
        .stTabs [data-baseweb="tab"]:hover {
            background-color: #E6F3FF;
            color: #0369A1;
        }
        
        /* ONGLET ACTIF - BLEU ROYAL */
        .stTabs [aria-selected="true"] {
            background-color: #0369A1 !important;
            color: white !important;
        }
        
        /* TEXTE DES ONGLETS */
        .stTabs [data-baseweb="tab"] p {
            font-size: 16px;
            font-weight: 600;
            margin: 0;
        }
        
        /* COULEUR DU TEXTE POUR ONGLET ACTIF */
        .stTabs [aria-selected="true"] p {
            color: white !important;
        }
        
        /* Style général du tableau */
        .custom-table {
            width: 100%;
            border-collapse: collapse;
            font-family: Arial, sans-serif;
            font-size: 14px;
            box-shadow: 0 2px 8px rgba(0,0,0,0.1);
            border-radius: 8px;
            overflow: hidden;
        }
        
        /* En-têtes du tableau - BLEU ROYAL SANS DÉGRADÉ */
        .custom-table th {
            background-color: #0369A1;
            color: white;
            padding: 12px 8px;
            text-align: center;
            border: 2px solid #4682B4;
            font-weight: normal;
            font-size: 13px;
            vertical-align: middle;
        }
        
        /* Cellules du tableau - TOUTES EN BLANC */
        .custom-table td {
            padding: 10px 8px;
            text-align: center;
            border: 1px solid #B0C4DE;
            background-color: white;
            color: #000000;
            vertical-align: middle;
            font-weight: normal;
        }
        
        /* Bordures visibles pour toutes les cellules */
        .custom-table th, 
        .custom-table td {
            border: 1px solid #B0C4DE !important;
        }
        
        /* Bordures épaisses pour l'extérieur du tableau */
        .custom-table {
            border: 2px solid #4682B4 !important;
        }
        
        /* Conteneur du tableau avec défilement horizontal */
        .table-container {
            overflow-x: auto;
            margin: 1rem 0;
            border-radius: 8px;
            border: 2px solid #4682B4;
        }
        
        /* Supprimer l'alternance des couleurs - TOUTES LES LIGNES BLANCHES */
        .custom-table tr:nth-child(even) td {
            background-color: white !important;
        }
        
        /* Survol des lignes - léger effet */
        .custom-table tr:hover td {
            background-color: #F0F8FF !important;
        }
    </style>
    """, unsafe_allow_html=True)
    
    # Onglets pour différents types d'analyse
    tab_grouped, tab_city, tab_zone_group, tab_zone_summary, tab_charts = st.tabs([
        "Livraisons Client/Ville", 
        "Besoin Estafette par Ville", 
        "Livraisons Client/Zone", 
        "Besoin Estafette par Zone",
        "Graphiques"
    ])
    
    # --- Onglet Livraisons Client/Ville ---
    with tab_grouped:
        st.subheader("Livraisons par Client & Ville")
        
        # Créer une copie du DataFrame et FILTRER TRIPOLI
        df_liv = st.session_state.df_grouped.drop(columns=["Zone"], errors='ignore').copy()
        df_liv = df_liv[df_liv["Ville"] != "TRIPOLI"]  # ← FILTRE TRIPOLI
        
        # Vérifier si le DataFrame n'est pas vide après filtrage
        if df_liv.empty:
            st.info("ℹ️ Aucune livraison à afficher (TRIPOLI exclue)")
        else:
            # Préparer les données pour l'affichage HTML
            if "Article" in df_liv.columns:
                # Transformer les articles avec retours à la ligne HTML - SANS "•"
                df_liv["Article"] = df_liv["Article"].astype(str).apply(
                    lambda x: "<br>".join(a.strip() for a in x.split(",") if a.strip())
                )
            
            # Formater les nombres - 3 chiffres après la virgule
            if "Poids total" in df_liv.columns:
                df_liv["Poids total"] = df_liv["Poids total"].map(lambda x: f"{x:.3f} kg" if pd.notna(x) else "")
            if "Volume total" in df_liv.columns:
                df_liv["Volume total"] = df_liv["Volume total"].map(lambda x: f"{x:.3f} m³" if pd.notna(x) else "")
            
            # Afficher le tableau avec le style CSS
            html_table = df_liv.to_html(
                escape=False, 
                index=False, 
                classes="custom-table",
                border=0
            )
            
            st.markdown(f"""
            <div class="table-container">
                {html_table}
            </div>
            """, unsafe_allow_html=True)
        
        # Métriques résumées
        st.markdown("---")
        col1, col2, col3, col4 = st.columns(4)
        
        with col1:
            total_livraisons = len(df_liv) if not df_liv.empty else 0
            st.metric("📦 Total Livraisons", total_livraisons)
        
        with col2:
            total_clients = df_liv["Client"].nunique() if not df_liv.empty else 0
            st.metric("👥 Clients Uniques", total_clients)
        
        with col3:
            # Calculer le poids total à partir des données filtrées
            df_liv_original = st.session_state.df_grouped[st.session_state.df_grouped["Ville"] != "TRIPOLI"]
            total_poids = df_liv_original["Poids total"].sum() if not df_liv_original.empty else 0
            st.metric("⚖️ Poids Total", f"{total_poids:.3f} kg")
        
        with col4:
            # Calculer le volume total à partir des données filtrées
            total_volume = df_liv_original["Volume total"].sum() if not df_liv_original.empty else 0
            st.metric("📏 Volume Total", f"{total_volume:.3f} m³")
        
        # Bouton de téléchargement
        from io import BytesIO
        excel_buffer_grouped = BytesIO()
        with pd.ExcelWriter(excel_buffer_grouped, engine='openpyxl') as writer:
            st.session_state.df_grouped.drop(columns=["Zone"], errors='ignore').to_excel(writer, index=False, sheet_name="Livraisons Client Ville")
        excel_buffer_grouped.seek(0)
        
        st.download_button(
            label="💾 Télécharger Livraisons Client/Ville",
            data=excel_buffer_grouped,
            file_name="Livraisons_Client_Ville.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        
        # Stockage pour la section transfert
        if "df_livraisons" not in st.session_state:
            st.session_state.df_livraisons = df_liv.copy()
    
    # --- Onglet Besoin Estafette par Ville ---
    with tab_city:
        st.subheader("Besoin Estafette par Ville")
        
        # Créer une copie du DataFrame et FILTRER TRIPOLI
        df_city_display = st.session_state.df_city.copy()
        df_city_display = df_city_display[df_city_display["Ville"] != "TRIPOLI"]
        
        # Formater les nombres - 3 chiffres après la virgule
        if "Poids total" in df_city_display.columns:
            df_city_display["Poids total"] = df_city_display["Poids total"].map(lambda x: f"{x:.3f} kg" if pd.notna(x) else "")
        if "Volume total" in df_city_display.columns:
            df_city_display["Volume total"] = df_city_display["Volume total"].map(lambda x: f"{x:.3f} m³" if pd.notna(x) else "")
        if "Besoin estafette réel" in df_city_display.columns:
            df_city_display["Besoin estafette réel"] = df_city_display["Besoin estafette réel"].map(lambda x: f"{x:.1f}" if pd.notna(x) else "")
        
        # Vérifier si le DataFrame n'est pas vide
        if df_city_display.empty:
            st.info("ℹ️ Aucune ville à afficher (TRIPOLI exclue)")
        else:
            # Afficher le tableau avec le style CSS
            html_table_city = df_city_display.to_html(
                escape=False, 
                index=False, 
                classes="custom-table",
                border=0
            )
            
            st.markdown(f"""
            <div class="table-container">
                {html_table_city}
            </div>
            """, unsafe_allow_html=True)
        
        # Métriques résumées
        st.markdown("---")
        col1, col2, col3 = st.columns(3)
        
        with col1:
            total_villes = len(df_city_display)
            st.metric("🏙️ Total Villes", total_villes)
        
        with col2:
            # Calculer le total des BLs
            df_city_original_filtered = st.session_state.df_city[st.session_state.df_city["Ville"] != "TRIPOLI"]
            total_bls = df_city_original_filtered["Nombre de BLs"].sum() if "Nombre de BLs" in df_city_original_filtered.columns else 0
            st.metric("📦 Total BLs", int(total_bls))
        
        with col3:
            # Calculer le total des estafettes nécessaires
            total_estafettes = df_city_original_filtered["Besoin estafette réel"].sum() if "Besoin estafette réel" in df_city_original_filtered.columns else 0
            st.metric("🚐 Besoin Estafettes", f"{total_estafettes:.1f}")

        # Bouton de téléchargement
        excel_buffer_city = BytesIO()
        with pd.ExcelWriter(excel_buffer_city, engine='openpyxl') as writer:
            st.session_state.df_city.to_excel(writer, index=False, sheet_name="Besoin Estafette Ville")
        excel_buffer_city.seek(0)
        
        st.download_button(
            label="💾 Télécharger Besoin par Ville",
            data=excel_buffer_city,
            file_name="Besoin_Estafette_Ville.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    
    # --- Onglet Livraisons Client & Ville + Zone ---
    with tab_zone_group:
        st.subheader("Livraisons par Client & Ville + Zone")
        
        # Créer une copie du DataFrame
        df_liv_zone = st.session_state.df_grouped_zone.copy()
        
        # Préparer les données pour l'affichage HTML
        if "Article" in df_liv_zone.columns:
            # Transformer les articles avec retours à la ligne HTML - SANS "•"
            df_liv_zone["Article"] = df_liv_zone["Article"].astype(str).apply(
                lambda x: "<br>".join(a.strip() for a in x.split(",") if a.strip())
            )
        
        # Formater les nombres - 3 chiffres après la virgule
        if "Poids total" in df_liv_zone.columns:
            df_liv_zone["Poids total"] = df_liv_zone["Poids total"].map(lambda x: f"{x:.3f} kg" if pd.notna(x) else "")
        if "Volume total" in df_liv_zone.columns:
            df_liv_zone["Volume total"] = df_liv_zone["Volume total"].map(lambda x: f"{x:.3f} m³" if pd.notna(x) else "")
        
        # Afficher le tableau avec le style CSS
        html_table_zone_group = df_liv_zone.to_html(
            escape=False, 
            index=False, 
            classes="custom-table",
            border=0
        )
        
        st.markdown(f"""
        <div class="table-container">
            {html_table_zone_group}
        </div>
        """, unsafe_allow_html=True)
        
        # Métriques résumées
        st.markdown("---")
        col1, col2, col3 = st.columns(3)
        
        with col1:
            total_livraisons_zone = len(df_liv_zone)
            st.metric("📦 Total Livraisons", total_livraisons_zone)
        
        with col2:
            zones_count = df_liv_zone["Zone"].nunique()
            st.metric("🌍 Zones", zones_count)
        
        with col3:
            villes_count = df_liv_zone["Ville"].nunique()
            st.metric("🏙️ Villes", villes_count)
        
        # Bouton de téléchargement
        excel_buffer_zone_group = BytesIO()
        with pd.ExcelWriter(excel_buffer_zone_group, engine='openpyxl') as writer:
            st.session_state.df_grouped_zone.to_excel(writer, index=False, sheet_name="Livraisons Client Ville Zone")
        excel_buffer_zone_group.seek(0)
        
        st.download_button(
            label="💾 Télécharger Livraisons Client/Ville/Zone",
            data=excel_buffer_zone_group,
            file_name="Livraisons_Client_Ville_Zone.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    
    # --- Onglet Besoin Estafette par Zone ---
    with tab_zone_summary:
        st.subheader("Besoin Estafette par Zone")
        
        # Créer une copie du DataFrame et renommer la colonne
        df_zone_display = st.session_state.df_zone.copy()
        
        # RENOMMER LA COLONNE "Nombre livraisons" en "Nombre de BLs"
        df_zone_display = df_zone_display.rename(columns={"Nombre livraisons": "Nombre de BLs"})
        
        # Formater les nombres
        if "Poids total" in df_zone_display.columns:
            df_zone_display["Poids total"] = df_zone_display["Poids total"].map(lambda x: f"{x:.3f} kg" if pd.notna(x) else "")
        if "Volume total" in df_zone_display.columns:
            df_zone_display["Volume total"] = df_zone_display["Volume total"].map(lambda x: f"{x:.3f} m³" if pd.notna(x) else "")
        if "Besoin estafette réel" in df_zone_display.columns:
            df_zone_display["Besoin estafette réel"] = df_zone_display["Besoin estafette réel"].map(lambda x: f"{x:.1f}" if pd.notna(x) else "")
        if "Nombre de BLs" in df_zone_display.columns:
            df_zone_display["Nombre de BLs"] = df_zone_display["Nombre de BLs"].map(lambda x: f"{int(x)}" if pd.notna(x) else "")
        
        # Afficher le tableau avec le style CSS
        html_table_zone = df_zone_display.to_html(
            escape=False, 
            index=False, 
            classes="custom-table",
            border=0
        )
        
        st.markdown(f"""
        <div class="table-container">
            {html_table_zone}
        </div>
        """, unsafe_allow_html=True)
        
        # Métriques résumées
        st.markdown("---")
        col1, col2, col3 = st.columns(3)
        
        with col1:
            total_zones = len(df_zone_display)
            st.metric("🌍 Total Zones", total_zones)
        
        with col2:
            # Calculer le total des BLs
            if "Nombre livraisons" in st.session_state.df_zone.columns:
                total_bls_zone = st.session_state.df_zone["Nombre livraisons"].sum()
            else:
                total_bls_zone = 0
            st.metric("📦 Total BLs", int(total_bls_zone))
        
        with col3:
            # Calculer le total des estafettes nécessaires
            total_estafettes_zone = st.session_state.df_zone["Besoin estafette réel"].sum() if "Besoin estafette réel" in st.session_state.df_zone.columns else 0
            st.metric("🚐 Besoin Estafettes", f"{total_estafettes_zone:.1f}")
        
        # Bouton de téléchargement
        excel_buffer_zone = BytesIO()
        with pd.ExcelWriter(excel_buffer_zone, engine='openpyxl') as writer:
            st.session_state.df_zone.to_excel(writer, index=False, sheet_name="Besoin Estafette Zone")
        excel_buffer_zone.seek(0)
        
        st.download_button(
            label="💾 Télécharger Besoin par Zone",
            data=excel_buffer_zone,
            file_name="Besoin_Estafette_Zone.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    
    # --- Onglet Graphiques ---
    with tab_charts:
        st.subheader("Statistiques par Ville")
        
        # FILTRER LES DONNÉES POUR EXCLURE TRIPOLI
        df_filtered = st.session_state.df_city[st.session_state.df_city["Ville"] != "TRIPOLI"]
        
        if not df_filtered.empty:
            # Configuration commune pour tous les graphiques
            chart_config = {
                'color_discrete_sequence': ['#0369A1'],  # BLEU ROYAL
                'template': 'plotly_white',
            }
            
            col1, col2 = st.columns(2)
            with col1:
                fig1 = px.bar(df_filtered, x="Ville", y="Poids total", **chart_config)
                fig1.update_layout(title_text="Poids total livré par ville", title_x=0.5)
                st.plotly_chart(fig1, use_container_width=True)
                
            with col2:
                fig2 = px.bar(df_filtered, x="Ville", y="Volume total", **chart_config)
                fig2.update_layout(title_text="Volume total livré par ville (m³)", title_x=0.5)
                st.plotly_chart(fig2, use_container_width=True)

            col3, col4 = st.columns(2)
            with col3:
                # DIAGRAMME CORRIGÉ : Nombre de BL par ville
                df_chart = df_filtered.rename(columns={"Nombre livraisons": "Nombre de BLs"})
                fig3 = px.bar(df_chart, x="Ville", y="Nombre de BLs", **chart_config)
                fig3.update_layout(title_text="Nombre de BL par ville", title_x=0.5)
                st.plotly_chart(fig3, use_container_width=True)
                
            with col4:
                fig4 = px.bar(df_filtered, x="Ville", y="Besoin estafette réel", **chart_config)
                fig4.update_layout(title_text="Besoin en Estafettes par ville", title_x=0.5)
                st.plotly_chart(fig4, use_container_width=True)
        else:
            st.info("ℹ️ Aucune donnée disponible pour les graphiques (TRIPOLI exclue)")
    
    # Navigation entre pages
    st.markdown("---")
    col_nav1, col_nav2, col_nav3 = st.columns(3)
    
    with col_nav1:
        if st.button("← Retour à l'importation", use_container_width=True):
            st.session_state.page = "import"
            st.rerun()
    
    with col_nav2:
        if st.button("📊 Exporter toute l'analyse", use_container_width=True):
            # Créer un fichier Excel avec tous les onglets
            from io import BytesIO
            
            excel_buffer = BytesIO()
            with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
                if st.session_state.df_grouped is not None:
                    st.session_state.df_grouped.drop(columns=["Zone"], errors='ignore').to_excel(writer, sheet_name="Livraisons Client Ville", index=False)
                if st.session_state.df_city is not None:
                    st.session_state.df_city.to_excel(writer, sheet_name="Besoin par Ville", index=False)
                if st.session_state.df_grouped_zone is not None:
                    st.session_state.df_grouped_zone.to_excel(writer, sheet_name="Livraisons Client Zone", index=False)
                if st.session_state.df_zone is not None:
                    st.session_state.df_zone.to_excel(writer, sheet_name="Besoin par Zone", index=False)
            
            excel_buffer.seek(0)
            
            st.download_button(
                label="💾 Télécharger l'analyse complète",
                data=excel_buffer,
                file_name="Analyse_Complete_Livraisons.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
    
    with col_nav3:
        if st.button("🚚 Passer à l'optimisation →", type="primary", use_container_width=True):
            st.session_state.page = "optimisation"
            st.rerun()

# =====================================================
# FONCTIONS DE CALLBACK POUR LA LOCATION
# =====================================================

def update_propositions_view():
    """Met à jour le DataFrame de propositions après une action."""
    if st.session_state.rental_processor:
        st.session_state.propositions = st.session_state.rental_processor.detecter_propositions()
        
        # Vérifier si le DataFrame de propositions n'est pas vide et contient la colonne 'Client'
        if (st.session_state.propositions is not None and 
            not st.session_state.propositions.empty and 
            'Client' in st.session_state.propositions.columns):
            
            # Réinitialiser la sélection si le client n'est plus dans les propositions ouvertes
            if (st.session_state.selected_client is not None and 
                st.session_state.selected_client not in st.session_state.propositions['Client'].astype(str).tolist()):
                st.session_state.selected_client = None
    else:
        st.session_state.propositions = pd.DataFrame()

def handle_location_action(accepter):
    """Gère l'acceptation ou le refus de la proposition de location."""
    if st.session_state.rental_processor and st.session_state.selected_client:
        try:
            # Assurer que le client est une chaîne valide
            client_to_process = str(st.session_state.selected_client)
            ok, msg, _ = st.session_state.rental_processor.appliquer_location(
                client_to_process, accepter=accepter
            )
            st.session_state.message = msg
            update_propositions_view()
            st.rerun()
        except Exception as e:
            st.session_state.message = f"❌ Erreur lors du traitement : {str(e)}"
    elif not st.session_state.selected_client:
        st.session_state.message = "⚠️ Veuillez sélectionner un client à traiter."
    else:
        st.session_state.message = "⚠️ Le processeur de location n'est pas initialisé."

def accept_location_callback():
    handle_location_action(True)

def refuse_location_callback():
    handle_location_action(False)

# =====================================================
# PAGE 3: Optimisation & location
# =====================================================
def page_optimisation():
    st.markdown("<h1 class='main-header'>3. 🚚 Optimisation & location</h1>", unsafe_allow_html=True)
    
    if not st.session_state.data_processed:
        st.warning("⚠️ Veuillez d'abord importer les données dans la page 1.")
        if st.button("📥 Retour à l'importation"):
            st.session_state.page = "import"
            st.rerun()
        return
    
    # CSS POUR LES TABLEAUX DE LA SECTION 3
    st.markdown("""
    <style>
        /* Style général du tableau */
        .custom-table-rental {
            width: 100%;
            border-collapse: collapse;
            font-family: Arial, sans-serif;
            font-size: 14px;
            box-shadow: 0 2px 8px rgba(0,0,0,0.1);
            border-radius: 8px;
            overflow: hidden;
        }
        
        /* En-têtes du tableau - BLEU ROYAL SANS DÉGRADÉ */
        .custom-table-rental th {
            background-color: #0369A1;
            color: white;
            padding: 12px 8px;
            text-align: center;
            border: 2px solid #4682B4;
            font-weight: normal;
            font-size: 13px;
            vertical-align: middle;
        }
        
        /* Cellules du tableau - TOUTES EN BLANC */
        .custom-table-rental td {
            padding: 10px 8px;
            text-align: center;
            border: 1px solid #B0C4DE;
            background-color: white;
            color: #000000;
            vertical-align: middle;
            font-weight: normal;
        }
        
        /* Bordures visibles pour toutes les cellules */
        .custom-table-rental th, 
        .custom-table-rental td {
            border: 1px solid #B0C4DE !important;
        }
        
        /* Bordures épaisses pour l'extérieur du tableau */
        .custom-table-rental {
            border: 2px solid #4682B4 !important;
        }
        
        /* Style pour les cellules numériques - SANS GRAS */
        .custom-table-rental td:nth-child(2),
        .custom-table-rental td:nth-child(3),
        .custom-table-rental td:nth-child(4),
        .custom-table-rental td:nth-child(5),
        .custom-table-rental td:nth-child(6) {
            font-weight: normal;
            color: #000000 !important;
            vertical-align: middle;
        }
        
        /* Conteneur du tableau avec défilement horizontal */
        .table-container-rental {
            overflow-x: auto;
            margin: 1rem 0;
            border-radius: 8px;
            border: 2px solid #4682B4;
        }
        
        /* Supprimer l'alternance des couleurs - TOUTES LES LIGNES BLANCHES */
        .custom-table-rental tr:nth-child(even) td {
            background-color: white !important;
        }
        
        /* Survol des lignes - léger effet */
        .custom-table-rental tr:hover td {
            background-color: #F0F8FF !important;
        }
        
        /* Style spécifique pour les cellules multilignes (BL inclus) */
        .multiline-cell {
            line-height: 1.4;
            text-align: left !important;
            padding: 8px !important;
            font-weight: normal;
        }
    </style>
    """, unsafe_allow_html=True)
    
    st.markdown(f"🔸 Si un client dépasse **{SEUIL_POIDS} kg** ou **{SEUIL_VOLUME} m³**, une location est proposée (si non déjà décidée).")
    
    # Initialiser propositions si nécessaire
    if st.session_state.propositions is None and st.session_state.rental_processor:
        update_propositions_view()
    
    # Onglets pour différentes fonctionnalités
    tab1, tab2, tab3, tab4, tab5 = st.tabs([
        "📋 Propositions Location", 
        "🔄 Transfert BLs", 
        "📦 Ajout Objets",
        "✅ VALIDATION DES VOYAGES APRÈS TRANSFERT",
        "🚛 Attribution des véhicules et chauffeurs"
    ])
    
    # --- Onglet 1: Propositions de Location ---
    with tab1:
        st.subheader("Propositions de Location de Camion")
        
        if st.session_state.propositions is not None and not st.session_state.propositions.empty:
            col_prop, col_details = st.columns([2, 3])
            
            with col_prop:
                st.markdown("### Propositions ouvertes")
                
                # Vérifier si la colonne 'Client' existe
                if 'Client' in st.session_state.propositions.columns:
                    # FORMATAGE DU TABLEAU DES PROPOSITIONS AVEC STYLE CSS
                    propositions_display = st.session_state.propositions.copy()
                    
                    # Formater les nombres
                    if "Poids total (kg)" in propositions_display.columns:
                        propositions_display["Poids total (kg)"] = propositions_display["Poids total (kg)"].map(
                            lambda x: f"{float(x):.3f}" if pd.notna(x) else ""
                        )
                    if "Volume total (m³)" in propositions_display.columns:
                        propositions_display["Volume total (m³)"] = propositions_display["Volume total (m³)"].map(
                            lambda x: f"{float(x):.3f}" if pd.notna(x) else ""
                        )
                    
                    # Afficher le tableau avec le style CSS
                    html_table_propositions = propositions_display.to_html(
                        escape=False, 
                        index=False, 
                        classes="custom-table-rental",
                        border=0
                    )
                    
                    st.markdown(f"""
                    <div class="table-container-rental">
                        {html_table_propositions}
                    </div>
                    """, unsafe_allow_html=True)
                    
                    # MÉTRIQUES RÉSUMÉES
                    st.markdown("---")
                    col_metric1, col_metric2, col_metric3 = st.columns(3)

                    with col_metric1:
                        total_propositions = len(st.session_state.propositions)
                        st.metric("📋 Propositions ouvertes", total_propositions)

                    with col_metric2:
                        # Calculer le nombre de clients dépassant le seuil de POIDS
                        clients_poids = len(st.session_state.propositions[
                            st.session_state.propositions["Poids total (kg)"] >= SEUIL_POIDS
                        ]) if "Poids total (kg)" in st.session_state.propositions.columns else 0
                        st.metric("⚖️ Dépassement poids", clients_poids)

                    with col_metric3:
                        # Calculer le nombre de clients dépassant le seuil de VOLUME
                        clients_volume = len(st.session_state.propositions[
                            st.session_state.propositions["Volume total (m³)"] >= SEUIL_VOLUME
                        ]) if "Volume total (m³)" in st.session_state.propositions.columns else 0
                        st.metric("📦 Dépassement volume", clients_volume)

                    # Sélection du client
                    client_options = st.session_state.propositions['Client'].astype(str).tolist()
                    client_options_with_empty = [""] + client_options
                    
                    # Index de sélection par défaut
                    default_index = 0
                    if st.session_state.selected_client in client_options:
                        default_index = client_options_with_empty.index(st.session_state.selected_client)
                    elif len(client_options) > 0:
                        default_index = 1

                    st.session_state.selected_client = st.selectbox(
                        "Client à traiter :", 
                        options=client_options_with_empty, 
                        index=default_index,
                        key='client_select_opt' 
                    )
                else:
                    st.warning("⚠️ Format de données incorrect dans les propositions.")
                    st.session_state.selected_client = None

                # Boutons d'action
                col_btn_acc, col_btn_ref = st.columns(2)
                is_client_selected = st.session_state.selected_client != "" and st.session_state.selected_client is not None
                
                with col_btn_acc:
                    st.button(
                        "✅ Accepter la location", 
                        on_click=accept_location_callback, 
                        disabled=not is_client_selected,
                        use_container_width=True,
                        type="primary"
                    )
                with col_btn_ref:
                    st.button(
                        "❌ Refuser la proposition", 
                        on_click=refuse_location_callback, 
                        disabled=not is_client_selected,
                        use_container_width=True,
                        type="secondary"
                    )
                
                # Afficher les messages
                if st.session_state.message:
                    if st.session_state.message.startswith("✅"):
                        st.success(st.session_state.message)
                    elif st.session_state.message.startswith("❌"):
                        st.error(st.session_state.message)
                    elif st.session_state.message.startswith("⚠️"):
                        st.warning(st.session_state.message)

            with col_details:
                st.markdown("### Détails de la commande client")
                is_client_selected = st.session_state.selected_client != "" and st.session_state.selected_client is not None
                
                if is_client_selected:
                    try:
                        resume, details_df = st.session_state.rental_processor.get_details_client(
                            st.session_state.selected_client
                        )
                        
                        # Afficher le résumé
                        st.markdown(f"**{resume}**")
                        
                        # FORMATAGE DU TABLEAU DES DÉTAILS AVEC STYLE CSS
                        if not details_df.empty:
                            details_display = details_df.copy()
                            
                            # Formatage simple et sécurisé des colonnes
                            def format_numeric_column(series, decimals, unit=""):
                                """Formate une colonne numérique avec le nombre de décimales et unité spécifiés"""
                                formatted_series = series.copy()
                                for i, value in enumerate(series):
                                    if pd.notna(value) and value != "":
                                        try:
                                            # Essayer de convertir en float
                                            if isinstance(value, str):
                                                # Nettoyer la valeur si c'est une string
                                                clean_value = value.replace(' kg', '').replace(' m³', '').replace('%', '').strip()
                                                num_value = float(clean_value)
                                            else:
                                                num_value = float(value)
                                            
                                            # Formater selon le nombre de décimales
                                            if decimals == 3:
                                                formatted_value = f"{num_value:.3f}"
                                            elif decimals == 2:
                                                formatted_value = f"{num_value:.2f}"
                                            elif decimals == 1:
                                                formatted_value = f"{num_value:.1f}"
                                            else:
                                                formatted_value = f"{num_value:.0f}"
                                            
                                            formatted_series.iloc[i] = f"{formatted_value}{unit}"
                                        except (ValueError, TypeError):
                                            # Si conversion échoue, garder la valeur originale
                                            formatted_series.iloc[i] = str(value)
                                    else:
                                        formatted_series.iloc[i] = ""
                                return formatted_series
                            
                            # Formater les colonnes numériques
                            if "Poids total" in details_display.columns:
                                details_display["Poids total"] = format_numeric_column(details_display["Poids total"], 3, " kg")
                            
                            if "Volume total" in details_display.columns:
                                details_display["Volume total"] = format_numeric_column(details_display["Volume total"], 3, " m³")
                            
                            if "Taux d'occupation (%)" in details_display.columns:
                                details_display["Taux d'occupation (%)"] = format_numeric_column(details_display["Taux d'occupation (%)"], 2, "%")
                            
                            # Gestion spéciale pour "BL inclus" - format multiligne
                            if "BL inclus" in details_display.columns:
                                details_display["BL inclus"] = details_display["BL inclus"].astype(str).apply(
                                    lambda x: "<br>".join(bl.strip() for bl in x.split(";")) if ";" in x else x
                                )
                            
                            # Afficher le tableau avec le style CSS
                            html_table_details = details_display.to_html(
                                escape=False, 
                                index=False, 
                                classes="custom-table-rental",
                                border=0
                            )
                            
                            st.markdown(f"""
                            <div class="table-container-rental">
                                {html_table_details}
                            </div>
                            """, unsafe_allow_html=True)
                            
                            # MÉTRIQUES POUR LES DÉTAILS
                            st.markdown("---")
                            col_det1, col_det2, col_det3 = st.columns(3)
                            
                            with col_det1:
                                total_camions = len(details_display)
                                st.metric("🚚 Nombre de camions", total_camions)
                            
                            with col_det2:
                                # Calculer le poids total à partir des données brutes
                                try:
                                    if "Poids total" in details_df.columns:
                                        poids_total = 0
                                        for value in details_df["Poids total"]:
                                            if pd.notna(value):
                                                try:
                                                    # Nettoyer la valeur si elle contient des unités
                                                    if isinstance(value, str):
                                                        clean_value = value.replace(' kg', '').replace('m³', '').strip()
                                                    else:
                                                        clean_value = str(value)
                                                    poids_total += float(clean_value)
                                                except (ValueError, TypeError):
                                                    continue
                                        st.metric("📦 Poids total", f"{poids_total:.1f} kg")
                                    else:
                                        st.metric("📦 Poids total", "N/A")
                                except Exception as e:
                                    st.metric("📦 Poids total", "Erreur")
                            
                            with col_det3:
                                # Calculer le volume total à partir des données brutes
                                try:
                                    if "Volume total" in details_df.columns:
                                        volume_total = 0
                                        for value in details_df["Volume total"]:
                                            if pd.notna(value):
                                                try:
                                                    # Nettoyer la valeur si elle contient des unités
                                                    if isinstance(value, str):
                                                        clean_value = value.replace(' kg', '').replace('m³', '').strip()
                                                    else:
                                                        clean_value = str(value)
                                                    volume_total += float(clean_value)
                                                except (ValueError, TypeError):
                                                    continue
                                        st.metric("📏 Volume total", f"{volume_total:.3f} m³")
                                    else:
                                        st.metric("📏 Volume total", "N/A")
                                except Exception as e:
                                    st.metric("📏 Volume total", "Erreur")
                                
                    except Exception as e:
                        st.error(f"❌ Erreur lors de la récupération des détails : {str(e)}")
                        # Debug information
                        st.write("Détails de l'erreur :")
                        if 'details_df' in locals():
                            st.write("Colonnes disponibles :", details_df.columns.tolist())
                            if not details_df.empty:
                                st.write("Aperçu des données :")
                                st.dataframe(details_df.head())
                else:
                    st.info("Sélectionnez un client pour afficher les détails de la commande/estafettes.")
        else:
            st.success("✅ Aucune proposition de location de camion en attente de décision.")
            
            # Bouton pour forcer la détection
            if st.button("🔍 Vérifier à nouveau les propositions"):
                if st.session_state.rental_processor:
                   update_propositions_view()
                st.rerun()
        # =====================================================
        # 4. VOYAGES PAR ESTAFETTE OPTIMISÉ (Section 4 - Résultat final)
        # =====================================================
        st.header("4. 🚐 Voyages par Estafette Optimisé (Inclut Camions Loués)")

        try:
            # Récupération sécurisée du DataFrame
            if st.session_state.rental_processor:
                df_optimized_estafettes = st.session_state.rental_processor.get_df_result()
            elif "df_voyages" in st.session_state:
                df_optimized_estafettes = st.session_state.df_voyages.copy()
            else:
                st.error("❌ Données non disponibles. Veuillez exécuter le traitement complet.")
                st.stop()
            
            # Vérifier que le DataFrame n'est pas vide
            if df_optimized_estafettes.empty:
                st.warning("⚠️ Aucune donnée à afficher.")
                st.stop()
            
            # CORRECTION : Nettoyer les colonnes en double
            df_clean = df_optimized_estafettes.loc[:, ~df_optimized_estafettes.columns.duplicated()]
            
            # CORRECTION : TRIER PAR ZONE D'ABORD
            if "Zone" in df_clean.columns:
                # Extraire le numéro de zone pour un tri numérique
                df_clean["Zone_Num"] = df_clean["Zone"].str.extract('(\d+)').astype(float)
                df_clean = df_clean.sort_values("Zone_Num").drop("Zone_Num", axis=1)
                # Alternative plus simple si l'extraction échoue :
                # df_clean = df_clean.sort_values("Zone")
            
            # Définir l'ordre des colonnes pour l'affichage
            colonnes_ordre = [
                "Zone", "Véhicule N°", "Poids total chargé", "Volume total chargé",
                "Client(s) inclus", "Représentant(s) inclus", "BL inclus", 
                "Taux d'occupation (%)", "Location_camion", "Location_proposee", "Code Véhicule"
            ]
            
            # Filtrer seulement les colonnes qui existent
            colonnes_finales = [col for col in colonnes_ordre if col in df_clean.columns]
            
            # Créer le DataFrame d'affichage avec retours à la ligne POUR STREAMLIT
            df_display = df_clean[colonnes_finales].copy()
            
            # Transformer les colonnes avec retours à la ligne HTML pour l'affichage Streamlit
            if "Client(s) inclus" in df_display.columns:
                df_display["Client(s) inclus"] = df_display["Client(s) inclus"].astype(str).apply(
                    lambda x: "<br>".join(client.strip() for client in x.split(",")) if x != "nan" else ""
                )
            
            if "Représentant(s) inclus" in df_display.columns:
                df_display["Représentant(s) inclus"] = df_display["Représentant(s) inclus"].astype(str).apply(
                    lambda x: "<br>".join(rep.strip() for rep in x.split(",")) if x != "nan" else ""
                )
            
            if "BL inclus" in df_display.columns:
                df_display["BL inclus"] = df_display["BL inclus"].astype(str).apply(
                    lambda x: "<br>".join(bl.strip() for bl in x.split(";")) if x != "nan" else ""
                )
            
            # Formater les colonnes numériques pour l'affichage
            if "Poids total chargé" in df_display.columns:
                df_display["Poids total chargé"] = df_display["Poids total chargé"].map(lambda x: f"{x:.3f} kg")
            if "Volume total chargé" in df_display.columns:
                df_display["Volume total chargé"] = df_display["Volume total chargé"].map(lambda x: f"{x:.3f} m³")
            if "Taux d'occupation (%)" in df_display.columns:
                df_display["Taux d'occupation (%)"] = df_display["Taux d'occupation (%)"].map(lambda x: f"{x:.3f}%")
            
            # CSS POUR UN TABLEAU PROFESSIONNEL (identique aux autres sections)
            st.markdown("""
            <style>
            /* Style général du tableau */
            .custom-table-voyages {
                width: 100%;
                border-collapse: collapse;
                font-family: Arial, sans-serif;
                font-size: 14px;
                box-shadow: 0 2px 8px rgba(0,0,0,0.1);
                border-radius: 8px;
                overflow: hidden;
            }
            
            /* En-têtes du tableau - BLEU ROYAL SANS DÉGRADÉ */
            .custom-table-voyages th {
                background-color: #0369A1;
                color: white;
                padding: 12px 8px;
                text-align: center;
                border: 2px solid #4682B4;
                font-weight: bold;
                font-size: 13px;
                vertical-align: middle;
            }
            
            /* Cellules du tableau - TOUTES EN BLANC */
            .custom-table-voyages td {
                padding: 10px 8px;
                text-align: center;
                border: 1px solid #B0C4DE;
                background-color: white;
                color: #000000;
                vertical-align: middle;
            }
            
            /* Bordures visibles pour toutes les cellules */
            .custom-table-voyages th, 
            .custom-table-voyages td {
                border: 1px solid #B0C4DE !important;
            }
            
            /* Bordures épaisses pour l'extérieur du tableau */
            .custom-table-voyages {
                border: 2px solid #4682B4 !important;
            }
            
            /* Conteneur du tableau avec défilement horizontal */
            .table-container-voyages {
                overflow-x: auto;
                margin: 1rem 0;
                border-radius: 8px;
                border: 2px solid #4682B4;
            }
            
            /* Supprimer l'alternance des couleurs - TOUTES LES LIGNES BLANCHES */
            .custom-table-voyages tr:nth-child(even) td {
                background-color: white !important;
            }
            
            /* Survol des lignes - léger effet */
            .custom-table-voyages tr:hover td {
                background-color: #F0F8FF !important;
            }
            
            /* Style pour les cellules multilignes */
            .custom-table-voyages td {
                line-height: 1.4;
            }
            </style>
            """, unsafe_allow_html=True)
            
            # Afficher le tableau avec le style CSS professionnel
            html_table = df_display.to_html(escape=False, index=False, classes="custom-table-voyages", border=0)
            
            st.markdown(f"""
            <div class="table-container-voyages">
                {html_table}
            </div>
            """, unsafe_allow_html=True)
            
            # MÉTRIQUES RÉSUMÉES
            st.markdown("---")
            col1, col2, col3, col4 = st.columns(4)
            
            with col1:
                total_voyages = len(df_display)
                st.metric("🚐 Total Voyages", total_voyages)
            
            with col2:
                total_zones = df_display["Zone"].nunique() if "Zone" in df_display.columns else 0
                st.metric("🌍 Zones couvertes", total_zones)
            
            with col3:
                camions_loues = df_display["Location_camion"].sum() if "Location_camion" in df_display.columns else 0
                st.metric("🚚 Camions loués", int(camions_loues))
            
            with col4:
                estafettes = total_voyages - camions_loues
                st.metric("📦 Estafettes", estafettes)
            
            # Préparer l'export Excel avec retours à la ligne \n
            df_export = df_clean.copy()
            
            # CORRECTION : S'assurer que l'export est aussi trié par zone
            if "Zone" in df_export.columns:
                df_export["Zone_Num"] = df_export["Zone"].str.extract('(\d+)').astype(float)
                df_export = df_export.sort_values("Zone_Num").drop("Zone_Num", axis=1)
            
            # Transformer les colonnes avec retours à la ligne \n pour Excel
            if "Client(s) inclus" in df_export.columns:
                df_export["Client(s) inclus"] = df_export["Client(s) inclus"].astype(str).apply(
                    lambda x: "\n".join(client.strip() for client in x.split(",")) if x != "nan" else ""
                )
            
            if "Représentant(s) inclus" in df_export.columns:
                df_export["Représentant(s) inclus"] = df_export["Représentant(s) inclus"].astype(str).apply(
                    lambda x: "\n".join(rep.strip() for rep in x.split(",")) if x != "nan" else ""
                )
            
            if "BL inclus" in df_export.columns:
                df_export["BL inclus"] = df_export["BL inclus"].astype(str).apply(
                    lambda x: "\n".join(bl.strip() for bl in x.split(";")) if x != "nan" else ""
                )
            
            # Formater les colonnes numériques pour l'export
            if "Poids total chargé" in df_export.columns:
                df_export["Poids total chargé"] = df_export["Poids total chargé"].round(3)
            if "Volume total chargé" in df_export.columns:
                df_export["Volume total chargé"] = df_export["Volume total chargé"].round(3)
            
            # Bouton de téléchargement avec formatage Excel
            from io import BytesIO
            import openpyxl
            from openpyxl.styles import Alignment
            
            excel_buffer = BytesIO()
            
            with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
                df_export.to_excel(writer, index=False, sheet_name="Voyages Optimisés")
                
                # Récupérer le workbook et worksheet pour appliquer le formatage
                workbook = writer.book
                worksheet = writer.sheets["Voyages Optimisés"]
                
                # Appliquer le style wrap_text aux colonnes avec retours à la ligne
                wrap_columns = []
                if "Client(s) inclus" in df_export.columns:
                    wrap_columns.append("Client(s) inclus")
                if "Représentant(s) inclus" in df_export.columns:
                    wrap_columns.append("Représentant(s) inclus")
                if "BL inclus" in df_export.columns:
                    wrap_columns.append("BL inclus")
                
                # Appliquer le format wrap_text à toutes les cellules des colonnes concernées
                for col_idx, col_name in enumerate(df_export.columns):
                    if col_name in wrap_columns:
                        col_letter = openpyxl.utils.get_column_letter(col_idx + 1)
                        for row in range(2, len(df_export) + 2):  # Commence à la ligne 2 (après l'en-tête)
                            cell = worksheet[f"{col_letter}{row}"]
                            cell.alignment = Alignment(wrap_text=True, vertical='top')
                
                # Ajuster la largeur des colonnes pour une meilleure visibilité
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = openpyxl.utils.get_column_letter(column[0].column)
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)  # Largeur max de 50
                    worksheet.column_dimensions[column_letter].width = adjusted_width
            
            excel_buffer.seek(0)
            
            st.download_button(
                label="💾 Télécharger Voyages Estafette Optimisés",
                data=excel_buffer,
                file_name="Voyages_Estafette_Optimises.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            
            # Mise à jour pour les sections suivantes
            st.session_state.df_voyages = df_clean

        except KeyError as e:
            st.error(f"❌ Erreur de colonne manquante : {e}")
            st.info("🔄 Tentative de récupération des données...")
            
            # Tentative de récupération
            if st.session_state.rental_processor:
                st.session_state.df_voyages = st.session_state.rental_processor.df_base.copy()
                st.rerun()
                
        except Exception as e:
            st.error(f"❌ Erreur lors de l'affichage des voyages optimisés: {str(e)}")
            # Afficher les données brutes pour debug
            st.write("Données brutes pour debug:")
            if st.session_state.rental_processor:
                st.write("Colonnes du df_base:", list(st.session_state.rental_processor.df_base.columns))
    # --- Onglet 2: Transfert BLs ---
    with tab2:
        st.subheader("🔁 Transfert de BLs entre véhicules")
        st.info("Réorganisez les livraisons entre estafettes/camions")
        
        MAX_POIDS = 1550  # kg
        MAX_VOLUME = 4.608  # m³
        
        if st.session_state.df_voyages is None and st.session_state.rental_processor:
            # Générer les voyages optimisés si pas encore fait
            df_optimized = st.session_state.rental_processor.get_df_result()
            st.session_state.df_voyages = df_optimized
        
        if st.session_state.df_voyages is not None and st.session_state.df_livraisons is not None:
            df_voyages = st.session_state.df_voyages.copy()
            df_livraisons = st.session_state.df_livraisons.copy()
            
            colonnes_requises = ["Zone", "Véhicule N°", "Poids total chargé", "Volume total chargé", "BL inclus"]
            
            if not all(col in df_voyages.columns for col in colonnes_requises):
                st.error(f"❌ Le DataFrame ne contient pas toutes les colonnes nécessaires : {', '.join(colonnes_requises)}")
            else:
                zones_disponibles = sorted(df_voyages["Zone"].dropna().unique().tolist())
                zone_selectionnee = st.selectbox("🌍 Sélectionner une zone", zones_disponibles)
                
                if zone_selectionnee:
                    df_zone = df_voyages[df_voyages["Zone"] == zone_selectionnee]
                    vehicules = sorted(df_zone["Véhicule N°"].dropna().unique().tolist())
                    
                    col1, col2 = st.columns(2)
                    with col1:
                        source = st.selectbox("🚐 Estafette / Camion source", vehicules)
                    with col2:
                        cible = st.selectbox("🎯 Estafette / Camion cible", [v for v in vehicules if v != source])
                    
                    if source and cible:
                        df_source = df_zone[df_zone["Véhicule N°"] == source]
                        if df_source.empty or df_source["BL inclus"].isna().all():
                            st.warning("⚠️ Aucun BL trouvé pour ce véhicule source.")
                        else:
                            st.subheader(f"📦 BLs actuellement assignés à {source}")
                            
                            # --- NOUVEAU : Créer un mapping BL → Client ---
                            bls_avec_clients = []
                            bls_simples = df_source["BL inclus"].iloc[0].split(";")
                            
                            for bl in bls_simples:
                                # Trouver le client correspondant à ce BL
                                client_info = df_livraisons[df_livraisons["No livraison"] == bl]
                                if not client_info.empty:
                                    client_nom = client_info["Client de l'estafette"].iloc[0]
                                    bl_affichage = f"{bl} - {client_nom}"
                                else:
                                    bl_affichage = f"{bl} - Client non trouvé"
                                bls_avec_clients.append(bl_affichage)
                            
                            # Affichage formaté avec clients
                            df_source_display = df_source[["Véhicule N°", "Poids total chargé", "Volume total chargé"]].copy()
                            df_source_display["BL inclus (avec clients)"] = "<br>".join(bls_avec_clients)
                            
                            df_source_display["Poids total chargé"] = df_source_display["Poids total chargé"].map(lambda x: f"{x:.3f} kg")
                            df_source_display["Volume total chargé"] = df_source_display["Volume total chargé"].map(lambda x: f"{x:.3f} m³")
                            
                            # CSS AMÉLIORÉ pour un tableau plus visible et bien centré
                            st.markdown("""
                            <style>
                            .centered-table {
                                margin-left: auto;
                                margin-right: auto;
                                display: table;
                                width: 100%;
                            }
                            .centered-table table {
                                margin: 0 auto;
                                border-collapse: collapse;
                                width: 100%;
                                font-family: Arial, sans-serif;
                                box-shadow: 0 2px 8px rgba(0,0,0,0.1);
                            }
                            .centered-table th {
                                background-color: #0369A1;
                                color: white;
                                padding: 12px 8px;
                                text-align: center;
                                border: 2px solid #555;
                                font-weight: bold;
                                font-size: 14px;
                                vertical-align: middle;
                            }
                            .centered-table td {
                                padding: 10px 8px;
                                text-align: center;
                                border: 2px solid #555;
                                background-color: #f9f9f9;
                                color: #333;
                                vertical-align: middle;
                            }
                            .centered-table tr:nth-child(even) td {
                                background-color: #f0f0f0;
                            }
                            .centered-table tr:hover td {
                                background-color: #e6f3ff;
                            }
                            </style>
                            """, unsafe_allow_html=True)
                            
                            # CSS SPÉCIFIQUE POUR LE MULTISELECT - VERSION GRIS
                            st.markdown("""
                            <style>
                            .stMultiSelect > div > div {
                                background-color: #F8FAFC !important;
                                border: 2px solid #CBD5E1 !important;
                                border-radius: 8px !important;
                            }
                            div[data-baseweb="select"] > div {
                                background-color: #F8FAFC !important;
                                border: 2px solid #CBD5E1 !important;
                                border-radius: 8px !important;
                            }
                            div[data-baseweb="select"] span[data-baseweb="tag"] {
                                background-color: #0369A1 !important;
                                color: white !important;
                                border-radius: 12px !important;
                                font-weight: bold;
                            }
                            div[role="listbox"] {
                                background-color: white !important;
                                border: 2px solid #CBD5E1 !important;
                            }
                            div[role="option"][aria-selected="true"] {
                                background-color: #F1F5F9 !important;
                                color: #475569 !important;
                            }
                            div[role="option"]:hover {
                                background-color: #E2E8F0 !important;
                            }
                            </style>
                            """, unsafe_allow_html=True)
                            
                            # Afficher le tableau avec CSS
                            html_content = f"""
                            <div class="centered-table">
                            {df_source_display.to_html(escape=False, index=False)}
                            </div>
                            """
                            st.markdown(html_content, unsafe_allow_html=True)
                            
                            # --- NOUVEAU : Sélection avec clients ---
                            st.subheader("📋 Sélectionner les BLs à transférer")
                            
                            # Créer les options avec format "BL - Client"
                            options_transfert = []
                            mapping_bl_original = {}  # Pour garder la correspondance BL original
                            
                            for bl in bls_simples:
                                client_info = df_livraisons[df_livraisons["No livraison"] == bl]
                                if not client_info.empty:
                                    client_nom = client_info["Client de l'estafette"].iloc[0]
                                    option_affichage = f"{bl} - {client_nom}"
                                else:
                                    option_affichage = f"{bl} - Client non trouvé"
                                
                                options_transfert.append(option_affichage)
                                mapping_bl_original[option_affichage] = bl
                            
                            # Multiselect avec clients
                            bls_selectionnes_affichage = st.multiselect(
                                "Sélectionnez les BLs à transférer (avec clients) :", 
                                options_transfert,
                                format_func=lambda x: x
                            )
                            
                            # Convertir la sélection en BLs simples pour le traitement
                            bls_selectionnes = [mapping_bl_original[bl_affichage] for bl_affichage in bls_selectionnes_affichage]
                            
                            if bls_selectionnes and st.button("🔁 Exécuter le transfert"):
                                df_bls_selection = df_livraisons[df_livraisons["No livraison"].isin(bls_selectionnes)]
                                poids_bls = df_bls_selection["Poids total"].sum()
                                volume_bls = df_bls_selection["Volume total"].sum()
                                
                                df_cible = df_zone[df_zone["Véhicule N°"] == cible]
                                poids_cible = df_cible["Poids total chargé"].sum()
                                volume_cible = df_cible["Volume total chargé"].sum()
                                
                                # Sauvegarder les BLs ORIGINAUX pour le debug
                                bls_source_avant = bls_simples.copy()  # BLs de la source avant transfert
                                
                                # Récupérer les BLs de la cible AVANT transfert
                                bls_cible_avant = []
                                if not df_cible.empty and "BL inclus" in df_cible.columns:
                                    bls_str = df_cible.iloc[0]["BL inclus"]
                                    if pd.notna(bls_str):
                                        bls_cible_avant = [b.strip() for b in bls_str.split(";") if b.strip()]
                                
                                if (poids_cible + poids_bls) > MAX_POIDS or (volume_cible + volume_bls) > MAX_VOLUME:
                                    st.warning("⚠️ Le transfert dépasse les limites de poids ou volume du véhicule cible.")
                                else:
                                    def transfer_bl(row):
                                        bls = row["BL inclus"].split(";") if pd.notna(row["BL inclus"]) else []
                                        
                                        if row["Véhicule N°"] == source:
                                            # RETIRER les BLs transférés de la source
                                            new_bls = [b for b in bls if b not in bls_selectionnes]
                                            row["BL inclus"] = ";".join(new_bls) if new_bls else ""
                                            row["Poids total chargé"] = max(0, row["Poids total chargé"] - poids_bls)
                                            row["Volume total chargé"] = max(0, row["Volume total chargé"] - volume_bls)
                                        
                                        elif row["Véhicule N°"] == cible:
                                            # AJOUTER les BLs transférés à la cible
                                            # Combiner BLs existants + BLs transférés
                                            all_bls = bls + bls_selectionnes
                                            # Supprimer les doublons
                                            all_bls = list(dict.fromkeys(all_bls))
                                            row["BL inclus"] = ";".join(all_bls) if all_bls else ""
                                            row["Poids total chargé"] = row["Poids total chargé"] + poids_bls
                                            row["Volume total chargé"] = row["Volume total chargé"] + volume_bls
                                        
                                        return row
                                    
                                    df_voyages = df_voyages.apply(transfer_bl, axis=1)
                                    st.session_state.df_voyages = df_voyages
                                    
                                    # Afficher un résumé du transfert avec clients
                                    clients_transferes = df_bls_selection["Client de l'estafette"].unique()
                                    st.success(f"""
                                    ✅ Transfert réussi !
                                    - **{len(bls_selectionnes)} BL(s)** déplacé(s) de **{source}** vers **{cible}**
                                    - **Clients concernés :** {', '.join(clients_transferes)}
                                    - **Poids transféré :** {poids_bls:.1f} kg
                                    - **Volume transféré :** {volume_bls:.3f} m³
                                    """)
                                    
                                    # --- VÉRIFICATION DÉTAILLÉE ---
                                    st.markdown("---")
                                   #st.subheader("🔍 Vérification du transfert")
                                    
                                    # Récupérer les données APRÈS transfert
                                    df_source_apres = df_voyages[df_voyages["Véhicule N°"] == source]
                                    df_cible_apres = df_voyages[df_voyages["Véhicule N°"] == cible]
                                    
                                    # BLs de la source APRÈS transfert
                                    bls_source_apres = []
                                    if not df_source_apres.empty and "BL inclus" in df_source_apres.columns:
                                        bls_str = df_source_apres.iloc[0]["BL inclus"]
                                        if pd.notna(bls_str):
                                            bls_source_apres = [b.strip() for b in bls_str.split(";") if b.strip()]
                                    
                                    # BLs de la cible APRÈS transfert  
                                    bls_cible_apres = []
                                    if not df_cible_apres.empty and "BL inclus" in df_cible_apres.columns:
                                        bls_str = df_cible_apres.iloc[0]["BL inclus"]
                                        if pd.notna(bls_str):
                                            bls_cible_apres = [b.strip() for b in bls_str.split(";") if b.strip()]
                                    
                                    # Afficher la vérification
                                   #col_verif1, col_verif2 = st.columns(2)
                                    
                                  # with col_verif1:
                                    #   st.markdown(f"""
                                     #  **✅ Vérification Source ({source}):**
                                      # - BLs avant: {', '.join(bls_source_avant) if bls_source_avant else 'Aucun'}
                                      # - BLs transférés: {', '.join(bls_selectionnes)}
                                      # - BLs après: {', '.join(bls_source_apres) if bls_source_apres else 'Aucun'}
                                      # - BLs retirés avec succès: {'✅ OUI' if all(bl not in bls_source_apres for bl in bls_selectionnes) else '❌ NON'}
                                     #  """)
                                    
                                   #with col_verif2:
                                    #   st.markdown(f"""
                                     #  **✅ Vérification Cible ({cible}):**
                                     #  - BLs avant: {', '.join(bls_cible_avant) if bls_cible_avant else 'Aucun'}
                                     #  - BLs ajoutés: {', '.join(bls_selectionnes)}
                                      # - BLs après: {', '.join(bls_cible_apres) if bls_cible_apres else 'Aucun'}
                                      # - BLs ajoutés avec succès: {'✅ OUI' if all(bl in bls_cible_apres for bl in bls_selectionnes) else '❌ NON'}
                                       #""")
                                    
                                    # --- AFFICHAGE DU TABLEAU COMPLET ---
                                   #st.markdown("---")
                                    st.subheader("📊 Voyages après transfert (toutes les zones)")
                                    
                                    # Préparer l'affichage
                                    df_display = df_voyages.sort_values(by=["Zone", "Véhicule N°"]).copy()
                                    
                                    # Mettre en évidence les véhicules modifiés
                                   #def highlight_transferred(row):
                                   #    if row["Véhicule N°"] == source:
                                    #       return ['background-color: #FFE6E6' if col == "BL inclus" else '' for col in df_display.columns]
                                    #   elif row["Véhicule N°"] == cible:
                                     #      return ['background-color: #E6FFE6' if col == "BL inclus" else '' for col in df_display.columns]
                                     #  return [''] * len(df_display.columns)
                                    
                                    # Formater pour l'affichage HTML
                                    if "BL inclus" in df_display.columns:
                                        df_display["BL inclus"] = df_display["BL inclus"].astype(str).apply(
                                            lambda x: "<br>".join(bl.strip() for bl in x.split(";")) if x != "nan" and x != "" else "Aucun BL"
                                        )
                                    
                                    df_display["Poids total chargé"] = df_display["Poids total chargé"].map(lambda x: f"{x:.3f} kg")
                                    df_display["Volume total chargé"] = df_display["Volume total chargé"].map(lambda x: f"{x:.3f} m³")
                                    
                                    # CSS pour mettre en évidence
                                    st.markdown("""
                                    <style>
                                    .highlight-source {
                                        border-left: 5px solid #FF6B6B !important;
                                    }
                                    .highlight-target {
                                        border-left: 5px solid #51CF66 !important;
                                    }
                                    </style>
                                    """, unsafe_allow_html=True)
                                    
                                    # Ajouter des classes CSS pour les lignes modifiées
                                    html_table = df_display.to_html(escape=False, index=False, classes="custom-table-voyages", border=0)
                                    
                                    # Modifier le HTML pour ajouter des classes
                                    for i, row in df_display.iterrows():
                                        if row["Véhicule N°"] == source:
                                            html_table = html_table.replace(f'<tr>', f'<tr class="highlight-source">', 1)
                                        elif row["Véhicule N°"] == cible:
                                            html_table = html_table.replace(f'<tr>', f'<tr class="highlight-target">', 1)
                                    
                                    st.markdown(f"""
                                            <div class="table-container-voyages">
                                                {html_table}
                                            </div>
                                            """, unsafe_allow_html=True)
                                    
                                   #st.info("**Légende :** 🔴 Véhicule source (BLs retirés) | 🟢 Véhicule cible (BLs ajoutés)")
                                    
                                    # --- Export Excel avec retours à la ligne \n ---
                                    df_export = df_voyages.copy()
                                    
                                    # Transformer les BL avec retours à la ligne \n pour Excel
                                    if "BL inclus" in df_export.columns:
                                        df_export["BL inclus"] = df_export["BL inclus"].astype(str).apply(
                                            lambda x: "\n".join(bl.strip() for bl in x.split(";")) if x != "nan" else ""
                                        )
                                    
                                    df_export["Poids total chargé"] = df_export["Poids total chargé"].round(3)
                                    df_export["Volume total chargé"] = df_export["Volume total chargé"].round(3)
                                    
                                    from io import BytesIO
                                    import openpyxl
                                    from openpyxl.styles import Alignment
                                    
                                    excel_buffer = BytesIO()
                                    with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
                                        df_export.to_excel(writer, index=False, sheet_name='Transfert BLs')
                                        
                                        # Appliquer le format wrap_text pour Excel
                                        workbook = writer.book
                                        worksheet = writer.sheets['Transfert BLs']
                                        
                                        # Appliquer le style wrap_text à la colonne BL inclus
                                        if "BL inclus" in df_export.columns:
                                            for col_idx, col_name in enumerate(df_export.columns):
                                                if col_name == "BL inclus":
                                                    col_letter = openpyxl.utils.get_column_letter(col_idx + 1)
                                                    for row in range(2, len(df_export) + 2):
                                                        cell = worksheet[f"{col_letter}{row}"]
                                                        cell.alignment = Alignment(wrap_text=True, vertical='top')
                                    
                                    excel_buffer.seek(0)
                                    
                                    st.download_button(
                                        label="💾 Télécharger le tableau mis à jour (XLSX)",
                                        data=excel_buffer,
                                        file_name="voyages_apres_transfert.xlsx",
                                        mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                                    )
                            else:
                                st.info("ℹ️ Générez d'abord les voyages optimisés dans l'onglet 1")
    
   # --- Onglet 3: Attribution véhicules/chauffeurs ---
    with tab3:
        # =====================================================
        # 6️⃣ AJOUT D'OBJETS MANUELS AUX VÉHICULES (CORRIGÉ)
        # =====================================================
        st.markdown("## 📦 AJOUT D'OBJETS MANUELS AUX VÉHICULES")

        # IMPORT SIMPLIFIÉ
        try:
            from backend import TruckTransferManager
            TM_IMPORTED = True
        except ImportError as e:
            st.error(f"⚠️ Import impossible: {e}")
            st.stop()

        # VÉRIFICATION DES DONNÉES REQUISES
        if "df_voyages" not in st.session_state or st.session_state.df_voyages is None:
            st.error("❌ df_voyages non disponible")
            st.info("Veuillez d'abord exécuter la section d'optimisation")
            st.stop()

        if "df_livraisons" not in st.session_state or st.session_state.df_livraisons is None:
            st.error("❌ df_livraisons non disponible")
            st.info("Veuillez d'abord exécuter la section d'optimisation")
            st.stop()

        # INITIALISATION DU TRANSFER MANAGER
        if "transfer_manager" not in st.session_state or st.session_state.transfer_manager is None:
            try:
                st.session_state.transfer_manager = TruckTransferManager(
                    st.session_state.df_voyages.copy(),
                    st.session_state.df_livraisons.copy()
                )
                st.success("✅ Gestionnaire de transfert initialisé")
            except Exception as e:
                st.error(f"❌ Erreur d'initialisation: {e}")
                st.stop()

        # VÉRIFICATION QUE LE GESTIONNAIRE EST VALIDE
        if not hasattr(st.session_state.transfer_manager, 'add_manual_object'):
            st.error("❌ Le gestionnaire de transfert n'a pas la méthode add_manual_object")
            st.stop()

        # CONTINUER AVEC L'INTERFACE UTILISATEUR
        df_voyages = st.session_state.df_voyages.copy()

        col1, col2, col3 = st.columns(3)

        with col1:
            # Sélection de la zone
            zones_disponibles = sorted(df_voyages["Zone"].dropna().unique().tolist())
            zone_objet = st.selectbox("🌍 Zone", zones_disponibles, key="select_zone_objet")

        with col2:
            # Sélection du véhicule dans la zone choisie
            if zone_objet:
                vehicules_zone = sorted(
                    df_voyages[df_voyages["Zone"] == zone_objet]["Véhicule N°"].dropna().unique().tolist()
                )
                vehicule_objet = st.selectbox("🚚 Véhicule", vehicules_zone, key="select_vehicule_objet")
            else:
                vehicule_objet = None

        with col3:
            # Informations sur le véhicule sélectionné
            if zone_objet and vehicule_objet:
                try:
                    vehicule_data = df_voyages[
                        (df_voyages["Zone"] == zone_objet) & 
                        (df_voyages["Véhicule N°"] == vehicule_objet)
                    ].iloc[0]
                    
                    # Détecter si c'est un camion loué
                    is_camion = "CAMION" in str(vehicule_data.get("Code Véhicule", "")).upper()
                    
                    # Utiliser les capacités par défaut si non spécifiées
                    capacite_poids = 30500 if is_camion else 1550
                    capacite_volume = 77.5 if is_camion else 4.608
                    
                    poids_actuel = vehicule_data.get("Poids total chargé", 0)
                    volume_actuel = vehicule_data.get("Volume total chargé", 0)
                    
                    st.metric(
                        "📊 Capacité utilisée", 
                        f"{poids_actuel:.1f}kg / {capacite_poids}kg",
                        f"{volume_actuel:.3f}m³ / {capacite_volume}m³"
                    )
                except Exception as e:
                    st.error(f"Erreur chargement véhicule: {e}")

        # Formulaire d'ajout d'objet
        st.markdown("### 📝 Détails de l'objet à ajouter")

        col4, col5, col6 = st.columns(3)

        with col4:
            nom_objet = st.text_input("🏷️ Nom de l'objet", 
                                    placeholder="Ex: Matériel urgent, Colis oublié...", 
                                    key="input_nom_objet")

        with col5:
            poids_objet = st.number_input("⚖️ Poids (kg)", 
                                        min_value=0.0, 
                                        max_value=10000.0, 
                                        value=10.0, 
                                        step=0.1, 
                                        key="input_poids_objet")

        with col6:
            volume_objet = st.number_input("📦 Volume (m³)", 
                                        min_value=0.0, 
                                        max_value=100.0, 
                                        value=0.1, 
                                        step=0.01, 
                                        key="input_volume_objet")

        # Bouton d'ajout
        if st.button("➕ Ajouter l'objet au véhicule", type="primary", key="btn_ajouter_objet"):
            
            if not nom_objet.strip():
                st.error("❌ Veuillez donner un nom à l'objet.")
            elif not zone_objet or not vehicule_objet:
                st.error("❌ Veuillez sélectionner une zone et un véhicule.")
            elif poids_objet <= 0 or volume_objet <= 0:
                st.error("❌ Le poids et le volume doivent être supérieurs à 0.")
            else:
                try:
                    # Appel de la méthode add_manual_object
                    with st.spinner("🔄 Ajout de l'objet en cours..."):
                        success, message, df_updated = st.session_state.transfer_manager.add_manual_object(
                            df_voyages=df_voyages,
                            vehicle=vehicule_objet,
                            zone=zone_objet,
                            name=nom_objet.strip(),
                            weight=poids_objet,
                            volume=volume_objet
                        )
                    
                    if success:
                        st.success(message)
                        
                        # Mettre à jour le DataFrame principal dans session_state
                        st.session_state.df_voyages = df_updated
                        st.session_state.transfer_manager.df_voyages = df_updated.copy()
                        
                        # Synchroniser le processeur de location si disponible
                        if 'rental_processor' in st.session_state and st.session_state.rental_processor:
                            try:
                                st.session_state.rental_processor.df_base = df_updated.copy()
                            except Exception as e:
                                st.warning(f"⚠️ Synchronisation partielle du processeur : {str(e)}")
                        
                        # FORCER L'ACTUALISATION
                        st.rerun()
                    else:
                        st.error(message)
                        
                except Exception as e:
                    st.error(f"❌ Erreur lors de l'ajout de l'objet : {str(e)}")

        # Affichage des objets ajoutés récemment
        st.markdown("### 📋 Historique des objets ajoutés")

        # Rechercher les objets manuels dans les BLs
        objets_manuels = []
        for idx, row in df_voyages.iterrows():
            bls = str(row.get("BL inclus", ""))
            if "OBJ-" in bls:
                for bl in bls.split(";"):
                    if bl.startswith("OBJ-"):
                        objets_manuels.append({
                            "Véhicule": row["Véhicule N°"],
                            "Zone": row["Zone"],
                            "Objet": bl,
                            "Poids total chargé": f"{row.get('Poids total chargé', 0):.1f} kg",
                            "Volume total chargé": f"{row.get('Volume total chargé', 0):.3f} m³",
                            "Type Véhicule": row.get("Code Véhicule", "Estafette")
                        })

        if objets_manuels:
            df_objets = pd.DataFrame(objets_manuels)
            st.dataframe(df_objets, use_container_width=True)
            
            # Bouton pour supprimer tous les objets
            if st.button("🗑️ Supprimer tous les objets", type="secondary", key="btn_supprimer_objets"):
                # Réinitialiser les données sans objets manuels
                df_sans_objets = st.session_state.df_voyages.copy()
                for idx, row in df_sans_objets.iterrows():
                    bls_originaux = str(row["BL inclus"]).split(";")
                    bls_filtres = [bl for bl in bls_originaux if not bl.startswith("OBJ-")]
                    df_sans_objets.at[idx, "BL inclus"] = ";".join(bls_filtres)
                
                # Réappliquer la mise à jour
                st.session_state.df_voyages = df_sans_objets
                st.session_state.transfer_manager.df_voyages = df_sans_objets.copy()
                if 'rental_processor' in st.session_state and st.session_state.rental_processor:
                    st.session_state.rental_processor.df_base = df_sans_objets.copy()
                
                st.success("✅ Tous les objets manuels ont été supprimés")
                st.rerun()
        else:
            st.info("ℹ️ Aucun objet manuel ajouté pour le moment.")

 # --- Onglet 4: ✅ VALIDATION DES VOYAGES APRÈS TRANSFERT
    with tab4:
        st.subheader("✅ VALIDATION DES VOYAGES APRÈS TRANSFERT")

    
        # =====================================================
        # 7️⃣ VALIDATION DES VOYAGES APRÈS TRANSFERT
        # =====================================================
        
        from io import BytesIO

        # --- Fonction pour exporter DataFrame en Excel avec arrondi ---
        def to_excel(df, sheet_name="Voyages Validés"):

            df_export = df.copy()
            if "Poids total chargé" in df_export.columns:
                df_export["Poids total chargé"] = df_export["Poids total chargé"].round(3)
            if "Volume total chargé" in df_export.columns:
                df_export["Volume total chargé"] = df_export["Volume total chargé"].round(3)

            output = BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                df_export.to_excel(writer, index=False, sheet_name=sheet_name)
            return output.getvalue()

        # --- CSS pour améliorer l'apparence ---
        st.markdown("""
        <style>
        .voyage-card {
            border: 1px solid #e0e0e0;
            border-radius: 10px;
            padding: 20px;
            margin: 10px 0;
            background: white;
            box-shadow: 0 2px 4px rgba(0,0,0,0.1);
        }
        .voyage-header {
            background: #0369A1;
            color: white;
            padding: 15px;
            border-radius: 8px;
            margin-bottom: 15px;
        }
        .metric-card {
            background: #f8f9fa;
            border-left: 4px solid #0369A1;
            padding: 12px;
            margin: 8px 0;
            border-radius: 5px;
        }
        .bl-list {
            background: #fff3cd;
            border: 1px solid #ffeaa7;
            border-radius: 5px;
            padding: 10px;
            margin: 10px 0;
            max-height: 150px;
            overflow-y: auto;
        }
        .validation-buttons {
            display: flex;
            gap: 10px;
            margin-top: 15px;
        }

        /* AJOUT : Boutons personnalisés */
        .stButton > button[data-testid*="btn_oui"] {
            background-color: #28a745 !important;  /* Vert pour Valider */
            color: white !important;
            border: none !important;
            font-weight: bold !important;
        }

        .stButton > button[data-testid*="btn_non"] {
            background-color: #dc3545 !important;  /* Rouge pour Rejeter */
            color: white !important;
            border: none !important;
            font-weight: bold !important;
        }

        /* Optionnel : Effet au survol */
        .stButton > button[data-testid*="btn_oui"]:hover {
            background-color: #218838 !important;
        }

        .stButton > button[data-testid*="btn_non"]:hover {
            background-color: #c82333 !important;
        }
        </style>
        """, unsafe_allow_html=True)

        # --- Création du DataFrame de validation à partir du df_voyages ---
        if "df_voyages" in st.session_state:
            voyages_apres_transfert = st.session_state.df_voyages.copy()
            df_validation = voyages_apres_transfert.copy()

            if "validations" not in st.session_state:
                st.session_state.validations = {}

            # --- Affichage amélioré des voyages ---
            st.markdown("### 📋 Liste des Voyages à Valider")
            
            for idx, row in df_validation.iterrows():
                # Création d'une carte pour chaque voyage
                with st.container():
                    st.markdown(f"""
                    <div class="voyage-card">
                        <div class="voyage-header">
                            <h4>🚚 Voyage {row['Véhicule N°']} | Zone: {row['Zone']}</h4>
                        </div>
                    """, unsafe_allow_html=True)
                    
                    # Métriques principales
                    col1, col2, col3 = st.columns(3)
                    
                    with col1:
                        st.markdown(f"""
                        <div class="metric-card">
                            <strong>⚖️ Poids Total</strong><br>
                            {row['Poids total chargé']:.3f} kg
                        </div>
                        """, unsafe_allow_html=True)
                    
                    with col2:
                        st.markdown(f"""
                        <div class="metric-card">
                            <strong>📏 Volume Total</strong><br>
                            {row['Volume total chargé']:.3f} m³
                        </div>
                        """, unsafe_allow_html=True)
                    
                    with col3:
                        taux_occupation = row.get('Taux d\'occupation (%)', 'N/A')
                        if taux_occupation != 'N/A':
                            taux_text = f"{taux_occupation:.1f}%"
                        else:
                            taux_text = "N/A"
                        st.markdown(f"""
                        <div class="metric-card">
                            <strong>📊 Taux d'Occupation</strong><br>
                            {taux_text}
                        </div>
                        """, unsafe_allow_html=True)
                    
                    # Informations détaillées
                    col4, col5 = st.columns(2)
                    
                    with col4:
                        clients = row.get('Client(s) inclus', '')
                        if clients:
                            st.markdown(f"**👥 Clients:** {clients}")
                        
                        representants = row.get('Représentant(s) inclus', '')
                        if representants:
                            st.markdown(f"**👨‍💼 Représentants:** {representants}")
                    
                    with col5:
                        location = "✅ Oui" if row.get('Location_camion') else "❌ Non"
                        st.markdown(f"**🚛 Location:** {location}")
                        
                        code_vehicule = row.get('Code Véhicule', 'N/A')
                        st.markdown(f"**🔧 Code Véhicule:** {code_vehicule}")
                    
                    # Liste des BL avec défilement
                    bls = row.get('BL inclus', '')
                    if bls:
                        bls_list = bls.split(';')
                        bls_html = "<br>".join([f"• {bl.strip()}" for bl in bls_list])
                        st.markdown(f"""
                        <div class="bl-list">
                            <strong>📋 BLs Inclus ({len(bls_list)}):</strong><br>
                            {bls_html}
                        </div>
                        """, unsafe_allow_html=True)
                    
                    # Boutons de validation côte à côte
                    st.markdown("**✅ Validation du voyage:**")
                    col_oui, col_non = st.columns(2)
                    
                    with col_oui:
                        if st.button(f"✅ Valider {row['Véhicule N°']}", key=f"btn_oui_{idx}", 
                                use_container_width=True, type="primary" if st.session_state.validations.get(idx) == "Oui" else "secondary"):
                            st.session_state.validations[idx] = "Oui"
                            st.rerun()
                    
                    with col_non:
                        if st.button(f"❌ Rejeter {row['Véhicule N°']}", key=f"btn_non_{idx}",
                                use_container_width=True, type="primary" if st.session_state.validations.get(idx) == "Non" else "secondary"):
                            st.session_state.validations[idx] = "Non"
                            st.rerun()
                    
                    # Afficher le statut actuel
                    statut = st.session_state.validations.get(idx)
                    if statut == "Oui":
                        st.success(f"✅ Voyage {row['Véhicule N°']} validé")
                    elif statut == "Non":
                        st.error(f"❌ Voyage {row['Véhicule N°']} rejeté")
                    else:
                        st.info("⏳ En attente de validation")
                    
                    st.markdown("</div>", unsafe_allow_html=True)
                    st.markdown("---")

            # --- Résumé des validations ---
            st.markdown("### 📊 Résumé des Validations")
            total_voyages = len(df_validation)
            valides = sum(1 for v in st.session_state.validations.values() if v == "Oui")
            rejetes = sum(1 for v in st.session_state.validations.values() if v == "Non")

            col1, col2, col3 = st.columns(3)

            with col1:
                st.metric("Total Voyages", total_voyages)
            with col2:
                st.metric("✅ Validés", valides)
            with col3:
                st.metric("❌ Rejetés", rejetes)

            # Information supplémentaire sur l'état des validations
            if valides + rejetes < total_voyages:
                st.info(f"ℹ️ {total_voyages - (valides + rejetes)} voyage(s) n'ont pas encore été validés")

            # --- Bouton pour appliquer les validations ---
            if st.button("🚀 Finaliser la Validation", type="primary", use_container_width=True):
                valid_indexes = [i for i, v in st.session_state.validations.items() if v == "Oui"]
                valid_indexes = [i for i in valid_indexes if i in df_validation.index]

                if valid_indexes:
                    df_voyages_valides = df_validation.loc[valid_indexes].reset_index(drop=True)
                    st.session_state.df_voyages_valides = df_voyages_valides

                    st.success(f"✅ {len(df_voyages_valides)} voyage(s) validé(s) avec succès!")
                    
                    # Affichage des voyages validés
                    st.markdown("### 🎉 Voyages Validés - Résumé Final")
                    
                    for idx, row_valide in df_voyages_valides.iterrows():
                        with st.expander(f"🚚 {row_valide['Véhicule N°']} - Zone {row_valide['Zone']}", expanded=True):
                            col1, col2 = st.columns(2)
                            with col1:
                                st.metric("Poids", f"{row_valide['Poids total chargé']:.3f} kg")
                                st.metric("Clients", row_valide.get('Client(s) inclus', 'N/A'))
                            with col2:
                                st.metric("Volume", f"{row_valide['Volume total chargé']:.3f} m³")
                                st.metric("Représentants", row_valide.get('Représentant(s) inclus', 'N/A'))

                    # --- Export Excel ---
                    excel_data = to_excel(df_voyages_valides)
                    st.download_button(
                        label="💾 Télécharger les voyages validés (XLSX)",
                        data=excel_data,
                        file_name="Voyages_valides.xlsx",
                        mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                        use_container_width=True
                    )
                else:
                    st.warning("⚠️ Aucun voyage n'a été validé. Veuillez valider au moins un voyage.")

        else:
            st.warning("⚠️ Vous devez d'abord exécuter la section 4 (Voyages par Estafette Optimisé).")


   # --- Onglet 5: 🚛 Attribution des véhicules et chauffeurs ---
    with tab5:
        st.subheader("🚛 Attribution des véhicules et chauffeurs")
        
        if 'df_voyages_valides' in st.session_state and st.session_state.df_voyages_valides is not None and not st.session_state.df_voyages_valides.empty:

            df_attribution = st.session_state.df_voyages_valides.copy()

            # Fonction pour formatter les colonnes avec retours à ligne POUR STREAMLIT
            def formater_colonnes_listes_streamlit(df):
                df_formate = df.copy()
                colonnes_a_formater = ['Client(s) inclus', 'Représentant(s) inclus', 'BL inclus']
                
                for col in colonnes_a_formater:
                    if col in df_formate.columns:
                        df_formate[col] = df_formate[col].apply(
                            lambda x: '\n'.join([elem.strip() for elem in str(x).replace(';', ',').split(',') if elem.strip()]) 
                            if pd.notna(x) else ""
                        )
                return df_formate

            if "attributions" not in st.session_state:
                st.session_state.attributions = {}

            for idx, row in df_attribution.iterrows():
                with st.expander(f"🚚 Voyage {row['Véhicule N°']} | Zone : {row['Zone']}"):
                    st.write("**Informations du voyage :**")
                    
                    # Créer un affichage personnalisé avec retours à ligne
                    col1, col2, col3 = st.columns(3)
                    
                    with col1:
                        st.write(f"**Zone:** {row['Zone']}")
                        st.write(f"**Véhicule N°:** {row['Véhicule N°']}")
                        if "Poids total chargé" in row:
                            st.write(f"**Poids total chargé:** {row['Poids total chargé']:.2f} kg")
                        if "Volume total chargé" in row:
                            st.write(f"**Volume total chargé:** {row['Volume total chargé']:.3f} m³")
                        if "Taux d'occupation (%)" in row:
                            st.write(f"**Taux d'occupation:** {row['Taux d\'occupation (%)']:.1f}%")
                    
                    with col2:
                        # Afficher les clients avec retours à ligne
                        if 'Client(s) inclus' in row and pd.notna(row['Client(s) inclus']):
                            st.write("**Clients:**")
                            clients = str(row['Client(s) inclus']).replace(';', ',').split(',')
                            for client in clients:
                                client_clean = client.strip()
                                if client_clean:
                                    st.write(f"- {client_clean}")
                        
                        # Afficher les représentants avec retours à ligne
                        if 'Représentant(s) inclus' in row and pd.notna(row['Représentant(s) inclus']):
                            st.write("**Représentants:**")
                            representants = str(row['Représentant(s) inclus']).replace(';', ',').split(',')
                            for rep in representants:
                                rep_clean = rep.strip()
                                if rep_clean:
                                    st.write(f"- {rep_clean}")
                    
                    with col3:
                        # Afficher les BL avec retours à ligne
                        if 'BL inclus' in row and pd.notna(row['BL inclus']):
                            st.write("**BL associés:**")
                            bls = str(row['BL inclus']).replace(';', ',').split(',')
                            for bl in bls:
                                bl_clean = bl.strip()
                                if bl_clean:
                                    st.write(f"- {bl_clean}")

                    col_veh, col_chauf = st.columns(2)
                    
                    with col_veh:
                        vehicule_selectionne = st.selectbox(
                            f"Véhicule pour le voyage {row['Véhicule N°']}",
                            VEHICULES_DISPONIBLES,
                            index=0 if st.session_state.attributions.get(idx, {}).get("Véhicule") else 0,
                            key=f"vehicule_{idx}"
                        )
                    
                    with col_chauf:
                        options_chauffeurs = [f"{matricule} - {nom}" for matricule, nom in CHAUFFEURS_DETAILS.items() if matricule != 'Matricule']
                        
                        default_index = 0
                        chauffeur_actuel = st.session_state.attributions.get(idx, {}).get("Chauffeur_complet")
                        if chauffeur_actuel and chauffeur_actuel in options_chauffeurs:
                            default_index = options_chauffeurs.index(chauffeur_actuel)
                        
                        chauffeur_selectionne_complet = st.selectbox(
                            f"Chauffeur pour le voyage {row['Véhicule N°']}",
                            options_chauffeurs,
                            index=default_index,
                            key=f"chauffeur_{idx}"
                        )
                        
                        if chauffeur_selectionne_complet:
                            matricule_chauffeur = chauffeur_selectionne_complet.split(" - ")[0]
                            nom_chauffeur = chauffeur_selectionne_complet.split(" - ")[1]
                        else:
                            matricule_chauffeur = ""
                            nom_chauffeur = ""

                    st.session_state.attributions[idx] = {
                        "Véhicule": vehicule_selectionne,
                        "Chauffeur_complet": chauffeur_selectionne_complet,
                        "Matricule_chauffeur": matricule_chauffeur,
                        "Nom_chauffeur": nom_chauffeur
                    }

            if st.button("✅ Appliquer les attributions", key="btn_apply_attributions"):

                df_attribution["Véhicule attribué"] = df_attribution.index.map(lambda i: st.session_state.attributions[i]["Véhicule"])
                df_attribution["Chauffeur attribué"] = df_attribution.index.map(lambda i: st.session_state.attributions[i]["Nom_chauffeur"])
                df_attribution["Matricule chauffeur"] = df_attribution.index.map(lambda i: st.session_state.attributions[i]["Matricule_chauffeur"])

                
                st.markdown("### 📦 Voyages avec Véhicule et Chauffeur")

                # --- Affichage Streamlit amélioré avec retours à ligne ---
                for idx, row in df_attribution.iterrows():
                    with st.expander(f"📋 Voyage {row['Véhicule N°']} - Zone {row['Zone']} - Véhicule: {row.get('Véhicule attribué', 'N/A')} - Chauffeur: {row.get('Chauffeur attribué', 'N/A')}"):
                        col1, col2, col3 = st.columns(3)
                        
                        with col1:
                            st.write("**Informations de base:**")
                            st.write(f"**Zone:** {row['Zone']}")
                            st.write(f"**Véhicule N°:** {row['Véhicule N°']}")
                            if "Poids total chargé" in row:
                                st.write(f"**Poids total chargé:** {row['Poids total chargé']:.3f} kg")
                            if "Volume total chargé" in row:
                                st.write(f"**Volume total chargé:** {row['Volume total chargé']:.3f} m³")
                            if "Taux d'occupation (%)" in row:
                                st.write(f"**Taux d'occupation:** {row['Taux d\'occupation (%)']:.3f}%")
                            if "Véhicule attribué" in row:
                                st.write(f"**Véhicule attribué:** {row['Véhicule attribué']}")
                            if "Chauffeur attribué" in row:
                                st.write(f"**Chauffeur attribué:** {row['Chauffeur attribué']}")
                            if "Matricule chauffeur" in row:
                                st.write(f"**Matricule chauffeur:** {row['Matricule chauffeur']}")
                        
                        with col2:
                            # Afficher les clients avec retours à ligne
                            if 'Client(s) inclus' in row and pd.notna(row['Client(s) inclus']):
                                st.write("**📋 Clients inclus:**")
                                clients = str(row['Client(s) inclus']).replace(';', ',').split(',')
                                for client in clients:
                                    client_clean = client.strip()
                                    if client_clean:
                                        st.write(f"• {client_clean}")
                            
                            # Afficher les représentants avec retours à ligne
                            if 'Représentant(s) inclus' in row and pd.notna(row['Représentant(s) inclus']):
                                st.write("**👤 Représentants inclus:**")
                                representants = str(row['Représentant(s) inclus']).replace(';', ',').split(',')
                                for rep in representants:
                                    rep_clean = rep.strip()
                                    if rep_clean:
                                        st.write(f"• {rep_clean}")
                        
                        with col3:
                            # Afficher les BL avec retours à ligne
                            if 'BL inclus' in row and pd.notna(row['BL inclus']):
                                st.write("**📄 BL associés:**")
                                bls = str(row['BL inclus']).replace(';', ',').split(',')
                                # Afficher en colonnes si beaucoup de BL
                                if len(bls) > 5:
                                    cols = st.columns(2)
                                    half = len(bls) // 2
                                    for i, bl in enumerate(bls):
                                        bl_clean = bl.strip()
                                        if bl_clean:
                                            col_idx = 0 if i < half else 1
                                            with cols[col_idx]:
                                                st.write(f"• {bl_clean}")
                                else:
                                    for bl in bls:
                                        bl_clean = bl.strip()
                                        if bl_clean:
                                            st.write(f"• {bl_clean}")

                # --- Export Excel avec retours à ligne et CENTRAGE ---
                from io import BytesIO
                import openpyxl

                def to_excel(df):
                    df_export = df.copy()
                    
                    # Formater les colonnes avec retours à ligne pour Excel
                    colonnes_a_formater = ['Client(s) inclus', 'Représentant(s) inclus', 'BL inclus']
                    for col in colonnes_a_formater:
                        if col in df_export.columns:
                            df_export[col] = df_export[col].apply(
                                lambda x: '\n'.join([elem.strip() for elem in str(x).replace(';', ',').split(',') if elem.strip()]) 
                                if pd.notna(x) else ""
                            )
                    
                    if "Poids total chargé" in df_export.columns:
                        df_export["Poids total chargé"] = df_export["Poids total chargé"].round(3)
                    if "Volume total chargé" in df_export.columns:
                        df_export["Volume total chargé"] = df_export["Volume total chargé"].round(3)
                    
                    output = BytesIO()
                    with pd.ExcelWriter(output, engine='openpyxl') as writer:
                        df_export.to_excel(writer, index=False, sheet_name='Voyages_Attribués')
                        
                        # Appliquer le formatage des retours à ligne et CENTRAGE dans Excel
                        workbook = writer.book
                        worksheet = writer.sheets['Voyages_Attribués']
                        
                        # Style de centrage avec retours à ligne
                        center_alignment = openpyxl.styles.Alignment(
                            horizontal='center', 
                            vertical='center', 
                            wrap_text=True
                        )
                        
                        # Appliquer le centrage à TOUTES les cellules
                        for row in worksheet.iter_rows(min_row=1, max_row=len(df_export) + 1, min_col=1, max_col=len(df_export.columns)):
                            for cell in row:
                                cell.alignment = center_alignment
                        
                        # Ajuster automatiquement la largeur des colonnes
                        for column in worksheet.columns:
                            max_length = 0
                            column_letter = column[0].column_letter
                            for cell in column:
                                try:
                                    if cell.value:
                                        # Calculer la longueur maximale en prenant en compte les retours à ligne
                                        lines = str(cell.value).split('\n')
                                        max_line_length = max(len(line) for line in lines)
                                        max_length = max(max_length, max_line_length)
                                except:
                                    pass
                            adjusted_width = min(50, (max_length + 2))  # Limiter à 50 caractères max
                            worksheet.column_dimensions[column_letter].width = adjusted_width
                        
                        # Ajuster la hauteur des lignes pour les retours à ligne
                        for row in range(2, len(df_export) + 2):  # Commencer à la ligne 2 (après l'en-tête)
                            worksheet.row_dimensions[row].height = 60  # Hauteur fixe pour accommoder les retours à ligne
                    
                    return output.getvalue()

                # --- Export PDF avec tableau ÉLARGI et ESPACES MINIMISÉS ---
                from fpdf import FPDF

                def to_pdf_better_centered(df, title="Voyages Attribués"):
                    pdf = FPDF(orientation='L')  # Paysage pour plus d'espace
                    pdf.add_page()
                    
                    # RÉDUCTION des marges pour utiliser TOUTE la largeur
                    pdf.set_left_margin(5)   # Marge gauche réduite
                    pdf.set_right_margin(5)  # Marge droite réduite
                    pdf.set_top_margin(10)   # Marge haut réduite
                    
                    # Titre PLUS PETIT et PLUS HAUT
                    pdf.set_font("Arial", 'B', 14)  # Taille réduite
                    pdf.cell(0, 8, title, ln=True, align="C")  # Hauteur réduite
                    pdf.ln(3)  # Espacement réduit après le titre
                    
                    # Créer une copie formatée pour le PDF
                    df_pdf = df.copy()
                    
                    # Formater les nombres avec 3 chiffres après la virgule SAUF le taux avec 2 chiffres
                    numeric_columns = {
                        'Poids total chargé': ('kg', 3),
                        'Volume total chargé': ('m³', 3), 
                        'Taux d\'occupation (%)': ('%', 2)  # 2 chiffres après la virgule
                    }
                    
                    for col, (unit, decimals) in numeric_columns.items():
                        if col in df_pdf.columns:
                            df_pdf[col] = df_pdf[col].apply(
                                lambda x: f"{float(x):.{decimals}f} {unit}" if x and str(x).strip() and str(x).strip() != 'nan' else ""
                            )
                    
                    # Configuration des colonnes AVEC LARGEURS MAXIMALISÉES
                    col_config = {
                        'Zone': {'width': 15, 'header': 'Zone'},
                        'Véhicule N°': {'width': 18, 'header': 'Véhicule'},
                        'Poids total chargé': {'width': 22, 'header': 'Poids (kg)'},
                        'Volume total chargé': {'width': 22, 'header': 'Volume (m³)'},
                        'Client(s) inclus': {'width': 30, 'header': 'Clients'},
                        'Représentant(s) inclus': {'width': 30, 'header': 'Représentants'},
                        'BL inclus': {'width': 35, 'header': 'BL associés'},
                        'Taux d\'occupation (%)': {'width': 18, 'header': 'Taux %'},
                        'Véhicule attribué': {'width': 25, 'header': 'Véhicule Attribué'},
                        'Chauffeur attribué': {'width': 25, 'header': 'Chauffeur'},
                        'Matricule chauffeur': {'width': 20, 'header': 'Matricule'}
                    }
                    
                    # Sélectionner seulement les colonnes existantes
                    colonnes_existantes = [col for col in df_pdf.columns if col in col_config]
                    widths = [col_config[col]['width'] for col in colonnes_existantes]
                    headers = [col_config[col]['header'] for col in colonnes_existantes]
                    
                    # Calculer la position de départ - DÉBUT PLUS À GAUCHE
                    total_width = sum(widths)
                    page_width = 297  # Largeur d'une page A4 en paysage (mm)
                    start_x = 5  # Commencer presque au bord gauche
                    
                    # Positionner le tableau AU DÉBUT
                    pdf.set_x(start_x)
                    
                    # En-têtes CENTRÉS avec police PLUS PETITE
                    pdf.set_font("Arial", 'B', 8)  # Taille réduite
                    for i, header in enumerate(headers):
                        pdf.cell(widths[i], 6, header, border=1, align='C')  # Hauteur réduite
                    pdf.ln()
                    
                    # Données avec centrage VERTICAL et HORIZONTAL
                    pdf.set_font("Arial", '', 7)  # Taille réduite pour les données
                    
                    for voyage_idx, (_, row) in enumerate(df_pdf.iterrows()):
                        # Vérifier si on dépasse la hauteur de page
                        if pdf.get_y() > 180:  # Si on approche du bas de page
                            pdf.add_page()  # Nouvelle page
                            pdf.set_x(start_x)
                            # Ré-afficher les en-têtes sur la nouvelle page
                            pdf.set_font("Arial", 'B', 8)
                            for i, header in enumerate(headers):
                                pdf.cell(widths[i], 6, header, border=1, align='C')
                            pdf.ln()
                            pdf.set_font("Arial", '', 7)
                        
                        # Déterminer le nombre de lignes nécessaires pour ce voyage
                        list_columns = ['Client(s) inclus', 'Représentant(s) inclus', 'BL inclus']
                        non_list_columns = [col for col in colonnes_existantes if col not in list_columns]
                        
                        max_lines = 1
                        list_contents = {}
                        
                        for col in list_columns:
                            if col in colonnes_existantes:
                                content = str(row[col]) if pd.notna(row[col]) and str(row[col]) != 'nan' else ""
                                elements = content.replace(';', ',').split(',')
                                elements = [elem.strip() for elem in elements if elem.strip()]
                                list_contents[col] = elements
                                max_lines = max(max_lines, len(elements))
                        
                        # Pour chaque ligne du voyage
                        for line_idx in range(max_lines):
                            # Vérifier si on dépasse la hauteur de page pour cette ligne
                            if pdf.get_y() > 190:  # Si on approche vraiment du bas
                                pdf.add_page()
                                pdf.set_x(start_x)
                                pdf.set_font("Arial", 'B', 8)
                                for i, header in enumerate(headers):
                                    pdf.cell(widths[i], 6, header, border=1, align='C')
                                pdf.ln()
                                pdf.set_font("Arial", '', 7)
                            
                            # Positionner au DÉBUT pour chaque ligne
                            pdf.set_x(start_x)
                            
                            for i, col in enumerate(colonnes_existantes):
                                if col in list_columns:
                                    # Colonnes de liste - afficher élément par élément
                                    elements = list_contents.get(col, [])
                                    content = elements[line_idx] if line_idx < len(elements) else ""
                                else:
                                    # Colonnes non-liste - afficher sur la première ligne seulement
                                    if line_idx == 0:
                                        content = str(row[col]) if pd.notna(row[col]) and str(row[col]) != 'nan' else ""
                                    else:
                                        content = ""
                                
                                # Bordures avec hauteur RÉDUITE
                                border = 'LR'
                                if line_idx == 0: border += 'T'
                                if line_idx == max_lines - 1: border += 'B'
                                if i == 0: border += 'L'
                                if i == len(colonnes_existantes) - 1: border += 'R'
                                
                                pdf.cell(widths[i], 5, content, border=border, align='C')  # Hauteur réduite à 5
                            
                            pdf.ln()
                    
                    return pdf.output(dest='S').encode('latin-1')
                
                # Afficher les boutons de téléchargement côte à côte
                col1, col2 = st.columns(2)

                with col1:
                    st.download_button(
                        label="💾 Télécharger le tableau final (XLSX)",
                        data=to_excel(df_attribution),
                        file_name="Voyages_attribues.xlsx",
                        mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                        key="btn_download_excel"
                    )

                with col2:
                    st.download_button(
                        label="📄 Télécharger le tableau final (PDF)",
                        data=to_pdf_better_centered(df_attribution),
                        file_name="Voyages_attribues.pdf",
                        mime='application/pdf',
                        key="btn_download_pdf"
                    )
                        
                # Mettre à jour le session state
                st.session_state.df_voyages_valides = df_attribution
                st.success("✅ Attributions appliquées avec succès !")
                
        else:
            st.warning("⚠️ Vous devez d'abord valider les voyages dans la section 7.")
        
        # =====================================================
        # SECTION: 📋Rapports et analyses avancés
        # =====================================================
        st.markdown("---")
        st.markdown("## 📋 Rapports et analyses avancés")
        
        if "df_voyages" in st.session_state and "df_livraisons_original" in st.session_state:
            
            try:
                from backend import AdvancedReportGenerator
                report_generator = AdvancedReportGenerator(
                    st.session_state.df_voyages, 
                    st.session_state.df_livraisons_original
                )
                
                tab_analytics1, tab_analytics2, tab_analytics3, tab_analytics4 = st.tabs([
                    "📈 Rapport Analytique", 
                    "👤 Rapport Client", 
                    "💰 Analyse Coûts", 
                    "✅ Validation Données"
                ])
                
                with tab_analytics1:
                    st.subheader("Rapport Analytique Complet")
                    if st.button("🔄 Générer le rapport analytique", key="btn_rapport_analytique"):
                        with st.spinner("Génération du rapport en cours..."):
                            rapport = report_generator.generer_rapport_analytique()
                            st.text_area("Rapport détaillé", rapport, height=400, key="text_rapport_analytique")
                
                with tab_analytics2:
                    st.subheader("Rapport Spécifique Client")
                    clients_disponibles = sorted(st.session_state.df_livraisons_original["Client de l'estafette"].unique())
                    client_rapport = st.selectbox("Sélectionner un client", clients_disponibles, key="select_client_rapport")
                    
                    if st.button("📋 Générer rapport client", key="btn_rapport_client"):
                        with st.spinner("Génération du rapport client..."):
                            rapport_client = report_generator.generer_rapport_client(client_rapport)
                            st.text_area(f"Rapport pour {client_rapport}", rapport_client, height=300, key="text_rapport_client")
                
                with tab_analytics3:
                    st.subheader("Analyse des Coûts")
                    col_cost1, col_cost2 = st.columns(2)
                    
                    with col_cost1:
                        cout_estafette = st.number_input("Coût unitaire estafette (TND)", value=150, min_value=50, max_value=500, key="cout_estafette")
                    with col_cost2:
                        cout_camion = st.number_input("Coût unitaire camion (TND)", value=800, min_value=300, max_value=2000, key="cout_camion")
                    
                    if st.button("💰 Calculer les coûts", key="btn_calcul_couts"):
                        try:
                            from backend import calculer_couts_estimation
                            couts = calculer_couts_estimation(
                                st.session_state.df_voyages, 
                                cout_estafette, 
                                cout_camion
                            )
                            
                            if 'erreur' not in couts:
                                st.success(couts['cout_estimation'])
                                
                                # Graphique des coûts
                                import plotly.express as px
                                df_couts = pd.DataFrame({
                                    'Type': ['Estafettes', 'Camions'],
                                    'Coût Total (TND)': [
                                        couts['estafettes'] * couts['cout_estafette_unitaire'],
                                        couts['camions'] * couts['cout_camion_unitaire']
                                    ]
                                })
                                
                                fig = px.pie(df_couts, values='Coût Total (TND)', names='Type', 
                                            title='Répartition des coûts par type de véhicule')
                                st.plotly_chart(fig, use_container_width=True)
                            else:
                                st.error(couts['erreur'])
                        except ImportError:
                            st.error("❌ La fonction 'calculer_couts_estimation' n'est pas disponible dans le backend")
                        except Exception as e:
                            st.error(f"❌ Erreur lors du calcul des coûts : {str(e)}")
                
                with tab_analytics4:
                    st.subheader("Validation d'Intégrité des Données")
                    if st.button("🔍 Vérifier l'intégrité des données", key="btn_verif_integrite"):
                        try:
                            from backend import verifier_integrite_donnees
                            resultat_validation = verifier_integrite_donnees(
                                st.session_state.df_voyages,
                                st.session_state.df_livraisons_original
                            )
                            
                            if "✅" in resultat_validation:
                                st.success(resultat_validation)
                            else:
                                st.warning(resultat_validation)
                        except ImportError:
                            st.error("❌ La fonction 'verifier_integrite_donnees' n'est pas disponible dans le backend")
                        except Exception as e:
                            st.error(f"❌ Erreur lors de la validation : {str(e)}")
            
            except ImportError:
                st.error("❌ La classe 'AdvancedReportGenerator' n'est pas disponible dans le backend")
                st.info("""
                ℹ️ Pour utiliser les rapports avancés, assurez-vous que :
                1. Le fichier 'backend.py' existe
                2. Il contient la classe 'AdvancedReportGenerator'
                3. Les fonctions nécessaires sont implémentées
                """)
            except Exception as e:
                st.error(f"❌ Erreur lors de l'initialisation du générateur de rapports : {str(e)}")

        else:
            st.warning("⚠️ Vous devez d'abord traiter les données dans la page d'importation.")
        
        # =====================================================
        # SECTION: 🏷️ GÉNÉRATION DES CODES VOYAGE
        # =====================================================
        st.markdown("---")
        st.markdown("## 🏷️ GÉNÉRATION DES CODES VOYAGE")
        
        if "df_voyages_valides" in st.session_state and st.session_state.df_voyages_valides is not None and not st.session_state.df_voyages_valides.empty:
            
            df_final = st.session_state.df_voyages_valides.copy()
            
            # Configuration des paramètres de génération
            col1, col2, col3 = st.columns(3)
            
            with col1:
                date_voyage = st.date_input(
                    "📅 Date de voyage",
                    value=pd.Timestamp.now().date(),
                    help="Date prévue pour les livraisons",
                    key="date_voyage_input"
                )
            
            with col2:
                numero_debut = st.number_input(
                    "🔢 Numéro séquentiel de départ",
                    min_value=1,
                    max_value=1000,
                    value=1,
                    help="Numéro de départ pour la séquence",
                    key="numero_debut_input"
                )
            
            with col3:
                st.markdown("<br>", unsafe_allow_html=True)
                generer_codes = st.button("🏷️ Générer les codes voyage", type="primary", key="btn_generer_codes")
            
            # CSS pour les codes voyage
            st.markdown("""
            <style>
            /* Style pour le tableau des codes voyage */
            .codes-table {
                width: 100%;
                border-collapse: collapse;
                font-family: Arial, sans-serif;
                font-size: 14px;
                box-shadow: 0 2px 8px rgba(0,0,0,0.1);
                border-radius: 8px;
                overflow: hidden;
                margin: 1rem 0;
            }
            
            .codes-table th {
                background-color: #0369A1;
                color: white;
                padding: 12px 8px;
                text-align: center;
                border: 2px solid #4682B4;
                font-weight: normal;
                font-size: 13px;
                vertical-align: middle;
            }
            
            .codes-table td {
                padding: 10px 8px;
                text-align: center;
                border: 1px solid #B0C4DE;
                background-color: white;
                color: #000000;
                vertical-align: middle;
                font-weight: normal;
            }
            
            .codes-container {
                overflow-x: auto;
                margin: 1rem 0;
                border-radius: 8px;
                border: 2px solid #4682B4;
            }
            
            .codes-table tr:hover td {
                background-color: #F0F8FF !important;
            }
            
            .codes-table tr:nth-child(even) td {
                background-color: #f9f9f9;
            }
            
            /* Style spécifique pour la colonne Code voyage */
            .code-voyage-cell {
                font-family: 'Courier New', monospace;
                font-weight: bold;
                color: #0369A1;
            }
            </style>
            """, unsafe_allow_html=True)
            
            if generer_codes:
                try:
                    # Préparation des données pour le code voyage
                    df_final['Date Voyage Format'] = date_voyage.strftime('%Y%m%d')
                    
                    # Création du numéro séquentiel pour chaque voyage
                    df_final['Numero Séquentiel'] = range(numero_debut, numero_debut + len(df_final))
                    df_final['Numero Séquentiel Formatted'] = df_final['Numero Séquentiel'].apply(lambda x: f"{x:03d}")
                    
                    # Création du Code voyage
                    df_final['Code voyage'] = (
                        df_final['Véhicule N°'].astype(str) + '/' +
                        df_final['Date Voyage Format'].astype(str) + '/' +
                        df_final['Numero Séquentiel Formatted'].astype(str)
                    )
                    
                    # Mettre à jour le session state
                    st.session_state.df_voyages_valides = df_final
                    
                    st.success(f"✅ {len(df_final)} codes voyage générés avec succès !")
                    
                    # APERÇU DES CODES VOYAGE
                    st.markdown("### 📋 Aperçu des codes voyage générés")
                    
                    # Afficher un aperçu des codes générés
                    df_apercu = df_final[['Véhicule N°', 'Zone', 'Code voyage']].copy()
                    
                    # Appliquer le style spécial pour la colonne Code voyage
                    df_apercu_display = df_apercu.copy()
                    df_apercu_display['Code voyage'] = df_apercu_display['Code voyage'].apply(
                        lambda x: f'<span class="code-voyage-cell">{x}</span>'
                    )
                    
                    # Convertir en HTML avec style
                    html_table = df_apercu_display.to_html(
                        index=False,
                        classes="codes-table",
                        border=0,
                        escape=False
                    )
                    
                    # Afficher le tableau stylisé
                    st.markdown(f"""
                    <div class="codes-container">
                        {html_table}
                    </div>
                    """, unsafe_allow_html=True)
                    
                except Exception as e:
                    st.error(f"❌ Erreur lors de la génération des codes voyage : {str(e)}")
            
            # Afficher les codes existants si déjà générés
            elif 'Code voyage' in df_final.columns:
                st.success("✅ Codes voyage déjà générés")
                
                # AFFICHAGE DES CODES EXISTANTS
                st.markdown("### 📋 Codes voyage générés")
                
                df_apercu = df_final[['Véhicule N°', 'Zone', 'Code voyage']].copy()
                
                # Appliquer le style spécial pour la colonne Code voyage
                df_apercu_display = df_apercu.copy()
                df_apercu_display['Code voyage'] = df_apercu_display['Code voyage'].apply(
                    lambda x: f'<span class="code-voyage-cell">{x}</span>'
                )
                
                # Convertir en HTML
                html_table = df_apercu_display.to_html(
                    index=False,
                    classes="codes-table",
                    border=0,
                    escape=False
                )
                
                # Afficher le tableau
                st.markdown(f"""
                <div class="codes-container">
                    {html_table}
                </div>
                """, unsafe_allow_html=True)
                
                # Option pour regénérer les codes
                st.markdown("<br>", unsafe_allow_html=True)
                
                col_regen1, col_regen2, col_regen3 = st.columns([1, 2, 1])
                with col_regen2:
                    if st.button("🔄 Regénérer les codes voyage", use_container_width=True, type="secondary", key="btn_regen_codes"):
                        columns_to_remove = ['Code voyage', 'Date Voyage Format', 'Numero Séquentiel', 'Numero Séquentiel Formatted']
                        for col in columns_to_remove:
                            if col in df_final.columns:
                                del df_final[col]
                        st.session_state.df_voyages_valides = df_final
                        st.rerun()
            
            # Bouton d'export avec les codes
            if 'Code voyage' in df_final.columns:
                st.markdown("---")
                st.markdown("### 💾 Exporter avec codes voyage")
                
                col_export1, col_export2 = st.columns(2)
                
                with col_export1:
                    # Fonction d'export Excel
                    def to_excel_with_codes(df):
                        output = BytesIO()
                        with pd.ExcelWriter(output, engine='openpyxl') as writer:
                            df.to_excel(writer, index=False, sheet_name='Voyages_avec_Codes')
                        return output.getvalue()
                    
                    excel_data = to_excel_with_codes(df_final)
                    st.download_button(
                        label="📥 Télécharger Excel avec codes",
                        data=excel_data,
                        file_name="voyages_avec_codes.xlsx",
                        mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                        use_container_width=True,
                        key="btn_download_excel_codes"
                    )
                
                with col_export2:
                    # Export PDF avec codes (optionnel)
                    from fpdf import FPDF
                    
                    def to_pdf_with_codes(df):
                        pdf = FPDF(orientation='L')
                        pdf.add_page()
                        pdf.set_font("Arial", 'B', 16)
                        pdf.cell(0, 10, "Liste des voyages avec codes", ln=True, align='C')
                        pdf.ln(10)
                        
                        pdf.set_font("Arial", '', 10)
                        # En-têtes
                        headers = ['Véhicule', 'Zone', 'Code Voyage', 'Chauffeur']
                        col_widths = [40, 40, 80, 60]
                        
                        for i, header in enumerate(headers):
                            pdf.cell(col_widths[i], 10, header, border=1)
                        pdf.ln()
                        
                        # Données
                        pdf.set_font("Arial", '', 9)
                        for _, row in df.iterrows():
                            pdf.cell(col_widths[0], 10, str(row.get('Véhicule N°', '')), border=1)
                            pdf.cell(col_widths[1], 10, str(row.get('Zone', '')), border=1)
                            pdf.cell(col_widths[2], 10, str(row.get('Code voyage', '')), border=1)
                            pdf.cell(col_widths[3], 10, str(row.get('Chauffeur attribué', '')), border=1)
                            pdf.ln()
                        
                        return pdf.output(dest='S').encode('latin-1')
                    
                    try:
                        pdf_data = to_pdf_with_codes(df_final)
                        st.download_button(
                            label="📄 Télécharger PDF avec codes",
                            data=pdf_data,
                            file_name="voyages_avec_codes.pdf",
                            mime='application/pdf',
                            use_container_width=True,
                            key="btn_download_pdf_codes"
                        )
                    except:
                        st.info("ℹ️ Export PDF non disponible")

        else:
            st.warning("⚠️ Vous devez d'abord valider les voyages.")

# =====================================================
# PAGE 4: VALIDATION & EXPORT (VERSION OPTIMISÉE)
# =====================================================
def page_finalisation():
    st.markdown("<h1 class='main-header'>4. 📊 Planning et KPIs</h1>", unsafe_allow_html=True)
    
    # =====================================================
    # 📤 EXPORT FINAL ET PLANNING COMPLET - VERSION OPTIMISÉE
    # =====================================================
    st.markdown("## 📤 EXPORT FINAL ET PLANNING COMPLET")

    if "df_voyages_valides" in st.session_state and not st.session_state.df_voyages_valides.empty:
        
        df_export_final = st.session_state.df_voyages_valides.copy()
        
        # =====================================================
        # GARANTIR QUE TOUTES LES COLONNES REQUISES EXISTENT
        # =====================================================
        
        # Vérifier et créer la colonne "Chauffeur" si nécessaire
        if "Chauffeur" not in df_export_final.columns:
            # Priorité 1 : Utiliser "Chauffeur attribué"
            if "Chauffeur attribué" in df_export_final.columns:
                df_export_final["Chauffeur"] = df_export_final["Chauffeur attribué"]
                st.success("✅ Colonne 'Chauffeur' créée à partir de 'Chauffeur attribué'")
            # Priorité 2 : Utiliser "Matricule chauffeur" avec format
            elif "Matricule chauffeur" in df_export_final.columns:
                df_export_final["Chauffeur"] = df_export_final["Matricule chauffeur"].apply(
                    lambda x: f"Chauffeur {x}" if pd.notna(x) and x != "" else "À attribuer"
                )
            # Fallback
            else:
                df_export_final["Chauffeur"] = "À attribuer"
                st.warning("⚠️ Colonne 'Chauffeur' créée vide")
        
        # Vérifier que "Code voyage" existe
        if "Code voyage" not in df_export_final.columns:
            st.error("❌ La colonne 'Code voyage' est manquante. Veuillez d'abord générer les codes voyage dans la section 10.")
            st.stop()
        
        # =====================================================
        # FONCTION POUR FORMATER LES COLONNES AVEC RETOURS À LA LIGNE
        # =====================================================
        def formater_colonnes_retours_ligne(df):
            df_formate = df.copy()
            colonnes_a_formater = ['BL inclus', 'Client(s) inclus', 'Représentant(s) inclus']
            
            for col in colonnes_a_formater:
                if col in df_formate.columns:
                    df_formate[col] = df_formate[col].apply(
                        lambda x: '\n'.join([elem.strip() for elem in str(x).replace(';', ',').split(',') if elem.strip()]) 
                        if pd.notna(x) else ""
                    )
            return df_formate
        
        # =====================================================
        # AFFICHAGE DÉTAILLÉ AVEC RETOURS À LA LIGNE
        # =====================================================
        st.markdown("### 📊 Planning de Livraisons Détaillé")
        
        # Appliquer le formatage pour l'affichage Streamlit
        df_affichage_formate = formater_colonnes_retours_ligne(df_export_final)
        
        # Afficher chaque voyage avec expanders détaillés
        for idx, row in df_affichage_formate.iterrows():
            with st.expander(f"🚚 Voyage {row['Véhicule N°']} | Zone : {row['Zone']} | Véhicule: {row.get('Véhicule attribué', 'N/A')} | Chauffeur: {row.get('Chauffeur', 'N/A')}"):
                col1, col2, col3 = st.columns(3)
                
                with col1:
                    st.write("**Informations de base:**")
                    st.write(f"**Code voyage:** {row['Code voyage']}")
                    st.write(f"**Zone:** {row['Zone']}")
                    st.write(f"**Véhicule N°:** {row['Véhicule N°']}")
                    if "Poids total chargé" in row:
                        st.write(f"**Poids total chargé:** {row['Poids total chargé']:.3f} kg")
                    if "Volume total chargé" in row:
                        st.write(f"**Volume total chargé:** {row['Volume total chargé']:.3f} m³")
                    if "Taux d'occupation (%)" in row:
                        st.write(f"**Taux d'occupation:** {row['Taux d\'occupation (%)']:.3f}%")
                    if "Véhicule attribué" in row:
                        st.write(f"**Véhicule attribué:** {row['Véhicule attribué']}")
                    if "Chauffeur" in row:
                        st.write(f"**Chauffeur:** {row['Chauffeur']}")
                
                with col2:
                    # Afficher les clients avec retours à ligne
                    if 'Client(s) inclus' in row and pd.notna(row['Client(s) inclus']):
                        st.write("**📋 Clients inclus:**")
                        clients = str(row['Client(s) inclus']).split('\n')
                        for client in clients:
                            client_clean = client.strip()
                            if client_clean:
                                st.write(f"• {client_clean}")
                    
                    # Afficher les représentants avec retours à ligne
                    if 'Représentant(s) inclus' in row and pd.notna(row['Représentant(s) inclus']):
                        st.write("**👤 Représentants inclus:**")
                        representants = str(row['Représentant(s) inclus']).split('\n')
                        for rep in representants:
                            rep_clean = rep.strip()
                            if rep_clean:
                                st.write(f"• {rep_clean}")
                
                with col3:
                    # Afficher les BL avec retours à ligne
                    if 'BL inclus' in row and pd.notna(row['BL inclus']):
                        st.write("**📄 BL associés:**")
                        bls = str(row['BL inclus']).split('\n')
                        # Afficher en colonnes si beaucoup de BL
                        if len(bls) > 5:
                            cols = st.columns(2)
                            half = len(bls) // 2
                            for i, bl in enumerate(bls):
                                bl_clean = bl.strip()
                                if bl_clean:
                                    col_idx = 0 if i < half else 1
                                    with cols[col_idx]:
                                        st.write(f"• {bl_clean}")
                        else:
                            for bl in bls:
                                bl_clean = bl.strip()
                                if bl_clean:
                                    st.write(f"• {bl_clean}")

        # =====================================================
        # EXPORT EXCEL AVEC RETOURS À LA LIGNE
        # =====================================================
        col_export1, col_export2 = st.columns(2)

        with col_export1:
            # Solution simple : UTC+1 (Tunis est en GMT+1)
            date_tunis = pd.Timestamp.now() + pd.Timedelta(hours=1)
            
            nom_fichier = st.text_input(
                "📝 Nom du fichier d'export", 
                value=f"Planning_Livraisons_{date_tunis.strftime('%Y%m%d_%H%M')}",
                help="Le fichier sera sauvegardé avec l'extension .xlsx"
            )

        with col_export2:
            st.markdown("<br>", unsafe_allow_html=True)
            generer_export = st.button("🚀 Générer l'export complet", type="primary")

        # =====================================================
        # CODE D'EXPORT (EN DEHORS DES COLONNES !)
        # =====================================================
        if generer_export:
            try:
                # Importer la fonction d'export depuis le backend
                from backend import exporter_planning_excel
                
                # Préparer les données supplémentaires
                donnees_supplementaires = {}
                
                # Ajouter les données de base si disponibles
                if st.session_state.df_grouped is not None:
                    donnees_supplementaires['Livraisons_Client_Ville'] = st.session_state.df_grouped
                if st.session_state.df_city is not None:
                    donnees_supplementaires['Besoin_Estafette_Ville'] = st.session_state.df_city
                if st.session_state.df_zone is not None:
                    donnees_supplementaires['Besoin_Estafette_Zone'] = st.session_state.df_zone
                
                # Appliquer le formatage avec retours à ligne avant l'export
                df_export_formate = formater_colonnes_retours_ligne(df_export_final)
                
                # Générer l'export
                success, message = exporter_planning_excel(
                    df_export_formate,  # Utiliser le DataFrame formaté avec retours à ligne
                    f"{nom_fichier}.xlsx",
                    donnees_supplementaires,
                    st.session_state.df_livraisons_original
                )
                
                if success:
                    st.success(message)
                    
                    # =====================================================
                    # APERÇU DU FORMAT D'EXPORT
                    # =====================================================

                    # Aperçu du format d'export
                    st.subheader("👁️ Aperçu du format d'export")
                    colonnes_apercu = ["Code voyage", "Zone", "Véhicule N°", "Chauffeur", "BL inclus", "Client(s) inclus", "Poids total chargé", "Volume total chargé"]
                    colonnes_apercu = [col for col in colonnes_apercu if col in df_export_formate.columns]

                    df_apercu = df_export_formate[colonnes_apercu].head(5).copy()

                    # Formater l'affichage
                    if "Poids total chargé" in df_apercu.columns:
                        df_apercu["Poids total chargé"] = df_apercu["Poids total chargé"].map(lambda x: f"{x:.1f} kg")
                    if "Volume total chargé" in df_apercu.columns:
                        df_apercu["Volume total chargé"] = df_apercu["Volume total chargé"].map(lambda x: f"{x:.3f} m³")

                    # =====================================================
                    # STYLE CSS AVEC CENTRAGE FORCÉ
                    # =====================================================
                    st.markdown("""
                    <style>
                    /* Style général du tableau */
                    .custom-table {
                        border-collapse: collapse;
                        font-family: Arial, sans-serif;
                        font-size: 14px;
                        box-shadow: 0 2px 8px rgba(0,0,0,0.1);
                        border-radius: 8px;
                        overflow: hidden;
                        margin: 0 auto !important;
                        display: table !important;
                    }

                    /* En-têtes du tableau - BLEU ROYAL */
                    .custom-table th {
                        background-color: #0369A1;
                        color: white;
                        padding: 12px 8px;
                        text-align: center;
                        border: 2px solid #4682B4;
                        font-weight: normal;
                        font-size: 13px;
                        vertical-align: middle;
                    }

                    /* Cellules du tableau */
                    .custom-table td {
                        padding: 10px 8px;
                        text-align: center;
                        border: 1px solid #B0C4DE;
                        background-color: white;
                        color: #000000;
                        vertical-align: middle;
                        font-weight: normal;
                    }

                    /* Conteneur du tableau avec CENTRAGE ABSOLU */
                    .table-container {
                        overflow-x: auto;
                        margin: 1rem auto !important;
                        border-radius: 8px;
                        border: 2px solid #4682B4;
                        width: 100% !important;
                        max-width: 100% !important;
                        display: flex !important;
                        justify-content: center !important;
                        padding: 10px !important;
                    }

                    /* Wrapper pour centrer tout */
                    .table-wrapper {
                        display: flex;
                        justify-content: center;
                        align-items: center;
                        width: 100%;
                        padding: 20px 0;
                    }

                    /* Survol des lignes */
                    .custom-table tr:hover td {
                        background-color: #F0F8FF !important;
                    }

                    /* Pour les cellules avec retours à ligne */
                    .custom-table td br {
                        display: block;
                        content: "";
                        margin-top: 2px;
                    }
                    </style>
                    """, unsafe_allow_html=True)

                    # =====================================================
                    # PRÉPARATION DES DONNÉES AVEC <br> POUR HTML
                    # =====================================================
                    # IMPORTANT : Pour l'affichage HTML, utiliser <br> au lieu de \n
                    if "BL inclus" in df_apercu.columns:
                        # Si la colonne contient déjà des \n, les remplacer par <br>
                        df_apercu["BL inclus"] = df_apercu["BL inclus"].astype(str).apply(
                            lambda x: "<br>".join(
                                bl.strip() 
                                for bl in str(x).replace('\n', ',').replace('\\n', ',').split(",") 
                                if bl.strip() and bl.strip() != 'nan'
                            )
                        )

                    if "Client(s) inclus" in df_apercu.columns:
                        df_apercu["Client(s) inclus"] = df_apercu["Client(s) inclus"].astype(str).apply(
                            lambda x: "<br>".join(
                                client.strip() 
                                for client in str(x).replace('\n', ',').replace('\\n', ',').split(",") 
                                if client.strip() and client.strip() != 'nan'
                            )
                        )

                    # Convertir le DataFrame en HTML avec escape=False pour que <br> fonctionne
                    html_table = df_apercu.to_html(
                        escape=False,  # TRÈS IMPORTANT : escape=False pour interpréter <br>
                        index=False, 
                        classes="custom-table",
                        border=0
                    )

                    # Afficher le tableau dans un conteneur centré
                    st.markdown(f"""
                    <div class="table-wrapper">
                        <div class="table-container">
                            {html_table}
                        </div>
                    </div>
                    """, unsafe_allow_html=True)
                    
                    # =====================================================
                    # TÉLÉCHARGEMENT (aussi centré)
                    # =====================================================
                    st.markdown("<br><br>", unsafe_allow_html=True)
                    
                    with open(f"{nom_fichier}.xlsx", "rb") as file:
                        btn = st.download_button(
                            label="💾 Télécharger le planning complet",
                            data=file,
                            file_name=f"{nom_fichier}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )
                else:
                    st.error(message)
                    
            except Exception as e:
                st.error(f"❌ Erreur lors de l'export : {str(e)}")
        
        # =====================================================
        # 🎯 RÉSUMÉ ET TABLEAU DE BORD FINAL
        # =====================================================
        st.markdown("## 🎯 RÉSUMÉ DU PLANNING")

        if "df_voyages" in st.session_state:
            df_final = st.session_state.df_voyages.copy()
            
            # Calcul des métriques principales
            total_vehicules = len(df_final)
            estafettes = len(df_final[df_final["Code Véhicule"] == "ESTAFETTE"])
            camions = len(df_final[df_final["Code Véhicule"] == "CAMION-LOUE"])
            poids_total = df_final["Poids total chargé"].sum()
            volume_total = df_final["Volume total chargé"].sum()
            taux_moyen = df_final["Taux d'occupation (%)"].mean() if "Taux d'occupation (%)" in df_final.columns else 0
            
            # Affichage des métriques
            col_metric1, col_metric2, col_metric3, col_metric4 = st.columns(4)
            
            with col_metric1:
                st.metric("🚚 Total Véhicules", total_vehicules)
            
            with col_metric2:
                st.metric("🚐 Estafettes", estafettes)
            
            with col_metric3:
                st.metric("🚛 Camions", camions)
            
            with col_metric4:
                st.metric("⚖️ Poids Total", f"{poids_total:.0f} kg")
            
            col_metric5, col_metric6, col_metric7, col_metric8 = st.columns(4)
            
            with col_metric5:
                st.metric("📦 Volume Total", f"{volume_total:.1f} m³")
            
            with col_metric6:
                st.metric("📊 Taux Occupation Moyen", f"{taux_moyen:.1f}%")
            
            with col_metric7:
                # Calcul de l'efficacité
                efficacite = " Bonne" if taux_moyen > 70 else " Moyenne" if taux_moyen > 50 else " Faible"
                st.metric("🎯 Efficacité", efficacite)
            
            with col_metric8:
                # Statut de complétion
                status = "✅ Complet" if 'df_voyages_valides' in st.session_state else "🟡 En cours"
                st.metric("📋 Statut", status)
            
            # Graphique de répartition par zone
            st.subheader("📊 Répartition par Zone")
            if 'Zone' in df_final.columns:
                repartition_zone = df_final.groupby("Zone").size().reset_index(name="Nombre de véhicules")
                
                if not repartition_zone.empty:
                    import plotly.express as px

                    fig_zone = px.bar(
                        repartition_zone, 
                        x="Zone", 
                        y="Nombre de véhicules",
                        title="Nombre de véhicules par zone",
                        color="Nombre de véhicules",
                        color_continuous_scale=[
                            "#ADE8F4",   # Bleu clair visible
                            "#90E0EF",
                            "#4EA8DE",
                            "#3A86FF",
                            "#1E6091",
                            "#0A3D62" 
                        ],
                        text="Nombre de véhicules"
                    )

                    fig_zone.update_layout(coloraxis_colorbar=dict(title="Nb Véhicules"))
                    st.plotly_chart(fig_zone, use_container_width=True)

            # Graphique de répartition par type de véhicule
            st.subheader("🚗 Répartition par Type de Véhicule")
            if "Code Véhicule" in df_final.columns:
                repartition_type = df_final["Code Véhicule"].value_counts().reset_index()
                repartition_type.columns = ["Type Véhicule", "Nombre"]
                
                fig_type = px.pie(
                    repartition_type, 
                    values="Nombre", 
                    names="Type Véhicule",
                    title="Répartition des types de véhicules"
                )
                st.plotly_chart(fig_type, use_container_width=True)

        else:
            st.warning("⚠️ Le planning n'est pas encore généré.")
        
        # =====================================================
        # NAVIGATION ENTRE PAGES
        # =====================================================
        st.markdown("---")
        col_nav1, col_nav2, col_nav3 = st.columns(3)
        
        with col_nav1:
            if st.button("← Retour à l'optimisation", use_container_width=True, key="btn_retour_optimisation_finalisation"):
                st.session_state.page = "optimisation"
                st.rerun()
        
        with col_nav2:
            voyages_valides = len(st.session_state.df_voyages_valides) if st.session_state.df_voyages_valides is not None else 0
            st.metric("📊 Voyages validés", voyages_valides)
        
        with col_nav3:
            if st.button("🔄 Recommencer", type="secondary", use_container_width=True):
                # Réinitialiser seulement certaines données
                keys_to_keep = ['page', 'data_processed', 'df_grouped', 'df_city', 'df_grouped_zone', 
                              'df_zone', 'df_livraisons_original', 'df_livraisons']
                
                for key in list(st.session_state.keys()):
                    if key not in keys_to_keep:
                        del st.session_state[key]
                
                st.session_state.rental_processor = TruckRentalProcessor(
                    st.session_state.df_optimized_estafettes, 
                    st.session_state.df_livraisons_original
                )
                st.success("✅ Application réinitialisée. Vous pouvez repartir de l'optimisation.")
                st.rerun()

    else:
        st.warning("⚠️ Vous devez d'abord valider les voyages et générer les codes voyage.")
        
        # Navigation vers la page d'optimisation
        st.markdown("---")
        if st.button("🚚 Retourner à l'optimisation pour valider les voyages", type="primary"):
            st.session_state.page = "optimisation"
            st.rerun()
# =====================================================
# NAVIGATION PRINCIPALE
# =====================================================
def main():
    # Initialiser la page courante
    if 'page' not in st.session_state:
        st.session_state.page = "import"
    
    # Sidebar avec navigation
    with st.sidebar:
        st.image("https://th.bing.com/th/id/OIP.NX4XkAk56j_1bs6CiYhdxQHaHa?pid=ImgDet&rs=1", width=120)
        st.markdown("<h2 style='text-align: center; color: #1E3A8A;'>🚚 Planning Livraisons</h2>", 
                   unsafe_allow_html=True)
        
        st.markdown("---")
        
        # Boutons de navigation avec icônes
        page_options = {
            "import": {"icon": "📥", "label": "Importation Données"},
            "analyse": {"icon": "🔍", "label": "Analyse Détaillée"},
            "optimisation": {"icon": "🚚", "label": "Optimisation"},
            "finalisation": {"icon": "📊", "label": "Planning et KPIs"}
        }
        
        for page_key, page_info in page_options.items():
            is_active = st.session_state.page == page_key
            button_type = "primary" if is_active else "secondary"
            
            if st.button(
                f"{page_info['icon']} {page_info['label']}",
                key=f"nav_{page_key}",
                use_container_width=True,
                type=button_type
            ):
                st.session_state.page = page_key
                st.rerun()
        
        st.markdown("---")
        
        # Statut de l'application
        st.markdown("### 📊 Statut")
        if st.session_state.data_processed:
            st.success("✅ Données chargées")
            if st.session_state.df_voyages_valides is not None:
                st.success("✅ Planning validé")
        else:
            st.warning("⏳ Données requises")
        
        # Pied de page sidebar
        st.markdown("---")
        st.markdown("""
        <div style='text-align: center; font-size: 12px; color: #666;'>
            <p>Développé par Zaineb KCHAOU</p>
            <p>📧 Zaineb.KCHAOU@sopal.com</p>
        </div>
        """, unsafe_allow_html=True)
    
    # Affichage de la page sélectionnée
    if st.session_state.page == "import":
        page_import()
    elif st.session_state.page == "analyse":
        page_analyse()
    elif st.session_state.page == "optimisation":
        page_optimisation()
    elif st.session_state.page == "finalisation":
        page_finalisation()

# =====================================================
# LANCEMENT DE L'APPLICATION
# =====================================================
if __name__ == "__main__":
    main()