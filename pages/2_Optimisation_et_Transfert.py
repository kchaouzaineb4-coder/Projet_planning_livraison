import streamlit as st
import pandas as pd
from backend import TruckRentalProcessor, TruckTransferManager, SEUIL_POIDS, SEUIL_VOLUME
from io import BytesIO
import openpyxl
from openpyxl.styles import Alignment

st.header("🚚 Optimisation et Transfert entre Véhicules")

# Vérification des prérequis
if not st.session_state.data_processed:
    st.warning("⚠️ Veuillez d'abord importer et traiter les données dans la page 'Import & Analyse'.")
    st.stop()

# CSS pour cette page
st.markdown("""
<style>
    .custom-table-rental {
        width: 100%;
        border-collapse: collapse;
        font-family: Arial, sans-serif;
        font-size: 14px;
        box-shadow: 0 2px 8px rgba(0,0,0,0.1);
        border-radius: 8px;
        overflow: hidden;
    }
    .custom-table-rental th {
        background-color: #0369A1;
        color: white;
        padding: 12px 8px;
        text-align: center;
        border: 2px solid #4682B4;
        font-weight: normal;
        font-size: 13px;
        vertical-align: middle;
    }
    .custom-table-rental td {
        padding: 10px 8px;
        text-align: center;
        border: 1px solid #B0C4DE;
        background-color: white;
        color: #000000;
        vertical-align: middle;
        font-weight: normal;
    }
    .table-container-rental {
        overflow-x: auto;
        margin: 1rem 0;
        border-radius: 8px;
        border: 2px solid #4682B4;
    }
    .centered-table {
        margin-left: auto;
        margin-right: auto;
        display: table;
        width: 100%;
    }
    .centered-table table {
        margin: 0 auto;
        border-collapse: collapse;
        width: 100%;
        font-family: Arial, sans-serif;
        box-shadow: 0 2px 8px rgba(0,0,0,0.1);
    }
    .centered-table th {
        background-color: #0369A1;
        color: white;
        padding: 12px 8px;
        text-align: center;
        border: 2px solid #555;
        font-weight: bold;
        font-size: 14px;
        vertical-align: middle;
    }
    .centered-table td {
        padding: 10px 8px;
        text-align: center;
        border: 2px solid #555;
        background-color: #f9f9f9;
        color: #333;
        vertical-align: middle;
    }
</style>
""", unsafe_allow_html=True)

# Section 3: Proposition de location
st.subheader("3. 🚚 Proposition de location de camion")
st.markdown(f"🔸 Si un client dépasse **{SEUIL_POIDS} kg** ou **{SEUIL_VOLUME} m³**, une location est proposée.")

# Initialisation du processeur de location si pas déjà fait
if st.session_state.rental_processor is None:
    st.session_state.rental_processor = TruckRentalProcessor(
        st.session_state.df_optimized_estafettes, 
        st.session_state.df_livraisons_original
    )

# Fonctions de callback
def update_propositions_view():
    if st.session_state.rental_processor:
        st.session_state.propositions = st.session_state.rental_processor.detecter_propositions()
        if (st.session_state.propositions is not None and 
            not st.session_state.propositions.empty and 
            'Client' in st.session_state.propositions.columns):
            if (st.session_state.selected_client is not None and 
                st.session_state.selected_client not in st.session_state.propositions['Client'].astype(str).tolist()):
                st.session_state.selected_client = None
    else:
        st.session_state.propositions = pd.DataFrame()

def handle_location_action(accepter):
    if st.session_state.rental_processor and st.session_state.selected_client:
        try:
            client_to_process = str(st.session_state.selected_client)
            ok, msg, _ = st.session_state.rental_processor.appliquer_location(
                client_to_process, accepter=accepter
            )
            st.session_state.message = msg
            update_propositions_view()
        except Exception as e:
            st.session_state.message = f"❌ Erreur lors du traitement : {str(e)}"
    elif not st.session_state.selected_client:
        st.session_state.message = "⚠️ Veuillez sélectionner un client à traiter."
    else:
        st.session_state.message = "⚠️ Le processeur de location n'est pas initialisé."

def accept_location_callback():
    handle_location_action(True)

def refuse_location_callback():
    handle_location_action(False)

# Initialiser les propositions si nécessaire
if st.session_state.propositions is None:
    update_propositions_view()

# Affichage des propositions
if st.session_state.propositions is not None and not st.session_state.propositions.empty:
    col_prop, col_details = st.columns([2, 3])
    
    with col_prop:
        st.markdown("### Propositions ouvertes")
        
        if 'Client' in st.session_state.propositions.columns:
            propositions_display = st.session_state.propositions.copy()
            
            # Formater les nombres
            if "Poids total (kg)" in propositions_display.columns:
                propositions_display["Poids total (kg)"] = propositions_display["Poids total (kg)"].map(
                    lambda x: f"{float(x):.3f}" if pd.notna(x) else ""
                )
            if "Volume total (m³)" in propositions_display.columns:
                propositions_display["Volume total (m³)"] = propositions_display["Volume total (m³)"].map(
                    lambda x: f"{float(x):.3f}" if pd.notna(x) else ""
                )
            
            # Afficher le tableau
            html_table_propositions = propositions_display.to_html(
                escape=False, 
                index=False, 
                classes="custom-table-rental",
                border=0
            )
            
            st.markdown(f"""
            <div class="table-container-rental">
                {html_table_propositions}
            </div>
            """, unsafe_allow_html=True)
            
            # Métriques résumées
            st.markdown("---")
            col_metric1, col_metric2, col_metric3 = st.columns(3)

            with col_metric1:
                total_propositions = len(st.session_state.propositions)
                st.metric("📋 Propositions ouvertes", total_propositions)

            with col_metric2:
                clients_poids = len(st.session_state.propositions[
                    st.session_state.propositions["Poids total (kg)"] >= SEUIL_POIDS
                ]) if "Poids total (kg)" in st.session_state.propositions.columns else 0
                st.metric("⚖️ Dépassement poids", clients_poids)

            with col_metric3:
                clients_volume = len(st.session_state.propositions[
                    st.session_state.propositions["Volume total (m³)"] >= SEUIL_VOLUME
                ]) if "Volume total (m³)" in st.session_state.propositions.columns else 0
                st.metric("📦 Dépassement volume", clients_volume)
        
            # Sélection du client
            client_options = st.session_state.propositions['Client'].astype(str).tolist()
            client_options_with_empty = [""] + client_options
            
            default_index = 0
            if st.session_state.selected_client in client_options:
                 default_index = client_options_with_empty.index(st.session_state.selected_client)
            elif len(client_options) > 0:
                 default_index = 1

            st.session_state.selected_client = st.selectbox(
                "Client à traiter :", 
                options=client_options_with_empty, 
                index=default_index,
                key='client_select' 
            )
        else:
            st.warning("⚠️ Format de données incorrect dans les propositions.")
            st.session_state.selected_client = None

        col_btn_acc, col_btn_ref = st.columns(2)
        is_client_selected = st.session_state.selected_client != "" and st.session_state.selected_client is not None
        
        with col_btn_acc:
            st.button(
                "✅ Accepter la location", 
                on_click=accept_location_callback, 
                disabled=not is_client_selected,
                use_container_width=True
            )
        with col_btn_ref:
            st.button(
                "❌ Refuser la proposition", 
                on_click=refuse_location_callback, 
                disabled=not is_client_selected,
                use_container_width=True
            )

    with col_details:
        st.markdown("### Détails de la commande client")
        is_client_selected = st.session_state.selected_client != "" and st.session_state.selected_client is not None
        
        if is_client_selected:
            try:
                resume, details_df = st.session_state.rental_processor.get_details_client(
                    st.session_state.selected_client
                )
                
                # Afficher le résumé
                st.markdown(f"**{resume}**")
                
                # Formatage du tableau des détails
                if not details_df.empty:
                    details_display = details_df.copy()
                    
                    def format_numeric_column(series, decimals, unit=""):
                        formatted_series = series.copy()
                        for i, value in enumerate(series):
                            if pd.notna(value) and value != "":
                                try:
                                    if isinstance(value, str):
                                        clean_value = value.replace(' kg', '').replace(' m³', '').replace('%', '').strip()
                                        num_value = float(clean_value)
                                    else:
                                        num_value = float(value)
                                    
                                    if decimals == 3:
                                        formatted_value = f"{num_value:.3f}"
                                    elif decimals == 2:
                                        formatted_value = f"{num_value:.2f}"
                                    elif decimals == 1:
                                        formatted_value = f"{num_value:.1f}"
                                    else:
                                        formatted_value = f"{num_value:.0f}"
                                    
                                    formatted_series.iloc[i] = f"{formatted_value}{unit}"
                                except (ValueError, TypeError):
                                    formatted_series.iloc[i] = str(value)
                            else:
                                formatted_series.iloc[i] = ""
                        return formatted_series
                    
                    # Formater les colonnes numériques
                    if "Poids total" in details_display.columns:
                        details_display["Poids total"] = format_numeric_column(details_display["Poids total"], 3, " kg")
                    
                    if "Volume total" in details_display.columns:
                        details_display["Volume total"] = format_numeric_column(details_display["Volume total"], 3, " m³")
                    
                    if "Taux d'occupation (%)" in details_display.columns:
                        details_display["Taux d'occupation (%)"] = format_numeric_column(details_display["Taux d'occupation (%)"], 2, "%")
                    
                    # Gestion spéciale pour "BL inclus"
                    if "BL inclus" in details_display.columns:
                        details_display["BL inclus"] = details_display["BL inclus"].astype(str).apply(
                            lambda x: "<br>".join(bl.strip() for bl in x.split(";")) if ";" in x else x
                        )
                    
                    # Afficher le tableau
                    html_table_details = details_display.to_html(
                        escape=False, 
                        index=False, 
                        classes="custom-table-rental",
                        border=0
                    )
                    
                    st.markdown(f"""
                    <div class="table-container-rental">
                        {html_table_details}
                    </div>
                    """, unsafe_allow_html=True)
                    
                    # Métriques pour les détails
                    st.markdown("---")
                    col_det1, col_det2, col_det3 = st.columns(3)
                    
                    with col_det1:
                        total_camions = len(details_display)
                        st.metric("🚚 Nombre de camions", total_camions)
                    
                    with col_det2:
                        if "Poids total" in details_df.columns:
                            poids_total = 0
                            for value in details_df["Poids total"]:
                                if pd.notna(value):
                                    try:
                                        if isinstance(value, str):
                                            clean_value = value.replace(' kg', '').replace('m³', '').strip()
                                        else:
                                            clean_value = str(value)
                                        poids_total += float(clean_value)
                                    except (ValueError, TypeError):
                                        continue
                            st.metric("📦 Poids total", f"{poids_total:.1f} kg")
                        else:
                            st.metric("📦 Poids total", "N/A")
                    
                    with col_det3:
                        if "Volume total" in details_df.columns:
                            volume_total = 0
                            for value in details_df["Volume total"]:
                                if pd.notna(value):
                                    try:
                                        if isinstance(value, str):
                                            clean_value = value.replace(' kg', '').replace('m³', '').strip()
                                        else:
                                            clean_value = str(value)
                                        volume_total += float(clean_value)
                                    except (ValueError, TypeError):
                                        continue
                            st.metric("📏 Volume total", f"{volume_total:.3f} m³")
                        else:
                            st.metric("📏 Volume total", "N/A")
                        
            except Exception as e:
                st.error(f"❌ Erreur lors de la récupération des détails : {str(e)}")
        else:
            st.info("Sélectionnez un client pour afficher les détails de la commande/estafettes.")
else:
    st.success("✅ Aucune proposition de location de camion en attente de décision.")

st.markdown("---")

# Section 4: Voyages optimisés
st.subheader("4. 🚐 Voyages par Estafette Optimisé (Inclut Camions Loués)")

try:
    # Récupération sécurisée du DataFrame
    if st.session_state.rental_processor:
        df_optimized_estafettes = st.session_state.rental_processor.get_df_result()
    elif "df_voyages" in st.session_state:
        df_optimized_estafettes = st.session_state.df_voyages.copy()
    else:
        st.error("❌ Données non disponibles. Veuillez exécuter le traitement complet.")
        st.stop()
    
    # Vérifier que le DataFrame n'est pas vide
    if df_optimized_estafettes.empty:
        st.warning("⚠️ Aucune donnée à afficher.")
        st.stop()
    
    # Nettoyer les colonnes en double
    df_clean = df_optimized_estafettes.loc[:, ~df_optimized_estafettes.columns.duplicated()]
    
    # TRIER PAR ZONE
    if "Zone" in df_clean.columns:
        df_clean["Zone_Num"] = df_clean["Zone"].str.extract('(\d+)').astype(float)
        df_clean = df_clean.sort_values("Zone_Num").drop("Zone_Num", axis=1)
    
    # Définir l'ordre des colonnes pour l'affichage
    colonnes_ordre = [
        "Zone", "Véhicule N°", "Poids total chargé", "Volume total chargé",
        "Client(s) inclus", "Représentant(s) inclus", "BL inclus", 
        "Taux d'occupation (%)", "Location_camion", "Location_proposee", "Code Véhicule"
    ]
    
    # Filtrer seulement les colonnes qui existent
    colonnes_finales = [col for col in colonnes_ordre if col in df_clean.columns]
    
    # Créer le DataFrame d'affichage avec retours à la ligne
    df_display = df_clean[colonnes_finales].copy()
    
    # Transformer les colonnes avec retours à la ligne HTML
    if "Client(s) inclus" in df_display.columns:
        df_display["Client(s) inclus"] = df_display["Client(s) inclus"].astype(str).apply(
            lambda x: "<br>".join(client.strip() for client in x.split(",")) if x != "nan" else ""
        )
    
    if "Représentant(s) inclus" in df_display.columns:
        df_display["Représentant(s) inclus"] = df_display["Représentant(s) inclus"].astype(str).apply(
            lambda x: "<br>".join(rep.strip() for rep in x.split(",")) if x != "nan" else ""
        )
    
    if "BL inclus" in df_display.columns:
        df_display["BL inclus"] = df_display["BL inclus"].astype(str).apply(
            lambda x: "<br>".join(bl.strip() for bl in x.split(";")) if x != "nan" else ""
        )
    
    # Formater les colonnes numériques pour l'affichage
    if "Poids total chargé" in df_display.columns:
        df_display["Poids total chargé"] = df_display["Poids total chargé"].map(lambda x: f"{x:.3f} kg")
    if "Volume total chargé" in df_display.columns:
        df_display["Volume total chargé"] = df_display["Volume total chargé"].map(lambda x: f"{x:.3f} m³")
    if "Taux d'occupation (%)" in df_display.columns:
        df_display["Taux d'occupation (%)"] = df_display["Taux d'occupation (%)"].map(lambda x: f"{x:.3f}%")
    
    # Afficher le tableau avec le style CSS professionnel
    html_table = df_display.to_html(escape=False, index=False, classes="custom-table", border=0)
    
    st.markdown(f"""
    <div class="table-container">
        {html_table}
    </div>
    """, unsafe_allow_html=True)
    
    # MÉTRIQUES RÉSUMÉES
    st.markdown("---")
    col1, col2, col3, col4 = st.columns(4)
    
    with col1:
        total_voyages = len(df_display)
        st.metric("🚐 Total Voyages", total_voyages)
    
    with col2:
        total_zones = df_display["Zone"].nunique() if "Zone" in df_display.columns else 0
        st.metric("🌍 Zones couvertes", total_zones)
    
    with col3:
        camions_loues = df_display["Location_camion"].sum() if "Location_camion" in df_display.columns else 0
        st.metric("🚚 Camions loués", int(camions_loues))
    
    with col4:
        estafettes = total_voyages - camions_loues
        st.metric("📦 Estafettes", estafettes)
    
    # Préparer l'export Excel avec retours à la ligne \n
    df_export = df_clean.copy()
    
    # S'assurer que l'export est aussi trié par zone
    if "Zone" in df_export.columns:
        df_export["Zone_Num"] = df_export["Zone"].str.extract('(\d+)').astype(float)
        df_export = df_export.sort_values("Zone_Num").drop("Zone_Num", axis=1)
    
    # Transformer les colonnes avec retours à la ligne \n pour Excel
    if "Client(s) inclus" in df_export.columns:
        df_export["Client(s) inclus"] = df_export["Client(s) inclus"].astype(str).apply(
            lambda x: "\n".join(client.strip() for client in x.split(",")) if x != "nan" else ""
        )
    
    if "Représentant(s) inclus" in df_export.columns:
        df_export["Représentant(s) inclus"] = df_export["Représentant(s) inclus"].astype(str).apply(
            lambda x: "\n".join(rep.strip() for rep in x.split(",")) if x != "nan" else ""
        )
    
    if "BL inclus" in df_export.columns:
        df_export["BL inclus"] = df_export["BL inclus"].astype(str).apply(
            lambda x: "\n".join(bl.strip() for bl in x.split(";")) if x != "nan" else ""
        )
    
    # Formater les colonnes numériques pour l'export
    if "Poids total chargé" in df_export.columns:
        df_export["Poids total chargé"] = df_export["Poids total chargé"].round(3)
    if "Volume total chargé" in df_export.columns:
        df_export["Volume total chargé"] = df_export["Volume total chargé"].round(3)
    
    # Bouton de téléchargement avec formatage Excel
    excel_buffer = BytesIO()
    
    with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
        df_export.to_excel(writer, index=False, sheet_name="Voyages Optimisés")
        
        # Récupérer le workbook et worksheet pour appliquer le formatage
        workbook = writer.book
        worksheet = writer.sheets["Voyages Optimisés"]
        
        # Appliquer le style wrap_text aux colonnes avec retours à la ligne
        wrap_columns = []
        if "Client(s) inclus" in df_export.columns:
            wrap_columns.append("Client(s) inclus")
        if "Représentant(s) inclus" in df_export.columns:
            wrap_columns.append("Représentant(s) inclus")
        if "BL inclus" in df_export.columns:
            wrap_columns.append("BL inclus")
        
        # Appliquer le format wrap_text à toutes les cellules des colonnes concernées
        for col_idx, col_name in enumerate(df_export.columns):
            if col_name in wrap_columns:
                col_letter = openpyxl.utils.get_column_letter(col_idx + 1)
                for row in range(2, len(df_export) + 2):
                    cell = worksheet[f"{col_letter}{row}"]
                    cell.alignment = Alignment(wrap_text=True, vertical='top')
        
        # Ajuster la largeur des colonnes pour une meilleure visibilité
        for column in worksheet.columns:
            max_length = 0
            column_letter = openpyxl.utils.get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
    
    excel_buffer.seek(0)
    
    st.download_button(
        label="💾 Télécharger Voyages Estafette Optimisés",
        data=excel_buffer,
        file_name="Voyages_Estafette_Optimises.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    
    # Mise à jour pour les sections suivantes
    st.session_state.df_voyages = df_clean

except KeyError as e:
    st.error(f"❌ Erreur de colonne manquante : {e}")
    if st.session_state.rental_processor:
        st.session_state.df_voyages = st.session_state.rental_processor.df_base.copy()
        st.rerun()
        
except Exception as e:
    st.error(f"❌ Erreur lors de l'affichage des voyages optimisés: {str(e)}")

st.markdown("---")

# Section 5: Transfert de BLs
st.subheader("5. 🔁 Transfert de BLs entre Estafettes / Camions")

MAX_POIDS = 1550
MAX_VOLUME = 4.608

if "df_voyages" not in st.session_state:
    st.warning("⚠️ Vous devez d'abord exécuter la section 4 (Voyages par Estafette Optimisé).")
elif "df_livraisons" not in st.session_state:
    st.warning("⚠️ Le DataFrame des livraisons détaillées n'est pas disponible.")
else:
    df_voyages = st.session_state.df_voyages.copy()
    df_livraisons = st.session_state.df_livraisons.copy()

    colonnes_requises = ["Zone", "Véhicule N°", "Poids total chargé", "Volume total chargé", "BL inclus"]

    if not all(col in df_voyages.columns for col in colonnes_requises):
        st.error(f"❌ Le DataFrame ne contient pas toutes les colonnes nécessaires : {', '.join(colonnes_requises)}")
    else:
        zones_disponibles = sorted(df_voyages["Zone"].dropna().unique().tolist())
        zone_selectionnee = st.selectbox("🌍 Sélectionner une zone", zones_disponibles)

        if zone_selectionnee:
            df_zone = df_voyages[df_voyages["Zone"] == zone_selectionnee]
            vehicules = sorted(df_zone["Véhicule N°"].dropna().unique().tolist())

            col1, col2 = st.columns(2)
            with col1:
                source = st.selectbox("🚐 Estafette / Camion source", vehicules)
            with col2:
                cible = st.selectbox("🎯 Estafette / Camion cible", [v for v in vehicules if v != source])

            if source and cible:
                df_source = df_zone[df_zone["Véhicule N°"] == source]
                if df_source.empty or df_source["BL inclus"].isna().all():
                    st.warning("⚠️ Aucun BL trouvé pour ce véhicule source.")
                else:
                    st.subheader(f"📦 BLs actuellement assignés à {source}")

                    # Créer un mapping BL → Client
                    bls_avec_clients = []
                    bls_simples = df_source["BL inclus"].iloc[0].split(";")
                    
                    for bl in bls_simples:
                        client_info = df_livraisons[df_livraisons["No livraison"] == bl]
                        if not client_info.empty:
                            client_nom = client_info["Client de l'estafette"].iloc[0]
                            bl_affichage = f"{bl} - {client_nom}"
                        else:
                            bl_affichage = f"{bl} - Client non trouvé"
                        bls_avec_clients.append(bl_affichage)
                    
                    # Affichage formaté avec clients
                    df_source_display = df_source[["Véhicule N°", "Poids total chargé", "Volume total chargé"]].copy()
                    df_source_display["BL inclus (avec clients)"] = "<br>".join(bls_avec_clients)
                    
                    df_source_display["Poids total chargé"] = df_source_display["Poids total chargé"].map(lambda x: f"{x:.3f} kg")
                    df_source_display["Volume total chargé"] = df_source_display["Volume total chargé"].map(lambda x: f"{x:.3f} m³")
                    
                    # Afficher le tableau source
                    html_content = f"""
                    <div class="centered-table">
                    {df_source_display.to_html(escape=False, index=False)}
                    </div>
                    """
                    st.markdown(html_content, unsafe_allow_html=True)
                    
                    # Sélectionner les BLs à transférer
                    st.subheader("📋 Sélectionner les BLs à transférer")
                    
                    options_transfert = []
                    mapping_bl_original = {}
                    
                    for bl in bls_simples:
                        client_info = df_livraisons[df_livraisons["No livraison"] == bl]
                        if not client_info.empty:
                            client_nom = client_info["Client de l'estafette"].iloc[0]
                            option_affichage = f"{bl} - {client_nom}"
                        else:
                            option_affichage = f"{bl} - Client non trouvé"
                        
                        options_transfert.append(option_affichage)
                        mapping_bl_original[option_affichage] = bl
                    
                    # CSS pour le multiselect
                    st.markdown("""
                    <style>
                    .stMultiSelect > div > div {
                        background-color: #F8FAFC !important;
                        border: 2px solid #CBD5E1 !important;
                        border-radius: 8px !important;
                    }
                    </style>
                    """, unsafe_allow_html=True)
                    
                    # Multiselect avec clients
                    bls_selectionnes_affichage = st.multiselect(
                        "Sélectionnez les BLs à transférer (avec clients) :", 
                        options_transfert,
                        format_func=lambda x: x
                    )
                    
                    # Convertir la sélection en BLs simples pour le traitement
                    bls_selectionnes = [mapping_bl_original[bl_affichage] for bl_affichage in bls_selectionnes_affichage]

                    if bls_selectionnes and st.button("🔁 Exécuter le transfert"):
                        df_bls_selection = df_livraisons[df_livraisons["No livraison"].isin(bls_selectionnes)]
                        poids_bls = df_bls_selection["Poids total"].sum()
                        volume_bls = df_bls_selection["Volume total"].sum()

                        df_cible = df_zone[df_zone["Véhicule N°"] == cible]
                        poids_cible = df_cible["Poids total chargé"].sum()
                        volume_cible = df_cible["Volume total chargé"].sum()

                        if (poids_cible + poids_bls) > MAX_POIDS or (volume_cible + volume_bls) > MAX_VOLUME:
                            st.warning("⚠️ Le transfert dépasse les limites de poids ou volume du véhicule cible.")
                        else:
                            def transfer_bl(row):
                                bls = row["BL inclus"].split(";") if pd.notna(row["BL inclus"]) else []
                                bls_to_move = [b for b in bls if b in bls_selectionnes]

                                if row["Véhicule N°"] == source:
                                    new_bls = [b for b in bls if b not in bls_to_move]
                                    row["BL inclus"] = ";".join(new_bls)
                                    row["Poids total chargé"] = max(0, row["Poids total chargé"] - poids_bls)
                                    row["Volume total chargé"] = max(0, row["Volume total chargé"] - volume_bls)
                                elif row["Véhicule N°"] == cible:
                                    new_bls = bls + bls_to_move
                                    row["BL inclus"] = ";".join(new_bls)
                                    row["Poids total chargé"] += poids_bls
                                    row["Volume total chargé"] += volume_bls
                                return row

                            df_voyages = df_voyages.apply(transfer_bl, axis=1)
                            st.session_state.df_voyages = df_voyages
                            
                            # Afficher un résumé du transfert avec clients
                            clients_transferes = df_bls_selection["Client de l'estafette"].unique()
                            st.success(f"""
                            ✅ Transfert réussi !
                            - **{len(bls_selectionnes)} BL(s)** déplacé(s) de **{source}** vers **{cible}**
                            - **Clients concernés :** {', '.join(clients_transferes)}
                            - **Poids transféré :** {poids_bls:.1f} kg
                            - **Volume transféré :** {volume_bls:.3f} m³
                            """)

                            # Affichage Streamlit avec retours à la ligne
                            st.subheader("📊 Voyages après transfert (toutes les zones)")
                            df_display = df_voyages.sort_values(by=["Zone", "Véhicule N°"]).copy()
                            
                            # Transformer les colonnes avec retours à la ligne HTML
                            if "BL inclus" in df_display.columns:
                                df_display["BL inclus"] = df_display["BL inclus"].astype(str).apply(
                                    lambda x: "<br>".join(bl.strip() for bl in x.split(";")) if x != "nan" else ""
                                )
                            
                            df_display["Poids total chargé"] = df_display["Poids total chargé"].map(lambda x: f"{x:.3f} kg")
                            df_display["Volume total chargé"] = df_display["Volume total chargé"].map(lambda x: f"{x:.3f} m³")
                            
                            # Affichage avec HTML amélioré
                            html_content_after = f"""
                            <div class="centered-table">
                            {df_display[colonnes_requises].to_html(escape=False, index=False)}
                            </div>
                            """
                            st.markdown(html_content_after, unsafe_allow_html=True)

                            # Export Excel avec retours à la ligne \n
                            df_export = df_voyages.copy()
                            
                            # Transformer les BL avec retours à la ligne \n pour Excel
                            if "BL inclus" in df_export.columns:
                                df_export["BL inclus"] = df_export["BL inclus"].astype(str).apply(
                                    lambda x: "\n".join(bl.strip() for bl in x.split(";")) if x != "nan" else ""
                                )
                            
                            df_export["Poids total chargé"] = df_export["Poids total chargé"].round(3)
                            df_export["Volume total chargé"] = df_export["Volume total chargé"].round(3)

                            excel_buffer = BytesIO()
                            with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
                                df_export.to_excel(writer, index=False, sheet_name='Transfert BLs')
                                
                                # Appliquer le format wrap_text pour Excel
                                workbook = writer.book
                                worksheet = writer.sheets['Transfert BLs']
                                
                                # Appliquer le style wrap_text à la colonne BL inclus
                                if "BL inclus" in df_export.columns:
                                    for col_idx, col_name in enumerate(df_export.columns):
                                        if col_name == "BL inclus":
                                            col_letter = openpyxl.utils.get_column_letter(col_idx + 1)
                                            for row in range(2, len(df_export) + 2):
                                                cell = worksheet[f"{col_letter}{row}"]
                                                cell.alignment = Alignment(wrap_text=True, vertical='top')
                            
                            excel_buffer.seek(0)

                            st.download_button(
                                label="💾 Télécharger le tableau mis à jour (XLSX)",
                                data=excel_buffer,
                                file_name="voyages_apres_transfert.xlsx",
                                mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                            )

st.markdown("---")

# Section 6: Ajout d'objets manuels
st.subheader("6. 📦 AJOUT D'OBJETS MANUELS AUX VÉHICULES")

if "df_voyages" in st.session_state:
    # Initialiser le gestionnaire de transfert si pas déjà fait
    if "transfer_manager" not in st.session_state:
        st.session_state.transfer_manager = TruckTransferManager(
            st.session_state.df_voyages, 
            st.session_state.df_livraisons
        )
    
    df_voyages = st.session_state.df_voyages.copy()
    
    col1, col2, col3 = st.columns(3)
    
    with col1:
        # Sélection de la zone
        zones_disponibles = sorted(df_voyages["Zone"].dropna().unique().tolist())
        zone_objet = st.selectbox("🌍 Zone", zones_disponibles, key="zone_objet")
    
    with col2:
        # Sélection du véhicule dans la zone choisie
        if zone_objet:
            vehicules_zone = sorted(
                df_voyages[df_voyages["Zone"] == zone_objet]["Véhicule N°"].dropna().unique().tolist()
            )
            vehicule_objet = st.selectbox("🚚 Véhicule", vehicules_zone, key="vehicule_objet")
        else:
            vehicule_objet = st.selectbox("🚚 Véhicule", [], key="vehicule_objet")
    
    with col3:
        # Informations sur le véhicule sélectionné
        if zone_objet and vehicule_objet:
            vehicule_data = df_voyages[
                (df_voyages["Zone"] == zone_objet) & 
                (df_voyages["Véhicule N°"] == vehicule_objet)
            ].iloc[0]
            
            is_camion = vehicule_data.get("Code Véhicule", "") == "CAMION-LOUE"
            capacite_poids = 30500 if is_camion else 1550
            capacite_volume = 77.5 if is_camion else 4.608
            
            poids_actuel = vehicule_data.get("Poids total chargé", 0)
            volume_actuel = vehicule_data.get("Volume total chargé", 0)
            
            st.metric(
                "📊 Capacité utilisée", 
                f"{poids_actuel:.1f}kg / {capacite_poids}kg",
                f"{volume_actuel:.3f}m³ / {capacite_volume}m³"
            )
    
    # Formulaire d'ajout d'objet
    st.markdown("### 📝 Détails de l'objet à ajouter")
    
    col4, col5, col6 = st.columns(3)
    
    with col4:
        nom_objet = st.text_input("🏷️ Nom de l'objet", placeholder="Ex: Matériel urgent, Colis oublié...")
    
    with col5:
        poids_objet = st.number_input("⚖️ Poids (kg)", min_value=0.0, max_value=1000.0, value=10.0, step=0.1)
    
    with col6:
        volume_objet = st.number_input("📦 Volume (m³)", min_value=0.0, max_value=10.0, value=0.1, step=0.01)
    
    if st.button("➕ Ajouter l'objet au véhicule", type="primary"):
        if not nom_objet:
            st.error("❌ Veuillez donner un nom à l'objet.")
        elif zone_objet and vehicule_objet:
            try:
                # Appel de la méthode add_manual_object
                success, message, df_updated = st.session_state.transfer_manager.add_manual_object(
                    df_voyages=df_voyages,
                    vehicle=vehicule_objet,
                    zone=zone_objet,
                    name=nom_objet,
                    weight=poids_objet,
                    volume=volume_objet
                )
                
                if success:
                    st.success(message)
                    
                    # Mise à jour des données
                    st.session_state.df_voyages = df_updated
                    st.session_state.transfer_manager.df_voyages = df_updated.copy()
                    
                    if st.session_state.rental_processor:
                        try:
                            st.session_state.rental_processor.df_base = df_updated.copy()
                            st.session_state.rental_processor = TruckRentalProcessor(
                                df_updated, 
                                st.session_state.df_livraisons_original
                            )
                            st.success("✅ Processeur de location synchronisé")
                        except Exception as e:
                            st.warning(f"⚠️ Synchronisation partielle du processeur : {str(e)}")
                    
                    # Mettre à jour les propositions de location si elles existent
                    if st.session_state.propositions is not None:
                        try:
                            st.session_state.propositions = st.session_state.rental_processor.detecter_propositions()
                        except:
                            pass
                    
                    # Mettre à jour les voyages validés si ils existent
                    if 'df_voyages_valides' in st.session_state:
                        try:
                            mask_valides = df_updated["Véhicule N°"].isin(
                                st.session_state.df_voyages_valides["Véhicule N°"]
                            )
                            st.session_state.df_voyages_valides = df_updated[mask_valides].copy()
                        except:
                            pass
                    
                    # Afficher le véhicule mis à jour
                    vehicule_update = df_updated[
                        (df_updated["Zone"] == zone_objet) & 
                        (df_updated["Véhicule N°"] == vehicule_objet)
                    ].iloc[0]
                    
                    st.info(f"""
                    **Véhicule mis à jour :**
                    - Poids total : {vehicule_update['Poids total chargé']:.1f} kg
                    - Volume total : {vehicule_update['Volume total chargé']:.3f} m³
                    - Taux d'occupation : {vehicule_update['Taux d\'occupation (%)']:.1f}%
                    - BLs inclus : {vehicule_update['BL inclus']}
                    """)
                    
                    st.success("🔄 Toutes les données ont été mises à jour avec succès !")
                    st.rerun()
                    
                else:
                    st.error(message)
                    
            except Exception as e:
                st.error(f"❌ Erreur lors de l'ajout de l'objet : {str(e)}")
        else:
            st.error("❌ Veuillez sélectionner une zone et un véhicule.")
    
    # Affichage des objets ajoutés récemment
    st.markdown("### 📋 Historique des objets ajoutés")
    
    # Rechercher les objets manuels dans les BLs
    objets_manuels = []
    for idx, row in df_voyages.iterrows():
        bls = str(row.get("BL inclus", ""))
        if "OBJ-" in bls:
            for bl in bls.split(";"):
                if bl.startswith("OBJ-"):
                    vehicule_info = df_voyages[
                        (df_voyages["Zone"] == row["Zone"]) & 
                        (df_voyages["Véhicule N°"] == row["Véhicule N°"])
                    ]
                    if not vehicule_info.empty:
                        poids_vehicule = vehicule_info["Poids total chargé"].iloc[0]
                        volume_vehicule = vehicule_info["Volume total chargé"].iloc[0]
                        
                        objets_manuels.append({
                            "Véhicule": row["Véhicule N°"],
                            "Zone": row["Zone"],
                            "Objet": bl,
                            "Poids Véhicule": f"{poids_vehicule:.1f} kg",
                            "Volume Véhicule": f"{volume_vehicule:.3f} m³",
                            "Type": "Camion" if row.get("Code Véhicule", "") == "CAMION-LOUE" else "Estafette"
                        })
    
    if objets_manuels:
        df_objets = pd.DataFrame(objets_manuels)
        
        # Fonction show_df simplifiée pour cette page
        def show_df_simple(df, **kwargs):
            st.dataframe(df, **kwargs)
        
        show_df_simple(df_objets, use_container_width=True)
        
        # Bouton pour supprimer tous les objets
        col_clear1, col_clear2 = st.columns([3, 1])
        with col_clear2:
            if st.button("🗑️ Supprimer tous les objets", type="secondary"):
                df_sans_objets = st.session_state.df_voyages.copy()
                for idx, row in df_sans_objets.iterrows():
                    bls_originaux = str(row["BL inclus"]).split(";")
                    bls_filtres = [bl for bl in bls_originaux if not bl.startswith("OBJ-")]
                    df_sans_objets.at[idx, "BL inclus"] = ";".join(bls_filtres)
                
                st.session_state.df_voyages = df_sans_objets
                st.session_state.transfer_manager.df_voyages = df_sans_objets.copy()
                if st.session_state.rental_processor:
                    st.session_state.rental_processor.df_base = df_sans_objets.copy()
                
                st.success("✅ Tous les objets manuels ont été supprimés")
                st.rerun()
    else:
        st.info("Aucun objet manuel ajouté pour le moment.")

else:
    st.warning("⚠️ Données non disponibles.")