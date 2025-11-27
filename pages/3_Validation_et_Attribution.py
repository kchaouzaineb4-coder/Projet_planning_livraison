import streamlit as st
import pandas as pd
from backend import VEHICULES_DISPONIBLES, CHAUFFEURS_DETAILS
from io import BytesIO
import openpyxl
from openpyxl.styles import Alignment
from fpdf import FPDF

st.header("✅ Validation des Voyages et Planning Final")

# Vérification des prérequis
if not st.session_state.data_processed:
    st.warning("⚠️ Veuillez d'abord importer et traiter les données dans la page 'Import & Analyse'.")
    st.stop()

# CSS pour cette page
st.markdown("""
<style>
    .voyage-card {
        border: 1px solid #e0e0e0;
        border-radius: 10px;
        padding: 20px;
        margin: 10px 0;
        background: white;
        box-shadow: 0 2px 4px rgba(0,0,0,0.1);
    }
    .voyage-header {
        background: #0369A1;
        color: white;
        padding: 15px;
        border-radius: 8px;
        margin-bottom: 15px;
    }
    .metric-card {
        background: #f8f9fa;
        border-left: 4px solid #0369A1;
        padding: 12px;
        margin: 8px 0;
        border-radius: 5px;
    }
    .bl-list {
        background: #fff3cd;
        border: 1px solid #ffeaa7;
        border-radius: 5px;
        padding: 10px;
        margin: 10px 0;
        max-height: 150px;
        overflow-y: auto;
    }
    .validation-buttons {
        display: flex;
        gap: 10px;
        margin-top: 15px;
    }
</style>
""", unsafe_allow_html=True)

# Section 7: Validation des voyages
st.subheader("7. ✅ VALIDATION DES VOYAGES APRÈS TRANSFERT")

# Fonction pour exporter DataFrame en Excel avec arrondi
def to_excel(df, sheet_name="Voyages Validés"):
    df_export = df.copy()
    if "Poids total chargé" in df_export.columns:
        df_export["Poids total chargé"] = df_export["Poids total chargé"].round(3)
    if "Volume total chargé" in df_export.columns:
        df_export["Volume total chargé"] = df_export["Volume total chargé"].round(3)

    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df_export.to_excel(writer, index=False, sheet_name=sheet_name)
    return output.getvalue()

# Création du DataFrame de validation
if "df_voyages" in st.session_state:
    voyages_apres_transfert = st.session_state.df_voyages.copy()
    df_validation = voyages_apres_transfert.copy()

    if "validations" not in st.session_state:
        st.session_state.validations = {}

    # Affichage amélioré des voyages
    st.markdown("### 📋 Liste des Voyages à Valider")
    
    for idx, row in df_validation.iterrows():
        # Création d'une carte pour chaque voyage
        with st.container():
            st.markdown(f"""
            <div class="voyage-card">
                <div class="voyage-header">
                    <h4>🚚 Voyage {row['Véhicule N°']} | Zone: {row['Zone']}</h4>
                </div>
            """, unsafe_allow_html=True)
            
            # Métriques principales
            col1, col2, col3 = st.columns(3)
            
            with col1:
                st.markdown(f"""
                <div class="metric-card">
                    <strong>⚖️ Poids Total</strong><br>
                    {row['Poids total chargé']:.3f} kg
                </div>
                """, unsafe_allow_html=True)
            
            with col2:
                st.markdown(f"""
                <div class="metric-card">
                    <strong>📏 Volume Total</strong><br>
                    {row['Volume total chargé']:.3f} m³
                </div>
                """, unsafe_allow_html=True)
            
            with col3:
                taux_occupation = row.get('Taux d\'occupation (%)', 'N/A')
                if taux_occupation != 'N/A':
                    taux_text = f"{taux_occupation:.1f}%"
                else:
                    taux_text = "N/A"
                st.markdown(f"""
                <div class="metric-card">
                    <strong>📊 Taux d'Occupation</strong><br>
                    {taux_text}
                </div>
                """, unsafe_allow_html=True)
            
            # Informations détaillées
            col4, col5 = st.columns(2)
            
            with col4:
                clients = row.get('Client(s) inclus', '')
                if clients:
                    st.markdown(f"**👥 Clients:** {clients}")
                
                representants = row.get('Représentant(s) inclus', '')
                if representants:
                    st.markdown(f"**👨‍💼 Représentants:** {representants}")
            
            with col5:
                location = "✅ Oui" if row.get('Location_camion') else "❌ Non"
                st.markdown(f"**🚛 Location:** {location}")
                
                code_vehicule = row.get('Code Véhicule', 'N/A')
                st.markdown(f"**🔧 Code Véhicule:** {code_vehicule}")
            
            # Liste des BL avec défilement
            bls = row.get('BL inclus', '')
            if bls:
                bls_list = bls.split(';')
                bls_html = "<br>".join([f"• {bl.strip()}" for bl in bls_list])
                st.markdown(f"""
                <div class="bl-list">
                    <strong>📋 BLs Inclus ({len(bls_list)}):</strong><br>
                    {bls_html}
                </div>
                """, unsafe_allow_html=True)
            
            # Boutons de validation côte à côte
            st.markdown("**✅ Validation du voyage:**")
            col_oui, col_non = st.columns(2)
            
            with col_oui:
                if st.button(f"✅ Valider {row['Véhicule N°']}", key=f"btn_oui_{idx}", 
                           use_container_width=True, type="primary" if st.session_state.validations.get(idx) == "Oui" else "secondary"):
                    st.session_state.validations[idx] = "Oui"
                    st.rerun()
            
            with col_non:
                if st.button(f"❌ Rejeter {row['Véhicule N°']}", key=f"btn_non_{idx}",
                           use_container_width=True, type="primary" if st.session_state.validations.get(idx) == "Non" else "secondary"):
                    st.session_state.validations[idx] = "Non"
                    st.rerun()
            
            # Afficher le statut actuel
            statut = st.session_state.validations.get(idx)
            if statut == "Oui":
                st.success(f"✅ Voyage {row['Véhicule N°']} validé")
            elif statut == "Non":
                st.error(f"❌ Voyage {row['Véhicule N°']} rejeté")
            else:
                st.info("⏳ En attente de validation")
            
            st.markdown("</div>", unsafe_allow_html=True)
            st.markdown("---")

    # Résumé des validations
    st.markdown("### 📊 Résumé des Validations")
    total_voyages = len(df_validation)
    valides = sum(1 for v in st.session_state.validations.values() if v == "Oui")
    rejetes = sum(1 for v in st.session_state.validations.values() if v == "Non")

    col1, col2, col3 = st.columns(3)

    with col1:
        st.metric("Total Voyages", total_voyages)
    with col2:
        st.metric("✅ Validés", valides)
    with col3:
        st.metric("❌ Rejetés", rejetes)

    # Information supplémentaire sur l'état des validations
    if valides + rejetes < total_voyages:
        st.info(f"ℹ️ {total_voyages - (valides + rejetes)} voyage(s) n'ont pas encore été validés")

    # Bouton pour appliquer les validations
    if st.button("🚀 Finaliser la Validation", type="primary", use_container_width=True):
        valid_indexes = [i for i, v in st.session_state.validations.items() if v == "Oui"]
        valid_indexes = [i for i in valid_indexes if i in df_validation.index]

        if valid_indexes:
            df_voyages_valides = df_validation.loc[valid_indexes].reset_index(drop=True)
            st.session_state.df_voyages_valides = df_voyages_valides

            st.success(f"✅ {len(df_voyages_valides)} voyage(s) validé(s) avec succès!")
            
            # Affichage des voyages validés
            st.markdown("### 🎉 Voyages Validés - Résumé Final")
            
            for idx, row_valide in df_voyages_valides.iterrows():
                with st.expander(f"🚚 {row_valide['Véhicule N°']} - Zone {row_valide['Zone']}", expanded=True):
                    col1, col2 = st.columns(2)
                    with col1:
                        st.metric("Poids", f"{row_valide['Poids total chargé']:.3f} kg")
                        st.metric("Clients", row_valide.get('Client(s) inclus', 'N/A'))
                    with col2:
                        st.metric("Volume", f"{row_valide['Volume total chargé']:.3f} m³")
                        st.metric("Représentants", row_valide.get('Représentant(s) inclus', 'N/A'))

            # Export Excel
            excel_data = to_excel(df_voyages_valides)
            st.download_button(
                label="💾 Télécharger les voyages validés (XLSX)",
                data=excel_data,
                file_name="Voyages_valides.xlsx",
                mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                use_container_width=True
            )
        else:
            st.warning("⚠️ Aucun voyage n'a été validé. Veuillez valider au moins un voyage.")

else:
    st.warning("⚠️ Vous devez d'abord exécuter la section 4 (Voyages par Estafette Optimisé).")

st.markdown("---")

# Section 8: Attribution véhicules et chauffeurs
st.subheader("8. 🚛 ATTRIBUTION DES VÉHICULES ET CHAUFFEURS")

if 'df_voyages_valides' in st.session_state and not st.session_state.df_voyages_valides.empty:

    df_attribution = st.session_state.df_voyages_valides.copy()

    # Fonction pour formatter les colonnes avec retours à la ligne POUR STREAMLIT
    def formater_colonnes_listes_streamlit(df):
        df_formate = df.copy()
        colonnes_a_formater = ['Client(s) inclus', 'Représentant(s) inclus', 'BL inclus']
        
        for col in colonnes_a_formater:
            if col in df_formate.columns:
                df_formate[col] = df_formate[col].apply(
                    lambda x: '\n'.join([elem.strip() for elem in str(x).replace(';', ',').split(',') if elem.strip()]) 
                    if pd.notna(x) else ""
                )
        return df_formate

    if "attributions" not in st.session_state:
        st.session_state.attributions = {}

    for idx, row in df_attribution.iterrows():
        with st.expander(f"🚚 Voyage {row['Véhicule N°']} | Zone : {row['Zone']}"):
            st.write("**Informations du voyage :**")
            
            # Créer un affichage personnalisé avec retours à ligne
            col1, col2, col3 = st.columns(3)
            
            with col1:
                st.write(f"**Zone:** {row['Zone']}")
                st.write(f"**Véhicule N°:** {row['Véhicule N°']}")
                if "Poids total chargé" in row:
                    st.write(f"**Poids total chargé:** {row['Poids total chargé']:.2f} kg")
                if "Volume total chargé" in row:
                    st.write(f"**Volume total chargé:** {row['Volume total chargé']:.3f} m³")
                if "Taux d'occupation (%)" in row:
                    st.write(f"**Taux d'occupation:** {row['Taux d\'occupation (%)']:.1f}%")
            
            with col2:
                # Afficher les clients avec retours à ligne
                if 'Client(s) inclus' in row and pd.notna(row['Client(s) inclus']):
                    st.write("**Clients:**")
                    clients = str(row['Client(s) inclus']).replace(';', ',').split(',')
                    for client in clients:
                        client_clean = client.strip()
                        if client_clean:
                            st.write(f"- {client_clean}")
                
                # Afficher les représentants avec retours à ligne
                if 'Représentant(s) inclus' in row and pd.notna(row['Représentant(s) inclus']):
                    st.write("**Représentants:**")
                    representants = str(row['Représentant(s) inclus']).replace(';', ',').split(',')
                    for rep in representants:
                        rep_clean = rep.strip()
                        if rep_clean:
                            st.write(f"- {rep_clean}")
            
            with col3:
                # Afficher les BL avec retours à ligne
                if 'BL inclus' in row and pd.notna(row['BL inclus']):
                    st.write("**BL associés:**")
                    bls = str(row['BL inclus']).replace(';', ',').split(',')
                    for bl in bls:
                        bl_clean = bl.strip()
                        if bl_clean:
                            st.write(f"- {bl_clean}")

            col_veh, col_chauf = st.columns(2)
            
            with col_veh:
                vehicule_selectionne = st.selectbox(
                    f"Véhicule pour le voyage {row['Véhicule N°']}",
                    VEHICULES_DISPONIBLES,
                    index=0 if st.session_state.attributions.get(idx, {}).get("Véhicule") else 0,
                    key=f"vehicule_{idx}"
                )
            
            with col_chauf:
                options_chauffeurs = [f"{matricule} - {nom}" for matricule, nom in CHAUFFEURS_DETAILS.items() if matricule != 'Matricule']
                
                default_index = 0
                chauffeur_actuel = st.session_state.attributions.get(idx, {}).get("Chauffeur_complet")
                if chauffeur_actuel and chauffeur_actuel in options_chauffeurs:
                    default_index = options_chauffeurs.index(chauffeur_actuel)
                
                chauffeur_selectionne_complet = st.selectbox(
                    f"Chauffeur pour le voyage {row['Véhicule N°']}",
                    options_chauffeurs,
                    index=default_index,
                    key=f"chauffeur_{idx}"
                )
                
                if chauffeur_selectionne_complet:
                    matricule_chauffeur = chauffeur_selectionne_complet.split(" - ")[0]
                    nom_chauffeur = chauffeur_selectionne_complet.split(" - ")[1]
                else:
                    matricule_chauffeur = ""
                    nom_chauffeur = ""

            st.session_state.attributions[idx] = {
                "Véhicule": vehicule_selectionne,
                "Chauffeur_complet": chauffeur_selectionne_complet,
                "Matricule_chauffeur": matricule_chauffeur,
                "Nom_chauffeur": nom_chauffeur
            }

    if st.button("✅ Appliquer les attributions"):

        df_attribution["Véhicule attribué"] = df_attribution.index.map(lambda i: st.session_state.attributions[i]["Véhicule"])
        df_attribution["Chauffeur attribué"] = df_attribution.index.map(lambda i: st.session_state.attributions[i]["Nom_chauffeur"])
        df_attribution["Matricule chauffeur"] = df_attribution.index.map(lambda i: st.session_state.attributions[i]["Matricule_chauffeur"])

        
        st.markdown("### 📦 Voyages avec Véhicule et Chauffeur")

        # Affichage Streamlit amélioré avec retours à ligne
        for idx, row in df_attribution.iterrows():
            with st.expander(f"📋 Voyage {row['Véhicule N°']} - Zone {row['Zone']} - Véhicule: {row.get('Véhicule attribué', 'N/A')} - Chauffeur: {row.get('Chauffeur attribué', 'N/A')}"):
                col1, col2, col3 = st.columns(3)
                
                with col1:
                    st.write("**Informations de base:**")
                    st.write(f"**Zone:** {row['Zone']}")
                    st.write(f"**Véhicule N°:** {row['Véhicule N°']}")
                    if "Poids total chargé" in row:
                        st.write(f"**Poids total chargé:** {row['Poids total chargé']:.3f} kg")
                    if "Volume total chargé" in row:
                        st.write(f"**Volume total chargé:** {row['Volume total chargé']:.3f} m³")
                    if "Taux d'occupation (%)" in row:
                        st.write(f"**Taux d'occupation:** {row['Taux d\'occupation (%)']:.3f}%")
                    if "Véhicule attribué" in row:
                        st.write(f"**Véhicule attribué:** {row['Véhicule attribué']}")
                    if "Chauffeur attribué" in row:
                        st.write(f"**Chauffeur attribué:** {row['Chauffeur attribué']}")
                    if "Matricule chauffeur" in row:
                        st.write(f"**Matricule chauffeur:** {row['Matricule chauffeur']}")
                
                with col2:
                    # Afficher les clients avec retours à ligne
                    if 'Client(s) inclus' in row and pd.notna(row['Client(s) inclus']):
                        st.write("**📋 Clients inclus:**")
                        clients = str(row['Client(s) inclus']).replace(';', ',').split(',')
                        for client in clients:
                            client_clean = client.strip()
                            if client_clean:
                                st.write(f"• {client_clean}")
                    
                    # Afficher les représentants avec retours à ligne
                    if 'Représentant(s) inclus' in row and pd.notna(row['Représentant(s) inclus']):
                        st.write("**👤 Représentants inclus:**")
                        representants = str(row['Représentant(s) inclus']).replace(';', ',').split(',')
                        for rep in representants:
                            rep_clean = rep.strip()
                            if rep_clean:
                                st.write(f"• {rep_clean}")
                
                with col3:
                    # Afficher les BL avec retours à ligne
                    if 'BL inclus' in row and pd.notna(row['BL inclus']):
                        st.write("**📄 BL associés:**")
                        bls = str(row['BL inclus']).replace(';', ',').split(',')
                        # Afficher en colonnes si beaucoup de BL
                        if len(bls) > 5:
                            cols = st.columns(2)
                            half = len(bls) // 2
                            for i, bl in enumerate(bls):
                                bl_clean = bl.strip()
                                if bl_clean:
                                    col_idx = 0 if i < half else 1
                                    with cols[col_idx]:
                                        st.write(f"• {bl_clean}")
                        else:
                            for bl in bls:
                                bl_clean = bl.strip()
                                if bl_clean:
                                    st.write(f"• {bl_clean}")

        # Export Excel avec retours à ligne et CENTRAGE
        def to_excel_attributions(df):
            df_export = df.copy()
            
            # Formater les colonnes avec retours à ligne pour Excel
            colonnes_a_formater = ['Client(s) inclus', 'Représentant(s) inclus', 'BL inclus']
            for col in colonnes_a_formater:
                if col in df_export.columns:
                    df_export[col] = df_export[col].apply(
                        lambda x: '\n'.join([elem.strip() for elem in str(x).replace(';', ',').split(',') if elem.strip()]) 
                        if pd.notna(x) else ""
                    )
            
            if "Poids total chargé" in df_export.columns:
                df_export["Poids total chargé"] = df_export["Poids total chargé"].round(3)
            if "Volume total chargé" in df_export.columns:
                df_export["Volume total chargé"] = df_export["Volume total chargé"].round(3)
            
            output = BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                df_export.to_excel(writer, index=False, sheet_name='Voyages_Attribués')
                
                # Appliquer le formatage des retours à ligne et CENTRAGE dans Excel
                workbook = writer.book
                worksheet = writer.sheets['Voyages_Attribués']
                
                # Style de centrage avec retours à ligne
                center_alignment = Alignment(
                    horizontal='center', 
                    vertical='center', 
                    wrap_text=True
                )
                
                # Appliquer le centrage à TOUTES les cellules
                for row in worksheet.iter_rows(min_row=1, max_row=len(df_export) + 1, min_col=1, max_col=len(df_export.columns)):
                    for cell in row:
                        cell.alignment = center_alignment
                
                # Ajuster automatiquement la largeur des colonnes
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if cell.value:
                                lines = str(cell.value).split('\n')
                                max_line_length = max(len(line) for line in lines)
                                max_length = max(max_length, max_line_length)
                        except:
                            pass
                    adjusted_width = min(50, (max_length + 2))
                    worksheet.column_dimensions[column_letter].width = adjusted_width
                
                # Ajuster la hauteur des lignes pour les retours à ligne
                for row in range(2, len(df_export) + 2):
                    worksheet.row_dimensions[row].height = 60
            
            return output.getvalue()

        # Export PDF avec tableau ÉLARGI et ESPACES MINIMISÉS
        def to_pdf_better_centered(df, title="Voyages Attribués"):
            pdf = FPDF(orientation='L')
            pdf.add_page()
            
            # RÉDUCTION des marges pour utiliser TOUTE la largeur
            pdf.set_left_margin(5)
            pdf.set_right_margin(5)
            pdf.set_top_margin(10)
            
            # Titre PLUS PETIT et PLUS HAUT
            pdf.set_font("Arial", 'B', 14)
            pdf.cell(0, 8, title, ln=True, align="C")
            pdf.ln(3)
            
            # Créer une copie formatée pour le PDF
            df_pdf = df.copy()
            
            # Formater les nombres avec 3 chiffres après la virgule SAUF le taux avec 2 chiffres
            numeric_columns = {
                'Poids total chargé': ('kg', 3),
                'Volume total chargé': ('m³', 3), 
                'Taux d\'occupation (%)': ('%', 2)
            }
            
            for col, (unit, decimals) in numeric_columns.items():
                if col in df_pdf.columns:
                    df_pdf[col] = df_pdf[col].apply(
                        lambda x: f"{float(x):.{decimals}f} {unit}" if x and str(x).strip() and str(x).strip() != 'nan' else ""
                    )
            
            # Configuration des colonnes AVEC LARGEURS MAXIMALISÉES
            col_config = {
                'Zone': {'width': 15, 'header': 'Zone'},
                'Véhicule N°': {'width': 18, 'header': 'Véhicule'},
                'Poids total chargé': {'width': 22, 'header': 'Poids (kg)'},
                'Volume total chargé': {'width': 22, 'header': 'Volume (m³)'},
                'Client(s) inclus': {'width': 30, 'header': 'Clients'},
                'Représentant(s) inclus': {'width': 30, 'header': 'Représentants'},
                'BL inclus': {'width': 35, 'header': 'BL associés'},
                'Taux d\'occupation (%)': {'width': 18, 'header': 'Taux %'},
                'Véhicule attribué': {'width': 25, 'header': 'Véhicule Attribué'},
                'Chauffeur attribué': {'width': 25, 'header': 'Chauffeur'},
                'Matricule chauffeur': {'width': 20, 'header': 'Matricule'}
            }
            
            # Sélectionner seulement les colonnes existantes
            colonnes_existantes = [col for col in df_pdf.columns if col in col_config]
            widths = [col_config[col]['width'] for col in colonnes_existantes]
            headers = [col_config[col]['header'] for col in colonnes_existantes]
            
            # Calculer la position de départ - DÉBUT PLUS À GAUCHE
            total_width = sum(widths)
            start_x = 5
            
            # Positionner le tableau AU DÉBUT
            pdf.set_x(start_x)
            
            # En-têtes CENTRÉS avec police PLUS PETITE
            pdf.set_font("Arial", 'B', 8)
            for i, header in enumerate(headers):
                pdf.cell(widths[i], 6, header, border=1, align='C')
            pdf.ln()
            
            # Données avec centrage VERTICAL et HORIZONTAL
            pdf.set_font("Arial", '', 7)
            
            for voyage_idx, (_, row) in enumerate(df_pdf.iterrows()):
                # Vérifier si on dépasse la hauteur de page
                if pdf.get_y() > 180:
                    pdf.add_page()
                    pdf.set_x(start_x)
                    pdf.set_font("Arial", 'B', 8)
                    for i, header in enumerate(headers):
                        pdf.cell(widths[i], 6, header, border=1, align='C')
                    pdf.ln()
                    pdf.set_font("Arial", '', 7)
                
                # Déterminer le nombre de lignes nécessaires pour ce voyage
                list_columns = ['Client(s) inclus', 'Représentant(s) inclus', 'BL inclus']
                non_list_columns = [col for col in colonnes_existantes if col not in list_columns]
                
                max_lines = 1
                list_contents = {}
                
                for col in list_columns:
                    if col in colonnes_existantes:
                        content = str(row[col]) if pd.notna(row[col]) and str(row[col]) != 'nan' else ""
                        elements = content.replace(';', ',').split(',')
                        elements = [elem.strip() for elem in elements if elem.strip()]
                        list_contents[col] = elements
                        max_lines = max(max_lines, len(elements))
                
                # Pour chaque ligne du voyage
                for line_idx in range(max_lines):
                    # Vérifier si on dépasse la hauteur de page pour cette ligne
                    if pdf.get_y() > 190:
                        pdf.add_page()
                        pdf.set_x(start_x)
                        pdf.set_font("Arial", 'B', 8)
                        for i, header in enumerate(headers):
                            pdf.cell(widths[i], 6, header, border=1, align='C')
                        pdf.ln()
                        pdf.set_font("Arial", '', 7)
                    
                    # Positionner au DÉBUT pour chaque ligne
                    pdf.set_x(start_x)
                    
                    for i, col in enumerate(colonnes_existantes):
                        if col in list_columns:
                            elements = list_contents.get(col, [])
                            content = elements[line_idx] if line_idx < len(elements) else ""
                        else:
                            if line_idx == 0:
                                content = str(row[col]) if pd.notna(row[col]) and str(row[col]) != 'nan' else ""
                            else:
                                content = ""
                        
                        border = 'LR'
                        if line_idx == 0: border += 'T'
                        if line_idx == max_lines - 1: border += 'B'
                        if i == 0: border += 'L'
                        if i == len(colonnes_existantes) - 1: border += 'R'
                        
                        pdf.cell(widths[i], 5, content, border=border, align='C')
                    
                    pdf.ln()
            
            return pdf.output(dest='S').encode('latin-1')

        # Afficher les boutons de téléchargement côte à côte
        col1, col2 = st.columns(2)

        with col1:
            st.download_button(
                label="💾 Télécharger le tableau final (XLSX)",
                data=to_excel_attributions(df_attribution),
                file_name="Voyages_attribues.xlsx",
                mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )

        with col2:
            st.download_button(
                label="📄 Télécharger le tableau final (PDF)",
                data=to_pdf_better_centered(df_attribution),
                file_name="Voyages_attribues.pdf",
                mime='application/pdf'
            )
                
        # Mettre à jour le session state
        st.session_state.df_voyages_valides = df_attribution
        st.success("✅ Attributions appliquées avec succès !")
        
else:
    st.warning("⚠️ Vous devez d'abord valider les voyages dans la section 7.")

st.markdown("---")

# Section 9: Rapports et analytics
st.subheader("9. 📊 RAPPORTS AVANCÉS ET ANALYTICS")

if "df_voyages" in st.session_state and "df_livraisons_original" in st.session_state:
    
    # Initialiser le générateur de rapports
    from backend import AdvancedReportGenerator
    report_generator = AdvancedReportGenerator(
        st.session_state.df_voyages, 
        st.session_state.df_livraisons_original
    )
    
    tab1, tab2, tab3, tab4 = st.tabs([
        "📈 Rapport Analytique", 
        "👤 Rapport Client", 
        "💰 Analyse Coûts", 
        "✅ Validation Données"
    ])
    
    with tab1:
        st.subheader("Rapport Analytique Complet")
        if st.button("🔄 Générer le rapport analytique"):
            with st.spinner("Génération du rapport en cours..."):
                rapport = report_generator.generer_rapport_analytique()
                st.text_area("Rapport détaillé", rapport, height=400)
    
    with tab2:
        st.subheader("Rapport Spécifique Client")
        clients_disponibles = sorted(st.session_state.df_livraisons_original["Client de l'estafette"].unique())
        client_rapport = st.selectbox("Sélectionner un client", clients_disponibles)
        
        if st.button("📋 Générer rapport client"):
            with st.spinner("Génération du rapport client..."):
                rapport_client = report_generator.generer_rapport_client(client_rapport)
                st.text_area(f"Rapport pour {client_rapport}", rapport_client, height=300)
    
    with tab3:
        st.subheader("Analyse des Coûts")
        col_cost1, col_cost2 = st.columns(2)
        
        with col_cost1:
            cout_estafette = st.number_input("Coût unitaire estafette (TND)", value=150, min_value=50, max_value=500)
        with col_cost2:
            cout_camion = st.number_input("Coût unitaire camion (TND)", value=800, min_value=300, max_value=2000)
        
        if st.button("💰 Calculer les coûts"):
            from backend import calculer_couts_estimation
            couts = calculer_couts_estimation(
                st.session_state.df_voyages, 
                cout_estafette, 
                cout_camion
            )
            
            if 'erreur' not in couts:
                st.success(couts['cout_estimation'])
                
                # Graphique des coûts
                import plotly.express as px
                df_couts = pd.DataFrame({
                    'Type': ['Estafettes', 'Camions'],
                    'Coût Total (TND)': [
                        couts['estafettes'] * couts['cout_estafette_unitaire'],
                        couts['camions'] * couts['cout_camion_unitaire']
                    ]
                })
                
                fig = px.pie(df_couts, values='Coût Total (TND)', names='Type', 
                            title='Répartition des coûts par type de véhicule')
                st.plotly_chart(fig, use_container_width=True)
            else:
                st.error(couts['erreur'])
    
    with tab4:
        st.subheader("Validation d'Intégrité des Données")
        if st.button("🔍 Vérifier l'intégrité des données"):
            from backend import verifier_integrite_donnees
            resultat_validation = verifier_integrite_donnees(
                st.session_state.df_voyages,
                st.session_state.df_livraisons_original
            )
            
            if "✅" in resultat_validation:
                st.success(resultat_validation)
            else:
                st.warning(resultat_validation)

else:
    st.warning("⚠️ Vous devez d'abord traiter les données.")

st.markdown("---")

# Section 10: Génération des codes voyage
st.subheader("10. 🏷️ GÉNÉRATION DES CODES VOYAGE")

if "df_voyages_valides" in st.session_state and not st.session_state.df_voyages_valides.empty:
    
    df_final = st.session_state.df_voyages_valides.copy()
    
    # Configuration des paramètres de génération
    col1, col2, col3 = st.columns(3)
    
    with col1:
        date_voyage = st.date_input(
            "📅 Date de voyage",
            value=pd.Timestamp.now().date(),
            help="Date prévue pour les livraisons"
        )
    
    with col2:
        numero_debut = st.number_input(
            "🔢 Numéro séquentiel de départ",
            min_value=1,
            max_value=1000,
            value=1,
            help="Numéro de départ pour la séquence"
        )
    
    with col3:
        st.markdown("<br>", unsafe_allow_html=True)
        generer_codes = st.button("🏷️ Générer les codes voyage", type="primary")
    
    if generer_codes:
        try:
            # Préparation des données pour le code voyage
            df_final['Date Voyage Format'] = date_voyage.strftime('%Y%m%d')
            
            # Création du numéro séquentiel pour chaque voyage
            df_final['Numero Séquentiel'] = range(numero_debut, numero_debut + len(df_final))
            df_final['Numero Séquentiel Formatted'] = df_final['Numero Séquentiel'].apply(lambda x: f"{x:03d}")
            
            # Création du Code voyage
            df_final['Code voyage'] = (
                df_final['Véhicule N°'].astype(str) + '/' +
                df_final['Date Voyage Format'].astype(str) + '/' +
                df_final['Numero Séquentiel Formatted'].astype(str)
            )
            
            # Mettre à jour le session state
            st.session_state.df_voyages_valides = df_final
            
            st.success(f"✅ {len(df_final)} codes voyage générés avec succès !")
            
            # Afficher un aperçu des codes générés
            st.markdown("### 📋 Aperçu des codes voyage générés")
            df_apercu = df_final[['Véhicule N°', 'Zone', 'Code voyage']].copy()
            st.dataframe(df_apercu, use_container_width=True)
            
        except Exception as e:
            st.error(f"❌ Erreur lors de la génération des codes voyage : {str(e)}")
    
    # Afficher les codes existants si déjà générés
    elif 'Code voyage' in df_final.columns:
        st.success("✅ Codes voyage déjà générés")
        df_apercu = df_final[['Véhicule N°', 'Zone', 'Code voyage']].copy()
        st.dataframe(df_apercu, use_container_width=True)
        
        # Option pour regénérer les codes
        if st.button("🔄 Regénérer les codes voyage"):
            columns_to_remove = ['Code voyage', 'Date Voyage Format', 'Numero Séquentiel', 'Numero Séquentiel Formatted']
            for col in columns_to_remove:
                if col in df_final.columns:
                    df_final.drop(col, axis=1, inplace=True)
            st.session_state.df_voyages_valides = df_final
            st.rerun()

else:
    st.warning("⚠️ Vous devez d'abord valider les voyages.")

st.markdown("---")

# Section 11: Export final
st.subheader("11. 📤 EXPORT FINAL ET PLANNING COMPLET")

if "df_voyages_valides" in st.session_state and not st.session_state.df_voyages_valides.empty:
    
    df_export_final = st.session_state.df_voyages_valides.copy()
    
    # GARANTIR QUE TOUTES LES COLONNES REQUISES EXISTENT
    if "Chauffeur" not in df_export_final.columns:
        if "Chauffeur attribué" in df_export_final.columns:
            df_export_final["Chauffeur"] = df_export_final["Chauffeur attribué"]
            st.success("✅ Colonne 'Chauffeur' créée à partir de 'Chauffeur attribué'")
        elif "Matricule chauffeur" in df_export_final.columns:
            df_export_final["Chauffeur"] = df_export_final["Matricule chauffeur"].apply(
                lambda x: f"Chauffeur {x}" if pd.notna(x) and x != "" else "À attribuer"
            )
        else:
            df_export_final["Chauffeur"] = "À attribuer"
            st.warning("⚠️ Colonne 'Chauffeur' créée vide")
    
    # Vérifier que "Code voyage" existe
    if "Code voyage" not in df_export_final.columns:
        st.error("❌ La colonne 'Code voyage' est manquante. Veuillez d'abord générer les codes voyage dans la section 10.")
        st.stop()
    
    # FONCTION POUR FORMATER LES COLONNES AVEC RETOURS À LA LIGNE
    def formater_colonnes_retours_ligne(df):
        df_formate = df.copy()
        colonnes_a_formater = ['BL inclus', 'Client(s) inclus', 'Représentant(s) inclus']
        
        for col in colonnes_a_formater:
            if col in df_formate.columns:
                df_formate[col] = df_formate[col].apply(
                    lambda x: '\n'.join([elem.strip() for elem in str(x).replace(';', ',').split(',') if elem.strip()]) 
                    if pd.notna(x) else ""
                )
        return df_formate
    
    # AFFICHAGE DÉTAILLÉ AVEC RETOURS À LA LIGNE
    st.markdown("### 📊 Planning de Livraisons Détaillé")
    
    # Appliquer le formatage pour l'affichage Streamlit
    df_affichage_formate = formater_colonnes_retours_ligne(df_export_final)
    
    # Afficher chaque voyage avec expanders détaillés
    for idx, row in df_affichage_formate.iterrows():
        with st.expander(f"🚚 Voyage {row['Véhicule N°']} | Zone : {row['Zone']} | Véhicule: {row.get('Véhicule attribué', 'N/A')} | Chauffeur: {row.get('Chauffeur', 'N/A')}"):
            col1, col2, col3 = st.columns(3)
            
            with col1:
                st.write("**Informations de base:**")
                st.write(f"**Code voyage:** {row['Code voyage']}")
                st.write(f"**Zone:** {row['Zone']}")
                st.write(f"**Véhicule N°:** {row['Véhicule N°']}")
                if "Poids total chargé" in row:
                    st.write(f"**Poids total chargé:** {row['Poids total chargé']:.3f} kg")
                if "Volume total chargé" in row:
                    st.write(f"**Volume total chargé:** {row['Volume total chargé']:.3f} m³")
                if "Taux d'occupation (%)" in row:
                    st.write(f"**Taux d'occupation:** {row['Taux d\'occupation (%)']:.3f}%")
                if "Véhicule attribué" in row:
                    st.write(f"**Véhicule attribué:** {row['Véhicule attribué']}")
                if "Chauffeur" in row:
                    st.write(f"**Chauffeur:** {row['Chauffeur']}")
            
            with col2:
                # Afficher les clients avec retours à ligne
                if 'Client(s) inclus' in row and pd.notna(row['Client(s) inclus']):
                    st.write("**📋 Clients inclus:**")
                    clients = str(row['Client(s) inclus']).split('\n')
                    for client in clients:
                        client_clean = client.strip()
                        if client_clean:
                            st.write(f"• {client_clean}")
                
                # Afficher les représentants avec retours à ligne
                if 'Représentant(s) inclus' in row and pd.notna(row['Représentant(s) inclus']):
                    st.write("**👤 Représentants inclus:**")
                    representants = str(row['Représentant(s) inclus']).split('\n')
                    for rep in representants:
                        rep_clean = rep.strip()
                        if rep_clean:
                            st.write(f"• {rep_clean}")
            
            with col3:
                # Afficher les BL avec retours à ligne
                if 'BL inclus' in row and pd.notna(row['BL inclus']):
                    st.write("**📄 BL associés:**")
                    bls = str(row['BL inclus']).split('\n')
                    # Afficher en colonnes si beaucoup de BL
                    if len(bls) > 5:
                        cols = st.columns(2)
                        half = len(bls) // 2
                        for i, bl in enumerate(bls):
                            bl_clean = bl.strip()
                            if bl_clean:
                                col_idx = 0 if i < half else 1
                                with cols[col_idx]:
                                    st.write(f"• {bl_clean}")
                    else:
                        for bl in bls:
                            bl_clean = bl.strip()
                            if bl_clean:
                                st.write(f"• {bl_clean}")

    # EXPORT EXCEL AVEC RETOURS À LA LIGNE
    col_export1, col_export2 = st.columns(2)
    
    with col_export1:
        nom_fichier = st.text_input(
            "📝 Nom du fichier d'export", 
            value=f"Planning_Livraisons_{pd.Timestamp.now().strftime('%Y%m%d_%H%M')}",
            help="Le fichier sera sauvegardé avec l'extension .xlsx"
        )
    
    with col_export2:
        st.markdown("<br>", unsafe_allow_html=True)
        if st.button("🚀 Générer l'export complet", type="primary"):
            try:
                from backend import exporter_planning_excel
                
                # Préparer les données supplémentaires
                donnees_supplementaires = {}
                
                # Ajouter les données de base si disponibles
                if st.session_state.df_grouped is not None:
                    donnees_supplementaires['Livraisons_Client_Ville'] = st.session_state.df_grouped
                if st.session_state.df_city is not None:
                    donnees_supplementaires['Besoin_Estafette_Ville'] = st.session_state.df_city
                if st.session_state.df_zone is not None:
                    donnees_supplementaires['Besoin_Estafette_Zone'] = st.session_state.df_zone
                
                # Appliquer le formatage avec retours à ligne avant l'export
                df_export_formate = formater_colonnes_retours_ligne(df_export_final)
                
                # Générer l'export
                success, message = exporter_planning_excel(
                    df_export_formate,
                    f"{nom_fichier}.xlsx",
                    donnees_supplementaires,
                    st.session_state.df_livraisons_original
                )
                                
                if success:
                    st.success(message)
                    
                    # Aperçu du format d'export
                    st.subheader("👁️ Aperçu du format d'export")
                    colonnes_apercu = ["Code voyage", "Zone", "Véhicule N°", "Chauffeur", "BL inclus", "Client(s) inclus", "Poids total chargé", "Volume total chargé"]
                    colonnes_apercu = [col for col in colonnes_apercu if col in df_export_formate.columns]
                    
                    df_apercu = df_export_formate[colonnes_apercu].head(5).copy()
                    
                    # Formater l'affichage
                    if "Poids total chargé" in df_apercu.columns:
                        df_apercu["Poids total chargé"] = df_apercu["Poids total chargé"].map(lambda x: f"{x:.1f} kg")
                    if "Volume total chargé" in df_apercu.columns:
                        df_apercu["Volume total chargé"] = df_apercu["Volume total chargé"].map(lambda x: f"{x:.3f} m³")
                    
                    st.dataframe(df_apercu, use_container_width=True)
                    
                    # Proposer le téléchargement
                    with open(f"{nom_fichier}.xlsx", "rb") as file:
                        btn = st.download_button(
                            label="💾 Télécharger le planning complet",
                            data=file,
                            file_name=f"{nom_fichier}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )
                else:
                    st.error(message)
                    
            except Exception as e:
                st.error(f"❌ Erreur lors de l'export : {str(e)}")

    # APERÇU DU PLANNING FINAL (TABLEAU SIMPLE)
    st.markdown("### 👁️ Aperçu du Planning Final (Vue Tableau)")
    
    df_apercu_final = df_export_final.copy()
    
    # Appliquer le formatage pour l'aperçu
    df_apercu_final = formater_colonnes_retours_ligne(df_apercu_final)
    
    # Colonnes à afficher (format d'export final)
    colonnes_apercu = ["Code voyage", "Zone", "Véhicule N°", "Chauffeur", "BL inclus", "Client(s) inclus", "Poids total chargé", "Volume total chargé"]
    colonnes_apercu = [col for col in colonnes_apercu if col in df_apercu_final.columns]
    
    # Formater l'affichage pour l'aperçu
    if "Poids total chargé" in df_apercu_final.columns:
        df_apercu_final["Poids total chargé"] = df_apercu_final["Poids total chargé"].map(lambda x: f"{x:.1f} kg")
    if "Volume total chargé" in df_apercu_final.columns:
        df_apercu_final["Volume total chargé"] = df_apercu_final["Volume total chargé"].map(lambda x: f"{x:.3f} m³")
    
    st.dataframe(df_apercu_final[colonnes_apercu], use_container_width=True)

else:
    st.warning("⚠️ Vous devez d'abord valider les voyages et générer les codes voyage.")