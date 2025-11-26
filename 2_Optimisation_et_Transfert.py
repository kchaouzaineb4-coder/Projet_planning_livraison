import streamlit as st
import pandas as pd
from backend import TruckRentalProcessor, TruckTransferManager, SEUIL_POIDS, SEUIL_VOLUME
import plotly.express as px
from io import BytesIO
import openpyxl
from openpyxl.styles import Alignment

# =====================================================
# CONFIGURATION DE LA PAGE
# =====================================================
st.set_page_config(page_title="Optimisation & Transfert", layout="wide")

# =====================================================
# FONCTIONS UTILITAIRES
# =====================================================
def show_df(df, **kwargs):
    """Affiche un DataFrame avec arrondi à 3 décimales"""
    if isinstance(df, pd.DataFrame):
        df_to_display = df.copy()
        df_to_display = df_to_display.round(3)
        st.dataframe(df_to_display, **kwargs)
    else:
        st.dataframe(df, **kwargs)

# =====================================================
# CSS PERSONNALISÉ
# =====================================================
st.markdown("""
<style>
/* Style pour les tableaux de location */
.custom-table-rental {
    width: 100%;
    border-collapse: collapse;
    font-family: Arial, sans-serif;
    font-size: 14px;
    box-shadow: 0 2px 8px rgba(0,0,0,0.1);
    border-radius: 8px;
    overflow: hidden;
}

.custom-table-rental th {
    background-color: #0369A1;
    color: white;
    padding: 12px 8px;
    text-align: center;
    border: 2px solid #4682B4;
    font-weight: normal;
    font-size: 13px;
    vertical-align: middle;
}

.custom-table-rental td {
    padding: 10px 8px;
    text-align: center;
    border: 1px solid #B0C4DE;
    background-color: white;
    color: #000000;
    vertical-align: middle;
    font-weight: normal;
}

.custom-table-rental th, 
.custom-table-rental td {
    border: 1px solid #B0C4DE !important;
}

.custom-table-rental {
    border: 2px solid #4682B4 !important;
}

.custom-table-rental td:nth-child(2),
.custom-table-rental td:nth-child(3),
.custom-table-rental td:nth-child(4),
.custom-table-rental td:nth-child(5),
.custom-table-rental td:nth-child(6) {
    font-weight: normal;
    color: #000000 !important;
    vertical-align: middle;
}

.table-container-rental {
    overflow-x: auto;
    margin: 1rem 0;
    border-radius: 8px;
    border: 2px solid #4682B4;
}

.custom-table-rental tr:nth-child(even) td {
    background-color: white !important;
}

.custom-table-rental tr:hover td {
    background-color: #F0F8FF !important;
}

.multiline-cell {
    line-height: 1.4;
    text-align: left !important;
    padding: 8px !important;
    font-weight: normal;
}

/* Style pour les tableaux de voyages */
.custom-table-voyages {
    width: 100%;
    border-collapse: collapse;
    font-family: Arial, sans-serif;
    font-size: 14px;
    box-shadow: 0 2px 8px rgba(0,0,0,0.1);
    border-radius: 8px;
    overflow: hidden;
}

.custom-table-voyages th {
    background-color: #0369A1;
    color: white;
    padding: 12px 8px;
    text-align: center;
    border: 2px solid #4682B4;
    font-weight: normal;
    font-size: 13px;
    vertical-align: middle;
}

.custom-table-voyages td {
    padding: 10px 8px;
    text-align: center;
    border: 1px solid #B0C4DE;
    background-color: white;
    color: #000000;
    vertical-align: middle;
    font-weight: normal;
}

.custom-table-voyages th, 
.custom-table-voyages td {
    border: 1px solid #B0C4DE !important;
}

.custom-table-voyages {
    border: 2px solid #4682B4 !important;
}

.table-container-voyages {
    overflow-x: auto;
    margin: 1rem 0;
    border-radius: 8px;
    border: 2px solid #4682B4;
}

.custom-table-voyages tr:nth-child(even) td {
    background-color: white !important;
}

.custom-table-voyages tr:hover td {
    background-color: #F0F8FF !important;
}

.custom-table-voyages td {
    line-height: 1.4;
}

/* Style pour le transfert */
.centered-table {
    margin-left: auto;
    margin-right: auto;
    display: table;
    width: 100%;
}

.centered-table table {
    margin: 0 auto;
    border-collapse: collapse;
    width: 100%;
    font-family: Arial, sans-serif;
    box-shadow: 0 2px 8px rgba(0,0,0,0.1);
}

.centered-table th {
    background-color: #0369A1;
    color: white;
    padding: 12px 8px;
    text-align: center;
    border: 2px solid #555;
    font-weight: normal;
    font-size: 14px;
    vertical-align: middle;
}

.centered-table td {
    padding: 10px 8px;
    text-align: center;
    border: 2px solid #555;
    background-color: #f9f9f9;
    color: #333;
    vertical-align: middle;
}

.centered-table tr:nth-child(even) td {
    background-color: #f0f0f0;
}

.centered-table tr:hover td {
    background-color: #e6f3ff;
}

/* Style pour le multiselect */
.stMultiSelect > div > div {
    background-color: #F8FAFC !important;
    border: 2px solid #CBD5E1 !important;
    border-radius: 8px !important;
}

div[data-baseweb="select"] > div {
    background-color: #F8FAFC !important;
    border: 2px solid #CBD5E1 !important;
    border-radius: 8px !important;
}

div[data-baseweb="select"] > div:first-child {
    background-color: #F8FAFC !important;
    border: 2px solid #CBD5E1 !important;
    border-radius: 8px !important;
}

div[data-baseweb="select"] span[data-baseweb="tag"] {
    background-color: #0369A1 !important;
    color: white !important;
    border-radius: 12px !important;
    font-weight: bold;
}

div[role="listbox"] {
    background-color: white !important;
    border: 2px solid #CBD5E1 !important;
}

div[role="option"][aria-selected="true"] {
    background-color: #F1F5F9 !important;
    color: #475569 !important;
}

div[role="option"]:hover {
    background-color: #E2E8F0 !important;
}

/* Style pour l'ajout d'objets */
.custom-border {
    border: 2px solid #1f77b4;
    border-radius: 5px;
    padding: 10px;
    margin: 5px 0px;
}

.custom-button {
    background-color: #1f77b4 !important;
    color: white !important;
    border: none !important;
}

.custom-button:hover {
    background-color: #1668a5 !important;
    color: white !important;
}

div.stButton > button:first-child {
    background-color: #1f77b4;
    color: white;
    border: none;
}

div.stButton > button:first-child:hover {
    background-color: #1668a5;
    color: white;
}
</style>
""", unsafe_allow_html=True)

# =====================================================
# VÉRIFICATION DES DONNÉES
# =====================================================
st.title("🚚 Optimisation des Tournées & Transfert")

if not st.session_state.get('data_processed', False):
    st.warning("⚠️ Veuillez d'abord importer et traiter les données dans la page 'Import & Analyse'")
    st.stop()

# =====================================================
# FONCTIONS DE CALLBACK POUR LA LOCATION
# =====================================================
def update_propositions_view():
    """Met à jour le DataFrame de propositions après une action."""
    if st.session_state.rental_processor:
        st.session_state.propositions = st.session_state.rental_processor.detecter_propositions()
        
        if (st.session_state.propositions is not None and 
            not st.session_state.propositions.empty and 
            'Client' in st.session_state.propositions.columns):
            
            if (st.session_state.selected_client is not None and 
                st.session_state.selected_client not in st.session_state.propositions['Client'].astype(str).tolist()):
                st.session_state.selected_client = None
    else:
        st.session_state.propositions = pd.DataFrame()

def handle_location_action(accepter):
    """Gère l'acceptation ou le refus de la proposition de location."""
    if st.session_state.rental_processor and st.session_state.selected_client:
        try:
            client_to_process = str(st.session_state.selected_client)
            ok, msg, _ = st.session_state.rental_processor.appliquer_location(
                client_to_process, accepter=accepter
            )
            st.session_state.message = msg
            update_propositions_view()
        except Exception as e:
            st.session_state.message = f"❌ Erreur lors du traitement : {str(e)}"
    elif not st.session_state.selected_client:
        st.session_state.message = "⚠️ Veuillez sélectionner un client à traiter."
    else:
        st.session_state.message = "⚠️ Le processeur de location n'est pas initialisé."

def accept_location_callback():
    handle_location_action(True)

def refuse_location_callback():
    handle_location_action(False)

# =====================================================
# INITIALISATION DU PROCESSOR DE LOCATION
# =====================================================
if 'rental_processor' not in st.session_state:
    st.session_state.rental_processor = TruckRentalProcessor(
        st.session_state.df_optimized_estafettes, 
        st.session_state.df_livraisons_original
    )
    update_propositions_view()

if 'selected_client' not in st.session_state:
    st.session_state.selected_client = None

if 'message' not in st.session_state:
    st.session_state.message = ""

# =====================================================
# 3. PROPOSITION DE LOCATION DE CAMION (Section 3)
# =====================================================
st.header("3. 🚚 Proposition de location de camion")
st.markdown(f"🔸 Si un client dépasse **{SEUIL_POIDS} kg** ou **{SEUIL_VOLUME} m³**, une location est proposée (si non déjà décidée).")

if st.session_state.propositions is not None and not st.session_state.propositions.empty:
    col_prop, col_details = st.columns([2, 3])
    
    with col_prop:
        st.markdown("### Propositions ouvertes")
        
        if 'Client' in st.session_state.propositions.columns:
            propositions_display = st.session_state.propositions.copy()
            
            if "Poids total (kg)" in propositions_display.columns:
                propositions_display["Poids total (kg)"] = propositions_display["Poids total (kg)"].map(
                    lambda x: f"{float(x):.3f}" if pd.notna(x) else ""
                )
            if "Volume total (m³)" in propositions_display.columns:
                propositions_display["Volume total (m³)"] = propositions_display["Volume total (m³)"].map(
                    lambda x: f"{float(x):.3f}" if pd.notna(x) else ""
                )
            
            html_table_propositions = propositions_display.to_html(
                escape=False, 
                index=False, 
                classes="custom-table-rental",
                border=0
            )
            
            st.markdown(f"""
            <div class="table-container-rental">
                {html_table_propositions}
            </div>
            """, unsafe_allow_html=True)
            
            # MÉTRIQUES RÉSUMÉES
            st.markdown("---")
            col_metric1, col_metric2, col_metric3 = st.columns(3)

            with col_metric1:
                total_propositions = len(st.session_state.propositions)
                st.metric("📋 Propositions ouvertes", total_propositions)

            with col_metric2:
                clients_poids = len(st.session_state.propositions[
                    st.session_state.propositions["Poids total (kg)"] >= SEUIL_POIDS
                ]) if "Poids total (kg)" in st.session_state.propositions.columns else 0
                st.metric("⚖️ Dépassement poids", clients_poids)

            with col_metric3:
                clients_volume = len(st.session_state.propositions[
                    st.session_state.propositions["Volume total (m³)"] >= SEUIL_VOLUME
                ]) if "Volume total (m³)" in st.session_state.propositions.columns else 0
                st.metric("📦 Dépassement volume", clients_volume)

            # Sélection du client
            client_options = st.session_state.propositions['Client'].astype(str).tolist()
            client_options_with_empty = [""] + client_options
            
            default_index = 0
            if st.session_state.selected_client in client_options:
                 default_index = client_options_with_empty.index(st.session_state.selected_client)
            elif len(client_options) > 0:
                 default_index = 1

            st.session_state.selected_client = st.selectbox(
                "Client à traiter :", 
                options=client_options_with_empty, 
                index=default_index,
                key='client_select' 
            )
        else:
            st.warning("⚠️ Format de données incorrect dans les propositions.")
            st.session_state.selected_client = None

        col_btn_acc, col_btn_ref = st.columns(2)
        is_client_selected = st.session_state.selected_client != "" and st.session_state.selected_client is not None
        
        with col_btn_acc:
            st.button(
                "✅ Accepter la location", 
                on_click=accept_location_callback, 
                disabled=not is_client_selected,
                use_container_width=True
            )
        with col_btn_ref:
            st.button(
                "❌ Refuser la proposition", 
                on_click=refuse_location_callback, 
                disabled=not is_client_selected,
                use_container_width=True
            )

    with col_details:
        st.markdown("### Détails de la commande client")
        if is_client_selected:
            try:
                resume, details_df = st.session_state.rental_processor.get_details_client(
                    st.session_state.selected_client
                )
                
                st.markdown(f"**{resume}**")
                
                if not details_df.empty:
                    details_display = details_df.copy()
                    
                    def format_numeric_column(series, decimals, unit=""):
                        formatted_series = series.copy()
                        for i, value in enumerate(series):
                            if pd.notna(value) and value != "":
                                try:
                                    if isinstance(value, str):
                                        clean_value = value.replace(' kg', '').replace(' m³', '').replace('%', '').strip()
                                        num_value = float(clean_value)
                                    else:
                                        num_value = float(value)
                                    
                                    if decimals == 3:
                                        formatted_value = f"{num_value:.3f}"
                                    elif decimals == 2:
                                        formatted_value = f"{num_value:.2f}"
                                    elif decimals == 1:
                                        formatted_value = f"{num_value:.1f}"
                                    else:
                                        formatted_value = f"{num_value:.0f}"
                                    
                                    formatted_series.iloc[i] = f"{formatted_value}{unit}"
                                except (ValueError, TypeError):
                                    formatted_series.iloc[i] = str(value)
                            else:
                                formatted_series.iloc[i] = ""
                        return formatted_series
                    
                    if "Poids total" in details_display.columns:
                        details_display["Poids total"] = format_numeric_column(details_display["Poids total"], 3, " kg")
                    
                    if "Volume total" in details_display.columns:
                        details_display["Volume total"] = format_numeric_column(details_display["Volume total"], 3, " m³")
                    
                    if "Taux d'occupation (%)" in details_display.columns:
                        details_display["Taux d'occupation (%)"] = format_numeric_column(details_display["Taux d'occupation (%)"], 2, "%")
                    
                    if "BL inclus" in details_display.columns:
                        details_display["BL inclus"] = details_display["BL inclus"].astype(str).apply(
                            lambda x: "<br>".join(bl.strip() for bl in x.split(";")) if ";" in x else x
                        )
                    
                    html_table_details = details_display.to_html(
                        escape=False, 
                        index=False, 
                        classes="custom-table-rental",
                        border=0
                    )
                    
                    st.markdown(f"""
                    <div class="table-container-rental">
                        {html_table_details}
                    </div>
                    """, unsafe_allow_html=True)
                    
                    # MÉTRIQUES POUR LES DÉTAILS
                    st.markdown("---")
                    col_det1, col_det2, col_det3 = st.columns(3)
                    
                    with col_det1:
                        total_camions = len(details_display)
                        st.metric("🚚 Nombre de camions", total_camions)
                    
                    with col_det2:
                        try:
                            if "Poids total" in details_df.columns:
                                poids_total = 0
                                for value in details_df["Poids total"]:
                                    if pd.notna(value):
                                        try:
                                            if isinstance(value, str):
                                                clean_value = value.replace(' kg', '').replace('m³', '').strip()
                                            else:
                                                clean_value = str(value)
                                            poids_total += float(clean_value)
                                        except (ValueError, TypeError):
                                            continue
                                st.metric("📦 Poids total", f"{poids_total:.1f} kg")
                            else:
                                st.metric("📦 Poids total", "N/A")
                        except Exception as e:
                            st.metric("📦 Poids total", "Erreur")
                    
                    with col_det3:
                        try:
                            if "Volume total" in details_df.columns:
                                volume_total = 0
                                for value in details_df["Volume total"]:
                                    if pd.notna(value):
                                        try:
                                            if isinstance(value, str):
                                                clean_value = value.replace(' kg', '').replace('m³', '').strip()
                                            else:
                                                clean_value = str(value)
                                            volume_total += float(clean_value)
                                        except (ValueError, TypeError):
                                            continue
                                st.metric("📏 Volume total", f"{volume_total:.3f} m³")
                            else:
                                st.metric("📏 Volume total", "N/A")
                        except Exception as e:
                            st.metric("📏 Volume total", "Erreur")
                        
            except Exception as e:
                st.error(f"❌ Erreur lors de la récupération des détails : {str(e)}")
        else:
            st.info("Sélectionnez un client pour afficher les détails de la commande/estafettes.")
else:
    st.success("✅ Aucune proposition de location de camion en attente de décision.")

st.markdown("---")

# =====================================================
# 4. VOYAGES PAR ESTAFETTE OPTIMISÉ (Section 4)
# =====================================================
st.header("4. 🚐 Voyages par Estafette Optimisé (Inclut Camions Loués)")

try:
    if st.session_state.rental_processor:
        df_optimized_estafettes = st.session_state.rental_processor.get_df_result()
    elif "df_voyages" in st.session_state:
        df_optimized_estafettes = st.session_state.df_voyages.copy()
    else:
        st.error("❌ Données non disponibles. Veuillez exécuter le traitement complet.")
        st.stop()
    
    if df_optimized_estafettes.empty:
        st.warning("⚠️ Aucune donnée à afficher.")
        st.stop()
    
    df_clean = df_optimized_estafettes.loc[:, ~df_optimized_estafettes.columns.duplicated()]
    
    if "Zone" in df_clean.columns:
        df_clean["Zone_Num"] = df_clean["Zone"].str.extract('(\d+)').astype(float)
        df_clean = df_clean.sort_values("Zone_Num").drop("Zone_Num", axis=1)
    
    colonnes_ordre = [
        "Zone", "Véhicule N°", "Poids total chargé", "Volume total chargé",
        "Client(s) inclus", "Représentant(s) inclus", "BL inclus", 
        "Taux d'occupation (%)", "Location_camion", "Location_proposee", "Code Véhicule"
    ]
    
    colonnes_finales = [col for col in colonnes_ordre if col in df_clean.columns]
    
    df_display = df_clean[colonnes_finales].copy()
    
    if "Client(s) inclus" in df_display.columns:
        df_display["Client(s) inclus"] = df_display["Client(s) inclus"].astype(str).apply(
            lambda x: "<br>".join(client.strip() for client in x.split(",")) if x != "nan" else ""
        )
    
    if "Représentant(s) inclus" in df_display.columns:
        df_display["Représentant(s) inclus"] = df_display["Représentant(s) inclus"].astype(str).apply(
            lambda x: "<br>".join(rep.strip() for rep in x.split(",")) if x != "nan" else ""
        )
    
    if "BL inclus" in df_display.columns:
        df_display["BL inclus"] = df_display["BL inclus"].astype(str).apply(
            lambda x: "<br>".join(bl.strip() for bl in x.split(";")) if x != "nan" else ""
        )
    
    if "Poids total chargé" in df_display.columns:
        df_display["Poids total chargé"] = df_display["Poids total chargé"].map(lambda x: f"{x:.3f} kg")
    if "Volume total chargé" in df_display.columns:
        df_display["Volume total chargé"] = df_display["Volume total chargé"].map(lambda x: f"{x:.3f} m³")
    if "Taux d'occupation (%)" in df_display.columns:
        df_display["Taux d'occupation (%)"] = df_display["Taux d'occupation (%)"].map(lambda x: f"{x:.3f}%")
    
    html_table = df_display.to_html(escape=False, index=False, classes="custom-table-voyages", border=0)
    
    st.markdown(f"""
    <div class="table-container-voyages">
        {html_table}
    </div>
    """, unsafe_allow_html=True)
    
    # MÉTRIQUES RÉSUMÉES
    st.markdown("---")
    col1, col2, col3, col4 = st.columns(4)
    
    with col1:
        total_voyages = len(df_display)
        st.metric("🚐 Total Voyages", total_voyages)
    
    with col2:
        total_zones = df_display["Zone"].nunique() if "Zone" in df_display.columns else 0
        st.metric("🌍 Zones couvertes", total_zones)
    
    with col3:
        camions_loues = df_display["Location_camion"].sum() if "Location_camion" in df_display.columns else 0
        st.metric("🚚 Camions loués", int(camions_loues))
    
    with col4:
        estafettes = total_voyages - camions_loues
        st.metric("📦 Estafettes", estafettes)
    
    # Préparer l'export Excel
    df_export = df_clean.copy()
    
    if "Zone" in df_export.columns:
        df_export["Zone_Num"] = df_export["Zone"].str.extract('(\d+)').astype(float)
        df_export = df_export.sort_values("Zone_Num").drop("Zone_Num", axis=1)
    
    if "Client(s) inclus" in df_export.columns:
        df_export["Client(s) inclus"] = df_export["Client(s) inclus"].astype(str).apply(
            lambda x: "\n".join(client.strip() for client in x.split(",")) if x != "nan" else ""
        )
    
    if "Représentant(s) inclus" in df_export.columns:
        df_export["Représentant(s) inclus"] = df_export["Représentant(s) inclus"].astype(str).apply(
            lambda x: "\n".join(rep.strip() for rep in x.split(",")) if x != "nan" else ""
        )
    
    if "BL inclus" in df_export.columns:
        df_export["BL inclus"] = df_export["BL inclus"].astype(str).apply(
            lambda x: "\n".join(bl.strip() for bl in x.split(";")) if x != "nan" else ""
        )
    
    if "Poids total chargé" in df_export.columns:
        df_export["Poids total chargé"] = df_export["Poids total chargé"].round(3)
    if "Volume total chargé" in df_export.columns:
        df_export["Volume total chargé"] = df_export["Volume total chargé"].round(3)
    
    excel_buffer = BytesIO()
    
    with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
        df_export.to_excel(writer, index=False, sheet_name="Voyages Optimisés")
        
        workbook = writer.book
        worksheet = writer.sheets["Voyages Optimisés"]
        
        wrap_columns = []
        if "Client(s) inclus" in df_export.columns:
            wrap_columns.append("Client(s) inclus")
        if "Représentant(s) inclus" in df_export.columns:
            wrap_columns.append("Représentant(s) inclus")
        if "BL inclus" in df_export.columns:
            wrap_columns.append("BL inclus")
        
        for col_idx, col_name in enumerate(df_export.columns):
            if col_name in wrap_columns:
                col_letter = openpyxl.utils.get_column_letter(col_idx + 1)
                for row in range(2, len(df_export) + 2):
                    cell = worksheet[f"{col_letter}{row}"]
                    cell.alignment = Alignment(wrap_text=True, vertical='top')
        
        for column in worksheet.columns:
            max_length = 0
            column_letter = openpyxl.utils.get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
    
    excel_buffer.seek(0)
    
    st.download_button(
        label="💾 Télécharger Voyages Estafette Optimisés",
        data=excel_buffer,
        file_name="Voyages_Estafette_Optimises.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    
    # Mise à jour pour les sections suivantes
    st.session_state.df_voyages = df_clean

except KeyError as e:
    st.error(f"❌ Erreur de colonne manquante : {e}")
    if st.session_state.rental_processor:
        st.session_state.df_voyages = st.session_state.rental_processor.df_base.copy()
        st.rerun()
        
except Exception as e:
    st.error(f"❌ Erreur lors de l'affichage des voyages optimisés: {str(e)}")

st.markdown("---")

# =====================================================
# TRANSFERT DE BLs ENTRE ESTAFETTES / CAMIONS
# =====================================================
st.markdown("## 🔁 Transfert de BLs entre Estafettes / Camions")

MAX_POIDS = 1550  # kg
MAX_VOLUME = 4.608  # m³

if "df_voyages" not in st.session_state:
    st.warning("⚠️ Vous devez d'abord exécuter la section 4 (Voyages par Estafette Optimisé).")
elif "df_livraisons" not in st.session_state:
    st.warning("⚠️ Le DataFrame des livraisons détaillées n'est pas disponible.")
else:
    df_voyages = st.session_state.df_voyages.copy()
    df_livraisons = st.session_state.df_livraisons.copy()

    colonnes_requises = ["Zone", "Véhicule N°", "Poids total chargé", "Volume total chargé", "BL inclus"]

    if not all(col in df_voyages.columns for col in colonnes_requises):
        st.error(f"❌ Le DataFrame ne contient pas toutes les colonnes nécessaires : {', '.join(colonnes_requises)}")
    else:
        zones_disponibles = sorted(df_voyages["Zone"].dropna().unique().tolist())
        zone_selectionnee = st.selectbox("🌍 Sélectionner une zone", zones_disponibles)

        if zone_selectionnee:
            df_zone = df_voyages[df_voyages["Zone"] == zone_selectionnee]
            vehicules = sorted(df_zone["Véhicule N°"].dropna().unique().tolist())

            col1, col2 = st.columns(2)
            with col1:
                source = st.selectbox("🚐 Estafette / Camion source", vehicules)
            with col2:
                cible = st.selectbox("🎯 Estafette / Camion cible", [v for v in vehicules if v != source])

            if source and cible:
                df_source = df_zone[df_zone["Véhicule N°"] == source]
                if df_source.empty or df_source["BL inclus"].isna().all():
                    st.warning("⚠️ Aucun BL trouvé pour ce véhicule source.")
                else:
                    st.subheader(f"📦 BLs actuellement assignés à {source}")

                    bls_avec_clients = []
                    bls_simples = df_source["BL inclus"].iloc[0].split(";")
                    
                    for bl in bls_simples:
                        client_info = df_livraisons[df_livraisons["No livraison"] == bl]
                        if not client_info.empty:
                            client_nom = client_info["Client de l'estafette"].iloc[0]
                            bl_affichage = f"{bl} - {client_nom}"
                        else:
                            bl_affichage = f"{bl} - Client non trouvé"
                        bls_avec_clients.append(bl_affichage)
                    
                    df_source_display = df_source[["Véhicule N°", "Poids total chargé", "Volume total chargé"]].copy()
                    df_source_display["BL inclus (avec clients)"] = "<br>".join(bls_avec_clients)
                    
                    df_source_display["Poids total chargé"] = df_source_display["Poids total chargé"].map(lambda x: f"{x:.3f} kg")
                    df_source_display["Volume total chargé"] = df_source_display["Volume total chargé"].map(lambda x: f"{x:.3f} m³")
                    
                    html_content = f"""
                    <div class="centered-table">
                    {df_source_display.to_html(escape=False, index=False)}
                    </div>
                    """
                    st.markdown(html_content, unsafe_allow_html=True)

                    # Sélection avec clients
                    st.subheader("📋 Sélectionner les BLs à transférer")
                    
                    options_transfert = []
                    mapping_bl_original = {}
                    
                    for bl in bls_simples:
                        client_info = df_livraisons[df_livraisons["No livraison"] == bl]
                        if not client_info.empty:
                            client_nom = client_info["Client de l'estafette"].iloc[0]
                            option_affichage = f"{bl} - {client_nom}"
                        else:
                            option_affichage = f"{bl} - Client non trouvé"
                        
                        options_transfert.append(option_affichage)
                        mapping_bl_original[option_affichage] = bl
                    
                    bls_selectionnes_affichage = st.multiselect(
                        "Sélectionnez les BLs à transférer (avec clients) :", 
                        options_transfert,
                        format_func=lambda x: x
                    )
                    
                    bls_selectionnes = [mapping_bl_original[bl_affichage] for bl_affichage in bls_selectionnes_affichage]

                    if bls_selectionnes and st.button("🔁 Exécuter le transfert"):
                        df_bls_selection = df_livraisons[df_livraisons["No livraison"].isin(bls_selectionnes)]
                        poids_bls = df_bls_selection["Poids total"].sum()
                        volume_bls = df_bls_selection["Volume total"].sum()

                        df_cible = df_zone[df_zone["Véhicule N°"] == cible]
                        poids_cible = df_cible["Poids total chargé"].sum()
                        volume_cible = df_cible["Volume total chargé"].sum()

                        if (poids_cible + poids_bls) > MAX_POIDS or (volume_cible + volume_bls) > MAX_VOLUME:
                            st.warning("⚠️ Le transfert dépasse les limites de poids ou volume du véhicule cible.")
                        else:
                            def transfer_bl(row):
                                bls = row["BL inclus"].split(";") if pd.notna(row["BL inclus"]) else []
                                bls_to_move = [b for b in bls if b in bls_selectionnes]

                                if row["Véhicule N°"] == source:
                                    new_bls = [b for b in bls if b not in bls_to_move]
                                    row["BL inclus"] = ";".join(new_bls)
                                    row["Poids total chargé"] = max(0, row["Poids total chargé"] - poids_bls)
                                    row["Volume total chargé"] = max(0, row["Volume total chargé"] - volume_bls)
                                elif row["Véhicule N°"] == cible:
                                    new_bls = bls + bls_to_move
                                    row["BL inclus"] = ";".join(new_bls)
                                    row["Poids total chargé"] += poids_bls
                                    row["Volume total chargé"] += volume_bls
                                return row

                            df_voyages = df_voyages.apply(transfer_bl, axis=1)
                            st.session_state.df_voyages = df_voyages
                            
                            clients_transferes = df_bls_selection["Client de l'estafette"].unique()
                            st.success(f"""
                            ✅ Transfert réussi !
                            - **{len(bls_selectionnes)} BL(s)** déplacé(s) de **{source}** vers **{cible}**
                            - **Clients concernés :** {', '.join(clients_transferes)}
                            - **Poids transféré :** {poids_bls:.1f} kg
                            - **Volume transféré :** {volume_bls:.3f} m³
                            """)

                            # Affichage après transfert
                            st.subheader("📊 Voyages après transfert (toutes les zones)")
                            df_display = df_voyages.sort_values(by=["Zone", "Véhicule N°"]).copy()
                            
                            if "BL inclus" in df_display.columns:
                                df_display["BL inclus"] = df_display["BL inclus"].astype(str).apply(
                                    lambda x: "<br>".join(bl.strip() for bl in x.split(";")) if x != "nan" else ""
                                )
                            
                            df_display["Poids total chargé"] = df_display["Poids total chargé"].map(lambda x: f"{x:.3f} kg")
                            df_display["Volume total chargé"] = df_display["Volume total chargé"].map(lambda x: f"{x:.3f} m³")
                            
                            html_content_after = f"""
                            <div class="centered-table">
                            {df_display[colonnes_requises].to_html(escape=False, index=False)}
                            </div>
                            """
                            st.markdown(html_content_after, unsafe_allow_html=True)

                            # Export Excel
                            df_export = df_voyages.copy()
                            
                            if "BL inclus" in df_export.columns:
                                df_export["BL inclus"] = df_export["BL inclus"].astype(str).apply(
                                    lambda x: "\n".join(bl.strip() for bl in x.split(";")) if x != "nan" else ""
                                )
                            
                            df_export["Poids total chargé"] = df_export["Poids total chargé"].round(3)
                            df_export["Volume total chargé"] = df_export["Volume total chargé"].round(3)

                            excel_buffer = BytesIO()
                            with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
                                df_export.to_excel(writer, index=False, sheet_name='Transfert BLs')
                                
                                workbook = writer.book
                                worksheet = writer.sheets['Transfert BLs']
                                
                                if "BL inclus" in df_export.columns:
                                    for col_idx, col_name in enumerate(df_export.columns):
                                        if col_name == "BL inclus":
                                            col_letter = openpyxl.utils.get_column_letter(col_idx + 1)
                                            for row in range(2, len(df_export) + 2):
                                                cell = worksheet[f"{col_letter}{row}"]
                                                cell.alignment = Alignment(wrap_text=True, vertical='top')
                            
                            excel_buffer.seek(0)

                            st.download_button(
                                label="💾 Télécharger le tableau mis à jour (XLSX)",
                                data=excel_buffer,
                                file_name="voyages_apres_transfert.xlsx",
                                mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                            )

st.markdown("---")

# =====================================================
# AJOUT D'OBJETS MANUELS AUX VÉHICULES
# =====================================================
st.markdown("## 📦 AJOUT D'OBJETS MANUELS AUX VÉHICULES")

if "df_voyages" in st.session_state:
    if "transfer_manager" not in st.session_state:
        st.session_state.transfer_manager = TruckTransferManager(
            st.session_state.df_voyages, 
            st.session_state.df_livraisons
        )
    
    df_voyages = st.session_state.df_voyages.copy()
    
    col1, col2, col3 = st.columns(3)
    
    with col1:
        zones_disponibles = sorted(df_voyages["Zone"].dropna().unique().tolist())
        zone_objet = st.selectbox("🌍 Zone", zones_disponibles, key="zone_objet")
    
    with col2:
        if zone_objet:
            vehicules_zone = sorted(
                df_voyages[df_voyages["Zone"] == zone_objet]["Véhicule N°"].dropna().unique().tolist()
            )
            vehicule_objet = st.selectbox("🚚 Véhicule", vehicules_zone, key="vehicule_objet")
        else:
            vehicule_objet = st.selectbox("🚚 Véhicule", [], key="vehicule_objet")
    
    with col3:
        if zone_objet and vehicule_objet:
            vehicule_data = df_voyages[
                (df_voyages["Zone"] == zone_objet) & 
                (df_voyages["Véhicule N°"] == vehicule_objet)
            ].iloc[0]
            
            is_camion = vehicule_data.get("Code Véhicule", "") == "CAMION-LOUE"
            capacite_poids = 30500 if is_camion else 1550
            capacite_volume = 77.5 if is_camion else 4.608
            
            poids_actuel = vehicule_data.get("Poids total chargé", 0)
            volume_actuel = vehicule_data.get("Volume total chargé", 0)
            
            st.metric(
                "📊 Capacité utilisée", 
                f"{poids_actuel:.1f}kg / {capacite_poids}kg",
                f"{volume_actuel:.3f}m³ / {capacite_volume}m³"
            )
    
    st.markdown("### 📝 Détails de l'objet à ajouter")
    
    col4, col5, col6 = st.columns(3)
    
    with col4:
        st.markdown('<div class="custom-border">', unsafe_allow_html=True)
        nom_objet = st.text_input("🏷️ Nom de l'objet", placeholder="Ex: Matériel urgent, Colis oublié...")
        st.markdown('</div>', unsafe_allow_html=True)
    
    with col5:
        st.markdown('<div class="custom-border">', unsafe_allow_html=True)
        poids_objet = st.number_input("⚖️ Poids (kg)", min_value=0.0, max_value=1000.0, value=10.0, step=0.1)
        st.markdown('</div>', unsafe_allow_html=True)
    
    with col6:
        st.markdown('<div class="custom-border">', unsafe_allow_html=True)
        volume_objet = st.number_input("📦 Volume (m³)", min_value=0.0, max_value=10.0, value=0.1, step=0.01)
        st.markdown('</div>', unsafe_allow_html=True)
    
    if st.button("➕ Ajouter l'objet au véhicule", type="primary"):
        if not nom_objet:
            st.error("❌ Veuillez donner un nom à l'objet.")
        elif zone_objet and vehicule_objet:
            try:
                success, message, df_updated = st.session_state.transfer_manager.add_manual_object(
                    df_voyages=df_voyages,
                    vehicle=vehicule_objet,
                    zone=zone_objet,
                    name=nom_objet,
                    weight=poids_objet,
                    volume=volume_objet
                )
                
                if success:
                    st.success(message)
                    
                    st.session_state.df_voyages = df_updated
                    st.session_state.transfer_manager.df_voyages = df_updated.copy()
                    
                    if st.session_state.rental_processor:
                        try:
                            st.session_state.rental_processor.df_base = df_updated.copy()
                            st.session_state.rental_processor = TruckRentalProcessor(
                                df_updated, 
                                st.session_state.df_livraisons_original
                            )
                            st.success("✅ Processeur de location synchronisé")
                        except Exception as e:
                            st.warning(f"⚠️ Synchronisation partielle du processeur : {str(e)}")
                    
                    if st.session_state.propositions is not None:
                        try:
                            st.session_state.propositions = st.session_state.rental_processor.detecter_propositions()
                        except:
                            pass
                    
                    if 'df_voyages_valides' in st.session_state:
                        try:
                            mask_valides = df_updated["Véhicule N°"].isin(
                                st.session_state.df_voyages_valides["Véhicule N°"]
                            )
                            st.session_state.df_voyages_valides = df_updated[mask_valides].copy()
                        except:
                            pass
                    
                    vehicule_update = df_updated[
                        (df_updated["Zone"] == zone_objet) & 
                        (df_updated["Véhicule N°"] == vehicule_objet)
                    ].iloc[0]
                    
                    st.info(f"""
                    **Véhicule mis à jour :**
                    - Poids total : {vehicule_update['Poids total chargé']:.1f} kg
                    - Volume total : {vehicule_update['Volume total chargé']:.3f} m³
                    - Taux d'occupation : {vehicule_update['Taux d\'occupation (%)']:.1f}%
                    - BLs inclus : {vehicule_update['BL inclus']}
                    """)
                    
                    st.success("🔄 Toutes les données ont été mises à jour avec succès !")
                    st.rerun()
                    
                else:
                    st.error(message)
                    
            except Exception as e:
                st.error(f"❌ Erreur lors de l'ajout de l'objet : {str(e)}")
        else:
            st.error("❌ Veuillez sélectionner une zone et un véhicule.")
    
    st.markdown("### 📋 Historique des objets ajoutés")
    
    objets_manuels = []
    for idx, row in df_voyages.iterrows():
        bls = str(row.get("BL inclus", ""))
        if "OBJ-" in bls:
            for bl in bls.split(";"):
                if bl.startswith("OBJ-"):
                    vehicule_info = df_voyages[
                        (df_voyages["Zone"] == row["Zone"]) & 
                        (df_voyages["Véhicule N°"] == row["Véhicule N°"])
                    ]
                    if not vehicule_info.empty:
                        poids_vehicule = vehicule_info["Poids total chargé"].iloc[0]
                        volume_vehicule = vehicule_info["Volume total chargé"].iloc[0]
                        
                        objets_manuels.append({
                            "Véhicule": row["Véhicule N°"],
                            "Zone": row["Zone"],
                            "Objet": bl,
                            "Poids Véhicule": f"{poids_vehicule:.1f} kg",
                            "Volume Véhicule": f"{volume_vehicule:.3f} m³",
                            "Type": "Camion" if row.get("Code Véhicule", "") == "CAMION-LOUE" else "Estafette"
                        })
    
    if objets_manuels:
        df_objets = pd.DataFrame(objets_manuels)
        show_df(df_objets, use_container_width=True)
        
        col_clear1, col_clear2 = st.columns([3, 1])
        with col_clear2:
            if st.button("🗑️ Supprimer tous les objets", type="secondary"):
                df_sans_objets = st.session_state.df_voyages.copy()
                for idx, row in df_sans_objets.iterrows():
                    bls_originaux = str(row["BL inclus"]).split(";")
                    bls_filtres = [bl for bl in bls_originaux if not bl.startswith("OBJ-")]
                    df_sans_objets.at[idx, "BL inclus"] = ";".join(bls_filtres)
                
                st.session_state.df_voyages = df_sans_objets
                st.session_state.transfer_manager.df_voyages = df_sans_objets.copy()
                if st.session_state.rental_processor:
                    st.session_state.rental_processor.df_base = df_sans_objets.copy()
                
                st.success("✅ Tous les objets manuels ont été supprimés")
                st.rerun()
    else:
        st.info(" Aucun objet manuel ajouté pour le moment.")

# =====================================================
# PIED DE PAGE
# =====================================================
st.markdown("---")
st.markdown(
    """
    <div style='text-align: center; color: #666;'>
        <p>🚚 <strong>Système d'Optimisation des Livraisons</strong> - Développé par Zaineb KCHAOU</p>
        <p>📧 Support : Zaineb.KCHAOU@sopal.com | 📞 Hotline : +216 23 130 088</p>
    </div>
    """,
    unsafe_allow_html=True
)